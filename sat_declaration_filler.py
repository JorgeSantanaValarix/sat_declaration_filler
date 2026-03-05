#!/usr/bin/env python3
"""
SAT Declaration Filler — Fill SAT provisional declaration from Contaayuda Excel workpaper.
Reads Impuestos tab (D4:E29 ISR, D33:E58 IVA), logs in with e.firma from DB, fills form, checks totals, sends.
See PLAN_FORM_FILL_AUTOMATION.md.
"""

from __future__ import annotations

import argparse
from typing import Callable
import json
import logging
import os
import re
import signal
import sys
import tempfile
import threading
import time
from datetime import datetime
from pathlib import Path

# Used by SIGINT handler to logout from SAT when user presses Ctrl+C.
_run_context: dict | None = None

import openpyxl

try:
    import pyodbc
except ImportError:
    pyodbc = None

try:
    import requests
except ImportError:
    requests = None  # type: ignore[assignment]

from playwright.sync_api import Frame, Page, sync_playwright

# --- Constants from plan ---
IMPUESTOS_SHEET = "Impuestos"
ISR_RANGE = (4, 29)   # rows 4-29, cols D,E
IVA_RANGE = (33, 58)  # rows 33-58, cols D,E
TOLERANCE_PESOS = 1
# Declaración Provisional (pstcdypisr): after login, click "Presentar declaración" to open Configuración de la declaración.
SAT_PORTAL_URL = "https://pstcdypisr.clouda.sat.gob.mx/"
RETRY_WAIT_SECONDS = 60  # wait before single retry after error (e.g. SAT HTTP 500; give server time to recover)
# Phase 2 → 3: wait for "Cargando información" to disappear (max seconds), then for pre-fill pop-up CERRAR to appear
PHASE3_LOADING_MAX_WAIT_SEC = 90
PHASE3_POPUP_WAIT_FOR_CERRAR_SEC = 45   # wait for CERRAR button to appear (pop-up can take several seconds)
PHASE3_POPUP_CERRAR_CLICK_MS = 5000    # timeout when clicking CERRAR once it's visible
PHASE3_POPUP_CERRAR_SELECTOR_MS = 4000  # per mapping selector (fail fast)
PHASE3_SECTION_GAP_MS = 100  # 1s between ISR Ingresos sections (was 20–60ms; reduces perceived lag)
SP_GET_EFIRMA = "[GET_AUTOMATICTAXDECLARATION_CUSTOMERDATA]"
# SAT Automatic Declaration API (dynamic Excel/CER/KEY/password). See docs/PLAN_DYNAMIC_API_EXCEL_CER_KEY_PASSWORD.md.
SAT_AUTOMATIC_DECLARATION_API_BASE_URL = "https://app.valarix.com/Services/SatAutomaticDeclarationService.asmx"
LOG = logging.getLogger("sat_declaration_filler")


def _parse_currency(val) -> float:
    """Parse Excel currency (e.g. '$ 1,132,090' or '$ -') to float. Handles formula cached values and datetime (returns 0.0)."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    # Excel may store dates as datetime; don't treat as currency
    if hasattr(val, "year") and hasattr(val, "month"):
        return 0.0
    s = str(val).strip().replace("$", "").replace(",", "").strip()
    if not s or s == "-":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _cell_value(cell) -> str | float | None:
    v = cell.value
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    return str(v).strip()


def _parse_cell_ref(formula: str) -> tuple[int, int] | None:
    """Parse simple formula like =F8 or =+F8 to (row, col). Column A=1, B=2, ..., F=6. Returns None if not a simple cell ref."""
    if not formula or not isinstance(formula, str):
        return None
    s = formula.strip().lstrip("=+").strip().upper()
    m = re.match(r"^([A-Z]+)(\d+)$", s)
    if not m:
        return None
    col_letters, row_str = m.group(1), m.group(2)
    col = 0
    for c in col_letters:
        col = col * 26 + (ord(c) - ord("A") + 1)
    return (int(row_str), col)


def _row_col_to_a1(row: int, col: int) -> str:
    """Convert 1-based (row, col) to Excel A1 notation (e.g. (8, 6) -> 'F8')."""
    s = ""
    c = col
    while c:
        c, r = divmod(c - 1, 26)
        s = chr(ord("A") + r) + s
    return (s or "A") + str(row)


_XLCALC_IMPORT_FAILED: bool | None = None  # None=not tried, True=failed, False=ok
_XLCALC_TIMEOUT_SEC = 3   # skip xlcalculator if loading takes longer (fail fast; openpyxl usually sufficient)


def _xlcalculator_evaluator(workbook_path: str):
    """
    Build xlcalculator model once (can be slow/hang). Runs with timeout; returns (Evaluator, None) or (None, error_msg).
    """
    global _XLCALC_IMPORT_FAILED
    try:
        from xlcalculator import ModelCompiler, Evaluator
    except ImportError:
        if _XLCALC_IMPORT_FAILED is None:
            _XLCALC_IMPORT_FAILED = True
            LOG.info("xlcalculator not installed; formula fallback disabled. Install with: pip install xlcalculator")
        return None, "not installed"
    if _XLCALC_IMPORT_FAILED is None:
        _XLCALC_IMPORT_FAILED = False

    result: list = []  # [evaluator] or [] on error; error_msg in err[0]
    err: list = []

    def _build():
        try:
            path = os.path.abspath(workbook_path)
            compiler = ModelCompiler()
            model = compiler.read_and_parse_archive(path, build_code=True)
            result.append(Evaluator(model))
        except Exception as e:
            err.append(str(e))

    LOG.info("xlcalculator: loading workbook (timeout %ss)...", _XLCALC_TIMEOUT_SEC)
    t = threading.Thread(target=_build, daemon=True)
    t.start()
    t.join(timeout=_XLCALC_TIMEOUT_SEC)
    if t.is_alive():
        LOG.warning("xlcalculator: loading timed out after %ss; skipping formula fallback", _XLCALC_TIMEOUT_SEC)
        return None, "timeout"
    if err:
        return None, err[0]
    if result:
        return result[0], None
    return None, "unknown"


def _evaluate_cell(evaluator, sheet_name: str, row: int, col: int) -> float | None:
    """Evaluate one cell with an existing xlcalculator Evaluator. Returns float or None."""
    try:
        cell_ref = f"{sheet_name}!{_row_col_to_a1(row, col)}"
        result = evaluator.evaluate(cell_ref)
        if result is None:
            LOG.info("xlcalculator %s returned None", cell_ref)
            return None
        return _parse_currency(result)
    except Exception as e:
        LOG.info("xlcalculator evaluate %s!%s: %s", sheet_name, _row_col_to_a1(row, col), e)
        return None


def read_impuestos(workbook_path: str) -> dict:
    """
    Read Impuestos tab. Two layouts supported:
    - Layout 1: label in column D, value in column E (D4:E29 ISR, D33:E58 IVA).
    - Layout 2: label in column E, value in column F (same row ranges).
    Values in E or F can be numeric or formulas. We use data_only=True so formula cells
    return the cached result (what you see in Excel). If the file was not saved in Excel
    and the cache is missing, we resolve simple refs (e.g. =F8, =+E8) by following the reference.
    Returns dict with keys: label_map (label->value), year, month, periodicidad, tipo_declaracion.
    """
    wb = openpyxl.load_workbook(workbook_path, read_only=False, data_only=True)
    if IMPUESTOS_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{IMPUESTOS_SHEET}' not found in workbook. Sheets: {wb.sheetnames}")
    ws = wb[IMPUESTOS_SHEET]
    label_map = {}
    formula_cells: list[tuple[str, int, int]] = []  # (label, row, col) where value was None

    def add_label_value(label: str, raw, row: int | None = None, col: int | None = None) -> None:
        if not label or not str(label).strip():
            return
        label = str(label).strip()
        if raw is not None and not (isinstance(raw, str) and not raw.strip()):
            if isinstance(raw, (int, float)):
                label_map[label] = float(raw)
            else:
                label_map[label] = _parse_currency(raw)
        elif row is not None and col is not None:
            formula_cells.append((label, row, col))

    for start_row, end_row in (ISR_RANGE, IVA_RANGE):
        for row in range(start_row, end_row + 1):
            # Layout 1: D = label, E = value
            label_d = _cell_value(ws.cell(row=row, column=4))
            value_e = ws.cell(row=row, column=5).value
            add_label_value(label_d, value_e, row=row, col=5)
            # Layout 2: E = label, F = value (only when E looks like text)
            label_e_cell = ws.cell(row=row, column=5)
            value_f_cell = ws.cell(row=row, column=6)
            label_e = _cell_value(label_e_cell)
            value_f = value_f_cell.value
            if label_e is not None and str(label_e).strip():
                label_e_str = str(label_e).strip()
                if not re.match(r"^[\d\s\$,.\-]+$", label_e_str):
                    add_label_value(label_e, value_f, row=row, col=6)

    # Resolve formula cells to get value-as-displayed (e.g. =+F8 -> value of F8; resolve ref recursively if it's a formula)
    # With data_only=True, formula cells often have a cached value (what you see in Excel); if missing, we resolve simple refs.
    def _cell_display_value(ws_data, ws_formula, r: int, c: int, visited: set | None = None) -> float:
        visited = visited or set()
        key = (r, c)
        if key in visited:
            return 0.0
        visited.add(key)
        raw = ws_data.cell(row=r, column=c).value
        # Use cached value if present (what user sees when file was saved in Excel) — works for both literal and formula cells
        if raw is not None and not (isinstance(raw, str) and str(raw).strip().startswith("=")):
            return _parse_currency(raw)
        # No cached value: resolve formula or use literal from formula sheet
        formula = ws_formula.cell(row=r, column=c).value
        if isinstance(formula, (int, float)):
            return float(formula)
        ref = _parse_cell_ref(formula) if isinstance(formula, str) else None
        if ref:
            r2, c2 = ref
            return _cell_display_value(ws_data, ws_formula, r2, c2, visited)
        return 0.0

    if formula_cells:
        try:
            wb_formula = openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)
            ws_formula = wb_formula[IMPUESTOS_SHEET]
            for label, row, col in formula_cells:
                formula = ws_formula.cell(row=row, column=col).value
                ref = _parse_cell_ref(formula) if isinstance(formula, str) else None
                if ref:
                    r2, c2 = ref
                    val = _cell_display_value(ws, ws_formula, r2, c2)
                    label_map[label] = val
                    LOG.info("Excel formula resolved (value as displayed): %r at (%s,%s) -> ref %s = %.2f", formula, row, col, ref, val)
                    if val == 0.0 and label == "Base gravable del pago provisional":
                        fallback = label_map.get("Ingresos cobrados y amparados por factura del mes")
                        if fallback is not None and _parse_currency(fallback) != 0.0:
                            label_map[label] = _parse_currency(fallback)
                            LOG.info("Excel: using fallback for Base gravable: %.2f", label_map[label])
            wb_formula.close()
        except Exception as e:
            LOG.warning("Could not resolve formula cells: %s", e)

    wb.close()

    # Optional: when openpyxl gave 0.0 or no value, try xlcalculator (build model once, then evaluate each cell)
    evaluator, xlcalc_err = _xlcalculator_evaluator(workbook_path)
    if xlcalc_err and evaluator is None:
        if xlcalc_err != "not installed":
            LOG.warning("xlcalculator: %s", xlcalc_err)
    elif evaluator is not None:
        for label, row, col in formula_cells:
            current = label_map.get(label)
            if current is not None and abs(float(current)) > 1e-9:
                continue  # keep existing non-zero value from cache/simple-ref
            xval = _evaluate_cell(evaluator, IMPUESTOS_SHEET, row, col)
            # If formula cell (e.g. F9 =+F8) didn't evaluate, try the referenced cell (F8) directly
            if xval is None:
                try:
                    wb_f = openpyxl.load_workbook(workbook_path, read_only=False, data_only=False)
                    ws_f = wb_f[IMPUESTOS_SHEET]
                    formula = ws_f.cell(row=row, column=col).value
                    wb_f.close()
                    ref = _parse_cell_ref(formula) if isinstance(formula, str) else None
                    if ref:
                        r2, c2 = ref
                        xval = _evaluate_cell(evaluator, IMPUESTOS_SHEET, r2, c2)
                except Exception as e:
                    LOG.debug("xlcalculator ref fallback for (%s,%s): %s", row, col, e)
            if xval is not None:
                label_map[label] = xval
                LOG.info("Excel value (xlcalculator): %r at (%s,%s) = %.2f", label, row, col, xval)
            else:
                LOG.info("xlcalculator: no value for %r at (%s,%s) (tried cell and ref fallback)", label, row, col)

    LOG.info(
        "Excel opened: %s | Sheet=%s, layouts D/E and E/F, rows ISR %s–%s, IVA %s–%s | %s labels read. "
        "'Base gravable del pago provisional' in sheet: %s",
        workbook_path, IMPUESTOS_SHEET, ISR_RANGE[0], ISR_RANGE[1], IVA_RANGE[0], IVA_RANGE[1],
        len(label_map), "Base gravable del pago provisional" in label_map,
    )
    if label_map and "Base gravable del pago provisional" not in label_map:
        sample = [k for k in list(label_map.keys())[:15]]
        LOG.info("Impuestos column D labels (sample): %s", sample)

    # Period from filename: YYYYMM_CustomerRfc_Hoja de Trabajo.xlsx
    year, month = None, None
    basename = os.path.basename(workbook_path)
    m = re.match(r"^(\d{4})(\d{2})_", basename)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))

    # Periodicidad: default 1 mensual; could be read from sheet if present
    periodicidad = label_map.get("Periodicidad") or 1
    if isinstance(periodicidad, str):
        periodicidad = 1

    tipo_declaracion = "Normal"
    periodo_str = f"{month:02d}" if month is not None else "(none)"
    LOG.info(
        "Phase 2: initial form data to fill: Ejercicio=%s, Periodicidad=%s, Período=%s, Tipo=%s (year/month from filename YYYYMM_; periodicidad from sheet)",
        year, periodicidad, periodo_str, tipo_declaracion,
    )
    return {
        "label_map": label_map,
        "year": year,
        "month": month,
        "periodicidad": periodicidad,
        "tipo_declaracion": tipo_declaracion,
    }


def get_efirma_from_db(company_id: int, branch_id: int, config: dict) -> dict:
    """
    Call Contaayuda SP [GET_AUTOMATICTAXDECLARATION_CUSTOMERDATA].
    Returns dict with cer_path, key_path, password, rfc (TAXID).
    """
    if pyodbc is None:
        raise RuntimeError("pyodbc is required for DB access. Install with: pip install pyodbc")
    conn_str = config.get("db_connection_string")
    if not conn_str:
        raise ValueError("config must set db_connection_string")
    base = config.get("fiel_certificate_base_path", "").rstrip("/\\")
    if not base:
        raise ValueError("config must set fiel_certificate_base_path")

    conn = pyodbc.connect(conn_str)
    try:
        cursor = conn.cursor()
        cursor.execute(
            f"EXEC {SP_GET_EFIRMA} @CompanyId = ?, @BranchId = ?",
            (company_id, branch_id),
        )
        row = cursor.fetchone()
        if not row:
            raise ValueError(f"No e.firma data for CompanyId={company_id}, BranchId={branch_id}")
        columns = [col[0] for col in cursor.description]
        r = dict(zip(columns, row))
        cer_name = r.get("FIELXMLCERTIFICATE") or r.get("FielXmlCertificate")
        key_name = r.get("FIELXMLKEY") or r.get("FielXmlKey")
        password = r.get("FIELTIMBARDOPASSWORD") or r.get("FielTimbardoPassword")
        rfc = r.get("TAXID") or r.get("TaxId") or ""
        if not cer_name or not key_name:
            raise ValueError("SP did not return FIELXMLCERTIFICATE / FIELXMLKEY")
        folder = os.path.join(base, str(company_id), str(branch_id))
        cer_path = os.path.join(folder, cer_name)
        key_path = os.path.join(folder, key_name)
        if not os.path.isfile(cer_path):
            raise FileNotFoundError(f"Certificate file not found: {cer_path}")
        if not os.path.isfile(key_path):
            raise FileNotFoundError(f"Key file not found: {key_path}")
        return {
            "cer_path": os.path.abspath(cer_path),
            "key_path": os.path.abspath(key_path),
            "password": password or "",
            "rfc": rfc,
        }
    finally:
        conn.close()


def load_config(path: str | None) -> dict:
    """Load config JSON. If path is None, try config.json in script dir."""
    if path is None:
        path = os.path.join(os.path.dirname(__file__), "config.json")
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Config not found: {path}. Copy config.example.json to config.json and edit.")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_mapping(path: str | None) -> dict:
    """Load form_field_mapping.json."""
    if path is None:
        path = os.path.join(os.path.dirname(__file__), "form_field_mapping.json")
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Mapping not found: {path}")
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {k: v for k, v in data.items() if not k.startswith("_comment") and isinstance(v, list)}


# --- SAT Automatic Declaration API (dynamic Excel/CER/KEY/password) ---

def get_pending_declarations(company_id: int, branch_id: int, base_url: str) -> list[dict]:
    """
    GET GetPendingDeclarations; return list of declaration objects from response data.
    On non-200 or missing data, log and return [].
    """
    if requests is None:
        raise RuntimeError("requests is required for API mode. Install with: pip install requests")
    url = base_url.rstrip("/") + "/GetPendingDeclarations"
    params = {"companyId": str(company_id), "branchId": str(branch_id)}
    try:
        resp = requests.get(url, params=params, timeout=60)
        resp.raise_for_status()
        body = resp.json()
    except requests.RequestException as e:
        LOG.warning("GetPendingDeclarations request failed: %s", e)
        return []
    except (ValueError, KeyError) as e:
        LOG.warning("GetPendingDeclarations invalid JSON: %s", e)
        return []
    status = body.get("statusCode")
    if status != 200:
        LOG.warning("GetPendingDeclarations statusCode=%s message=%s", status, body.get("message", ""))
        return []
    data = body.get("data")
    if not isinstance(data, list):
        return []
    return data


def download_file(url: str, dest_path: str) -> None:
    """Download url to dest_path. Raises RuntimeError on non-2xx or IO error."""
    if requests is None:
        raise RuntimeError("requests is required. Install with: pip install requests")
    try:
        resp = requests.get(url, stream=True, timeout=60)
        resp.raise_for_status()
        os.makedirs(os.path.dirname(dest_path) or ".", exist_ok=True)
        with open(dest_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
    except requests.RequestException as e:
        raise RuntimeError(f"Download failed {url!r}: {e}") from e
    except OSError as e:
        raise RuntimeError(f"Write failed {dest_path!r}: {e}") from e


def prepare_declaration_from_api(declaration: dict, download_dir: str) -> tuple[str, dict]:
    """
    Download Excel, CER, KEY from declaration URLs; return (workbook_path, efirma_dict).
    efirma_dict has cer_path, key_path, password, rfc. Normalizes FIEL* / Fiel* key names.
    """
    def _get(d: dict, *keys: str) -> str:
        for k in keys:
            v = d.get(k)
            if v:
                return str(v).strip()
        return ""

    excel_url = _get(declaration, "EXCELFILEPATH", "ExcelFilePath")
    cer_url = _get(declaration, "FIELXMLCERTIFICATE", "FielXmlCertificate")
    key_url = _get(declaration, "FIELXMLKEY", "FielXmlKey")
    password = _get(declaration, "FIELTIMBARDOPASSWORD", "FielTimbardoPassword")
    decl_id = _get(declaration, "DECLARATIONID", "DeclarationId") or "0"
    rfc = _get(declaration, "TAXID", "TaxId")

    if not excel_url or not cer_url or not key_url:
        raise ValueError("Declaration missing EXCELFILEPATH, FIELXMLCERTIFICATE, or FIELXMLKEY")

    safe_id = re.sub(r"[^\w\-]", "_", decl_id)
    excel_path = os.path.join(download_dir, f"declaration_{safe_id}.xlsx")
    cer_path = os.path.join(download_dir, f"declaration_{safe_id}.cer")
    key_path = os.path.join(download_dir, f"declaration_{safe_id}.key")

    LOG.info("Downloading declaration files (DECLARATIONID=%s) to %s", decl_id, download_dir)
    LOG.info("  Excel: %s -> %s", excel_url, excel_path)
    download_file(excel_url, excel_path)
    LOG.info("  CER:   %s -> %s", cer_url, cer_path)
    download_file(cer_url, cer_path)
    LOG.info("  KEY:   %s -> %s", key_url, key_path)
    download_file(key_url, key_path)
    LOG.info("Downloaded workbook, CER, KEY for declaration %s", decl_id)

    return excel_path, {
        "cer_path": os.path.abspath(cer_path),
        "key_path": os.path.abspath(key_path),
        "password": password or "",
        "rfc": rfc or "",
    }


def mark_processing(company_id: int, branch_id: int, declaration_id: int, base_url: str) -> bool:
    """POST MarkProcessing; return True if statusCode == 200."""
    if requests is None:
        raise RuntimeError("requests is required for API mode. Install with: pip install requests")
    url = base_url.rstrip("/") + "/MarkProcessing"
    payload = {
        "companyId": str(company_id),
        "branchId": str(branch_id),
        "declarationId": str(declaration_id),
    }
    try:
        resp = requests.post(url, json=payload, timeout=30)
        resp.raise_for_status()
        body = resp.json()
    except requests.RequestException as e:
        LOG.warning("MarkProcessing request failed: %s", e)
        return False
    ok = body.get("statusCode") == 200
    if not ok:
        LOG.warning("MarkProcessing statusCode=%s message=%s", body.get("statusCode"), body.get("message", ""))
    return ok


def mark_completed(company_id: int, branch_id: int, declaration_id: int, pdf_file_path: str, base_url: str) -> bool:
    """
    POST MarkCompleted with pdfFilePath. Return True if statusCode == 200.
    Note: Backend may expect a server-relative path; upload may be required separately.
    """
    if requests is None:
        raise RuntimeError("requests is required for API mode. Install with: pip install requests")
    url = base_url.rstrip("/") + "/MarkCompleted"
    payload = {
        "companyId": str(company_id),
        "branchId": str(branch_id),
        "declarationId": str(declaration_id),
        "pdfFilePath": pdf_file_path,
    }
    try:
        resp = requests.post(url, json=payload, timeout=30)
        resp.raise_for_status()
        body = resp.json()
    except requests.RequestException as e:
        LOG.warning("MarkCompleted request failed: %s", e)
        return False
    ok = body.get("statusCode") == 200
    if not ok:
        LOG.warning("MarkCompleted statusCode=%s message=%s", body.get("statusCode"), body.get("message", ""))
    return ok


# Default SAT UI labels/patterns (config.json sat_ui). When SAT changes wording, update config; code merges with these.
DEFAULT_SAT_UI = {
    "isr_ingresos_copropiedad": "*¿Los ingresos fueron obtenidos a través de copropiedad?",
    "btn_administracion_declaracion": r"ADMINISTRACIÓN\s+DE\s+LA\s+DECLARACIÓN",
    "select_obligation_isr": r"ISR\s+simplificado\s+de\s+confianza\.\s*Personas\s+físicas",
    "select_obligation_iva": r"IVA\s+simplificado\s+de\s+confianza",
    "total_ingresos_cobrados": "Total de ingresos efectivamente cobrados",
    "capturar_button": "CAPTURAR",
    "isr_section_descuentos": "Descuentos",
    "isr_section_ingresos_disminuir": "Ingresos a disminuir",
    "isr_section_ingresos_adicionales": "Ingresos adicionales",
    "isr_section_total_percibidos": "Total de ingresos percibidos",
    "iva_section_acreditable": "IVA acreditable del periodo",
    "popup_cerrar": "CERRAR",
    "loading_text": "Cargando información",
    "formularios_no_enviados": "Formularios no enviados",
    "iniciar_nueva_declaracion": r"INICIAR\s*.*\s*NUEVA\s+DECLARACIÓN",
    "eliminar_confirm": "¿Deseas eliminar esta declaración?",
    "eliminar_si": "sí",
    "login_enviar": "Enviar",
    "login_e_firma": r"e\.firma",
    "nav_presentar_declaracion": "Presentar declaración",
    "nav_nuevo_portal": "Nuevo Portal de pagos provisionales",
    "nav_iniciar_nueva": "Iniciar una nueva declaración",
    "btn_siguiente": "SIGUIENTE",
    "initial_ejercicio_label": "Ejercicio",
    "initial_periodicidad_label": "Periodicidad",
    "initial_periodo_label": "Período",
    "initial_tipo_label": "Tipo de Declaración",
    "ver_detalle_button": "VER DETALLE",
    "isr_retenido_row_label": "ISR retenido por personas morales",
    "isr_retenido_no_acreditable_label": "ISR retenido no acreditable",
    "isr_retenido_excel_label": "ISR retenido",
    "determinacion_tab_name": "Determinación",
    "pago_tab_name": "Pago",
    "isr_pago_compensaciones_question": "compensaciones por aplicar",
    "isr_pago_estimulos_question": "estímulos por aplicar",
    "btn_guardar": "GUARDAR",
}
DEFAULT_ISR_DETERMINACION_LABELS = [
    "Ingresos nominales facturados",
    "Total de ingresos acumulados",
    "Base gravable del pago provisional",
    "Impuesto del periodo",
    "Total ISR retenido del periodo",
    "ISR a cargo",
]
DEFAULT_DECLARATION_FLOW = ["isr", "iva"]

DEFAULT_IVA_DETERMINACION_FIELDS = [
    ("Actividades gravadas a la tasa del 16%", "Actividades gravadas a la tasa del 16%"),
    ("Actividades gravadas a la tasa del 0% otros", "Actividades gravadas a la tasa del 0%"),
    ("Actividades exentas", "Actividades exentas"),
    ("Actividades no objeto de impuesto", "Actividades no objeto del impuesto"),
    ("IVA retenido a favor", "IVA retenido"),
]
DEFAULT_IVA_PAGO_LABELS = [
    "Actividades gravadas a la tasa del 16%",
    "Actividades gravadas a la tasa del 8%",
    "Actividades gravadas a la tasa del 0% otros",
    "Actividades exentas",
    "Actividades no objeto de impuesto",
    "IVA a cargo a la tasa del 16% y 8%",
    "Total IVA Trasladado",
    "IVA retenido a favor",
    "IVA acreditable del periodo",
    "Cantidad a cargo",
    "IVA a cargo",
    "IVA a favor",
]


def get_sat_ui(config: dict) -> dict:
    """Return merged sat_ui from config with defaults. Safe when config has no sat_ui."""
    ui = dict(DEFAULT_SAT_UI)
    ui.update(config.get("sat_ui") or {})
    return ui


def get_isr_determinacion_labels(config: dict) -> list[str]:
    """Return isr_determinacion_labels from config or default list."""
    labels = config.get("isr_determinacion_labels")
    return list(labels) if labels else list(DEFAULT_ISR_DETERMINACION_LABELS)


def get_declaration_flow(config: dict) -> list[str]:
    """Return declaration_flow from config or default ['isr', 'iva']."""
    flow = config.get("declaration_flow")
    return list(flow) if flow else list(DEFAULT_DECLARATION_FLOW)


def get_iva_determinacion_fields(config: dict) -> list[tuple[str, str]]:
    """Return iva_determinacion_fields from config or default (excel_label, form_label) tuples."""
    raw = config.get("iva_determinacion_fields")
    if not raw:
        return list(DEFAULT_IVA_DETERMINACION_FIELDS)
    out = []
    for item in raw:
        if isinstance(item, dict):
            ex = item.get("excel_label", "")
            fm = item.get("form_label", ex)
            out.append((ex, fm))
        else:
            out.append((item, item))
    return out if out else list(DEFAULT_IVA_DETERMINACION_FIELDS)


def get_iva_pago_labels(config: dict) -> list[str]:
    """Return iva_pago_labels from config or default list."""
    labels = config.get("iva_pago_labels")
    return list(labels) if labels else list(DEFAULT_IVA_PAGO_LABELS)


def _ui_pattern(sat_ui: dict, key: str, *, literal: bool = False) -> re.Pattern:
    """Compile sat_ui key as case-insensitive regex. If literal=True, escape special chars."""
    raw = sat_ui.get(key) or DEFAULT_SAT_UI.get(key) or ""
    if literal:
        raw = re.escape(raw)
    return re.compile(raw, re.I)


def setup_logging(log_file: str | None) -> None:
    """Print to console and append to log file."""
    if log_file is None:
        log_file = "sat_declaration_filler.log"
    log_path = os.path.abspath(log_file)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(log_path, encoding="utf-8", mode="a"),
        ],
    )
    LOG.info("Log file: %s", log_path)


_FILL_SELECTOR_TIMEOUT_MS = 1000
_INITIAL_FORM_SELECTOR_TIMEOUT_MS = 1000

# SAT Periodicidad dropdown: option labels shown in UI → we use value= for select_option.
# Options: 1-Mensual, 3-Trimestral, 4-Cuatrimestral, 5-Semestral (A), 6-Semestral (B) Liquidación,
#          7-Ajuste, 8-Del Ejercicio, 9-Sin Periodo.
_SAT_PERIODICIDAD_VALUE = {
    1: "M",   # 1-Mensual
    3: "T",   # 3-Trimestral
    4: "Q",   # 4-Cuatrimestral
    5: "S",   # 5-Semestral (A)
    6: "L",   # 6-Semestral (B) Liquidación
    7: "J",   # 7-Ajuste
    8: "Y",   # 8-Del Ejercicio
    9: "N",   # 9-Sin Periodo
}

# SAT Ejercicio dropdown: options are years 2022–2026 (we pass str(year), e.g. "2026").

# SAT Período dropdown: options are months Enero–Diciembre (labels in Spanish). We pass label for select_option.
_SAT_PERIODO_LABEL = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def _try_fill(scope: Page | Frame, page_for_wait: Page, mapping: dict, key: str, value: str | float, *, is_file: bool = False) -> bool:
    """Try selectors for key; fill first that exists. scope = page or frame where elements live; page_for_wait for timeouts."""
    selectors = mapping.get(key)
    if not selectors:
        return False
    # Login: short timeouts for 5s target. Initial form: longer. ISR/IVA form fields: fast (~1s per textbox).
    if key.startswith("_login_"):
        timeout_ms = 400
    elif key.startswith("initial_"):
        timeout_ms = _INITIAL_FORM_SELECTOR_TIMEOUT_MS
    else:
        timeout_ms = _TEXTBOX_FILL_VISIBLE_MS  # form fields (ISR/IVA): fast visible wait
    value_str = str(value)
    for sel in selectors:
        try:
            # SAT e.firma: two input[type='file'] with no name/accept; first=cer, second=key.
            if is_file and sel == "input[type='file']" and key in ("_login_cer_file_input", "_login_key_file_input"):
                loc = scope.locator(sel)
                if loc.count() < 2 and key == "_login_key_file_input":
                    continue
                target = loc.first if key == "_login_cer_file_input" else loc.nth(1)
                target.wait_for(state="attached", timeout=min(300, timeout_ms))
                target.set_input_files(value)
                return True
            # Label-based: find control by its visible label (works for dropdowns and inputs).
            if sel.startswith("label="):
                label_text = sel[6:].strip()
                loc = scope.get_by_label(label_text, exact=False)
                if loc.count() == 0:
                    continue
                first = loc.first
                first.wait_for(state="visible", timeout=timeout_ms or 1000)
                if is_file:
                    first.set_input_files(value)
                else:
                    tag = first.evaluate("el => el.tagName.toLowerCase()")
                    if tag == "select":
                        try:
                            first.select_option(value=value_str)
                        except Exception:
                            first.select_option(label=value_str)
                    else:
                        first.click()
                        page_for_wait.wait_for_timeout((100 if (timeout_ms or 0) <= 500 else 300) if key.startswith("initial_") else _OBLIGATION_CLICK_WAIT_MS)
                        first.fill(value_str)
                return True
            loc = scope.locator(sel)
            if loc.count() == 0:
                continue
            first = loc.first
            first.wait_for(state="visible", timeout=timeout_ms or 1000)
            if is_file:
                first.set_input_files(value)
            else:
                tag = first.evaluate("el => el.tagName.toLowerCase()")
                if tag == "select":
                    try:
                        first.select_option(value=value_str)
                    except Exception:
                        first.select_option(label=value_str)
                else:
                    first.click()
                    page_for_wait.wait_for_timeout((100 if (timeout_ms or 0) <= 500 else 300) if key.startswith("initial_") else _OBLIGATION_CLICK_WAIT_MS)
                    first.fill(value_str)
            return True
        except Exception:
            continue
    return False


# Phrases that indicate SAT portal error (page down, 500, maintenance). Checked case-insensitive.
# Use "http 500" / "error 500" instead of bare "500" to avoid false positives from amounts/IDs on page.
_SAT_ERROR_PHRASES = (
    "http 500",
    "error 500",
    "internal server error",
    "error del servidor",
    "servidor no disponible",
    "no disponible",
    "service unavailable",
    "unavailable",
    "error http",
    "mantenimiento",
    "maintenance",
    "try again later",
    "intente más tarde",
)


def _check_sat_page_error(page, *, response_status: int | None = None) -> str | None:
    """
    Check if the current page shows an SAT error (500, down, maintenance).
    Returns an error message to raise if an issue is detected, None otherwise.
    """
    if response_status is not None and response_status >= 400:
        return f"SAT returned HTTP {response_status}"
    try:
        body = page.locator("body").inner_text(timeout=3000) or ""
        text = body.lower()
        for phrase in _SAT_ERROR_PHRASES:
            if phrase in text:
                return f"SAT page shows error: found '{phrase}' on page"
    except Exception as e:
        LOG.debug("Could not read page body for error check: %s", e)
    return None


def _try_click(page, mapping: dict, key: str) -> bool:
    selectors = mapping.get(key)
    if not selectors:
        return False
    for sel in selectors:
        try:
            loc = page.locator(sel)
            if loc.count() == 0:
                continue
            first = loc.first
            first.wait_for(state="visible", timeout=1000)
            first.click()
            return True
        except Exception:
            continue
    return False


def login_sat(page, efirma: dict, mapping: dict, base_url: str = SAT_PORTAL_URL, sat_ui: dict | None = None) -> None:
    """Open SAT portal, click e.firma, fill .cer, .key, password, Enviar."""
    sat_ui = sat_ui or DEFAULT_SAT_UI
    t0 = time.perf_counter()

    def _ts() -> str:
        now = datetime.now()
        return now.strftime("%Y-%m-%d %H:%M:%S") + f",{now.microsecond // 1000:03d}"

    def _elapsed() -> float:
        return round(time.perf_counter() - t0, 2)

    # Use 'load' so we don't timeout: SAT page often never reaches networkidle (long-lived connections).
    response = page.goto(base_url, wait_until="load", timeout=60000)
    page.wait_for_load_state("domcontentloaded")
    err = _check_sat_page_error(page, response_status=response.status if response else None)
    if err:
        LOG.error("SAT issue on load: %s", err)
        print(err, file=sys.stderr)
        raise RuntimeError(err)
    LOG.info("Phase 1: [%.2fs] SAT page loaded, looking for e.firma button", _elapsed())

    if not _try_click(page, mapping, "_login_e_firma_button"):
        raise RuntimeError("Could not find e.firma button on SAT login page")
    LOG.info("Phase 1: [%.2fs] e.firma pressed", _elapsed())
    try:
        page.wait_for_url(re.compile(r".*id=fiel.*", re.I), timeout=200)
    except Exception:
        pass
    try:
        page.locator("input[type='file'], input[type='password']").first.wait_for(state="visible", timeout=300)
    except Exception:
        pass
    # Set cer, key, password with minimal delay.
    cer_input_ok = _try_fill(page, page, mapping, "_login_cer_file_input", efirma["cer_path"], is_file=True)
    if cer_input_ok:
        LOG.info("Phase 1: [%.2fs] filled .cer: %s", _elapsed(), os.path.basename(efirma['cer_path']))
    key_input_ok = _try_fill(page, page, mapping, "_login_key_file_input", efirma["key_path"], is_file=True)
    if key_input_ok:
        LOG.info("Phase 1: [%.2fs] filled .key: %s", _elapsed(), os.path.basename(efirma['key_path']))
    if not cer_input_ok or not key_input_ok:
        LOG.warning("One or both file inputs not found by selector; using generic input[type='file'] order (first=cer, second=key)")
        inputs = page.locator("input[type='file']").all()
        if len(inputs) >= 2:
            inputs[0].set_input_files(efirma["cer_path"])
            inputs[1].set_input_files(efirma["key_path"])
            LOG.info("Phase 1: [%.2fs] filled .cer (fallback): %s", _elapsed(), os.path.basename(efirma['cer_path']))
            LOG.info("Phase 1: [%.2fs] filled .key (fallback): %s", _elapsed(), os.path.basename(efirma['key_path']))
        elif len(inputs) == 1:
            inputs[0].set_input_files(efirma["cer_path"])
            LOG.info("Phase 1: [%.2fs] filled .cer (fallback): %s", _elapsed(), os.path.basename(efirma['cer_path']))
            LOG.warning("Only one file input found; key may need manual selection")

    pwd_ok = _try_fill(page, page, mapping, "_login_password_input", efirma["password"])
    if pwd_ok:
        LOG.info("Phase 1: [%.2fs] filled password: ***", _elapsed())
    else:
        LOG.warning("Phase 1: Password field not filled — check selectors")
        LOG.info("Phase 1: [%.2fs] password NOT filled (check selectors)", _elapsed())
    page.wait_for_timeout(50)
    LOG.info("Phase 1: [%.2fs] pressing Enviar", _elapsed())
    # Click Enviar: try mapping first, then fallbacks (SAT markup varies; Enviar can be button or input).
    enviar_pat = _ui_pattern(sat_ui, "login_enviar")
    if not _try_click(page, mapping, "_login_enviar_button"):
        enviar_clicked = False
        for try_fn in [
            lambda: page.get_by_role("button", name=enviar_pat).first.click(timeout=1200),
            lambda: page.locator("button").filter(has_text=enviar_pat).first.click(timeout=1200),
            lambda: page.locator("input[type='submit'][value='Enviar']").first.click(timeout=1200),
            lambda: page.locator("input[type='submit'][value*='Enviar']").first.click(timeout=1200),
        ]:
            try:
                try_fn()
                enviar_clicked = True
                LOG.info("Phase 1: Enviar pressed (fallback)")
                break
            except Exception:
                continue
        if not enviar_clicked:
            # Last resort: second submit button (first is often "Contraseña")
            try:
                page.locator("input[type='submit'], button[type='submit']").nth(1).click(timeout=1200)
                enviar_clicked = True
                LOG.info("Phase 1: Enviar pressed (nth(1))")
            except Exception:
                pass
        if not enviar_clicked:
            raise RuntimeError("Could not find Enviar button on e.firma form")
    else:
        LOG.info("Phase 1: [%.2fs] Enviar pressed", _elapsed())
    page.wait_for_load_state("domcontentloaded")
    LOG.info("Phase 1: [%.2fs] waiting for SAT post-login page...", _elapsed())
    post_login_timeout = 8000
    poll_ms = 150
    t_end = time.perf_counter() + (post_login_timeout / 1000.0)
    while time.perf_counter() < t_end:
        page.wait_for_timeout(poll_ms)
        try:
            url = page.url or ""
            if "clouda.sat.gob.mx" in url.lower():
                break
            body = (page.locator("body").inner_text(timeout=400) or "").lower()
            if "500" in body or "internal server error" in body or "error: http 500" in body:
                err = "SAT login returned HTTP 500 Internal Server Error (server-side). Try again later or check SAT status."
                LOG.error("SAT 500 after login: %s", err)
                print(err, file=sys.stderr)
                raise RuntimeError(err)
        except RuntimeError:
            raise
        except Exception:
            pass
    try:
        page.wait_for_url(re.compile(r"clouda\.sat\.gob\.mx", re.I), timeout=1000)
    except Exception:
        pass
    try:
        for sel in (mapping.get("_nav_presentar_declaracion") or []) + (mapping.get("_nav_nuevo_portal") or []) + ["text=Presentar declaración", "text=Cerrar Sesión", "text=Bienvenido"]:
            try:
                page.locator(sel).first.wait_for(state="visible", timeout=1000)
                break
            except Exception:
                continue
    except Exception:
        pass
    page.wait_for_timeout(80)
    LOG.info("Phase 1: [%.2fs] post-login page ready.", _elapsed())
    err = _check_sat_page_error(page)
    if err:
        LOG.error("SAT issue after login: %s", err)
        print(err, file=sys.stderr)
        raise RuntimeError(err)


def navigate_to_declaration(page, mapping: dict) -> None:
    """Click Nuevo Portal → Presentar Declaración → Iniciar una nueva declaración (legacy portal)."""
    _try_click(page, mapping, "_nav_nuevo_portal")
    page.wait_for_timeout(1000)
    _try_click(page, mapping, "_nav_presentar_declaracion")
    page.wait_for_timeout(1000)
    _try_click(page, mapping, "_nav_iniciar_nueva")
    page.wait_for_timeout(1000)


DRAFT_PAGE_WAIT_MS = 6000   # allow up to 6 s for draft page to load (was 2s; page can render slowly)
DRAFT_POLL_MS = 150          # poll page body every 150 ms
DRAFT_INITIAL_WAIT_MS = 400  # brief wait for navigation/content to start rendering before first check
DRAFT_BODY_TIMEOUT_MS = 800  # per-call timeout for body.inner_text so we can poll multiple times


def dismiss_draft_if_present(page: Page, mapping: dict, sat_ui: dict | None = None) -> bool:
    """If 'Formulario no concluido' is shown (saved draft), click trash icon and confirm 'Sí' to delete; then we can continue to initial form. Returns True if a draft was dismissed."""
    sat_ui = sat_ui or DEFAULT_SAT_UI
    LOG.info("Phase 2: Checking for draft declaration (Formulario no concluido) after Presentar declaración, before filling initial form...")
    page.wait_for_timeout(DRAFT_INITIAL_WAIT_MS)
    # Broader markers: SAT may show "Formulario no concluido", "Formularios no enviados", "borrador", "sin enviar", etc.
    draft_markers = (
        "formulario no concluido",
        "formularios no enviados",
        "formularios no concluidos",
        "declaraciones no enviadas",
        "no concluido",
        "no enviados",
        "borrador",
        "sin enviar",
    )
    t_end = (time.perf_counter() * 1000) + DRAFT_PAGE_WAIT_MS
    draft_found = False
    last_body_snippet = ""
    while (time.perf_counter() * 1000) < t_end:
        try:
            body = (page.locator("body").inner_text(timeout=DRAFT_BODY_TIMEOUT_MS) or "").lower()
            if any(m in body for m in draft_markers):
                draft_found = True
                break
            last_body_snippet = (body[:500] + "..." if len(body) > 500 else body) or "(empty)"
            # If content is in an iframe, check main content iframe(s) as well
            for frame in page.frames:
                if frame == page.main_frame:
                    continue
                try:
                    iframe_body = (frame.locator("body").inner_text(timeout=DRAFT_BODY_TIMEOUT_MS) or "").lower()
                    if any(m in iframe_body for m in draft_markers):
                        draft_found = True
                        break
                    if len(iframe_body) > len(body):
                        last_body_snippet = (iframe_body[:500] + "..." if len(iframe_body) > 500 else iframe_body) or "(empty)"
                except Exception:
                    pass
            if draft_found:
                break
        except Exception:
            pass
        page.wait_for_timeout(DRAFT_POLL_MS)
    if not draft_found:
        LOG.info("Phase 2: No draft declaration detected; proceeding to configuration form.")
        if last_body_snippet and LOG.isEnabledFor(logging.DEBUG):
            LOG.debug("Draft check: page text snippet (first 500 chars): %s", last_body_snippet[:500])
        return False
    LOG.info("Phase 2: Formulario no concluido detected; dismissing saved draft (trash → Sí)")
    try:
        # Click trash icon (first one next to a draft row)
        trash_clicked = False
        if mapping.get("_draft_trash"):
            for sel in mapping["_draft_trash"]:
                try:
                    loc = page.locator(sel).first
                    loc.wait_for(state="visible", timeout=1000)
                    loc.click()
                    trash_clicked = True
                    break
                except Exception:
                    continue
        if not trash_clicked:
            for fallback_selector in [
                "[aria-label*='eliminar']", "[aria-label*='Eliminar']",
                "[title*='eliminar']", "[title*='Eliminar']",
                "button:has(svg)", "a:has(svg)",
                "[class*='trash']", "[class*='eliminar']", "[class*='borrar']",
            ]:
                try:
                    loc = page.locator(fallback_selector).first
                    loc.wait_for(state="visible", timeout=800)
                    loc.click()
                    trash_clicked = True
                    break
                except Exception:
                    continue
        if not trash_clicked:
            form_no_env = sat_ui.get("formularios_no_enviados") or "Formularios no enviados"
            iniciar_pat = _ui_pattern(sat_ui, "iniciar_nueva_declaracion")
            try:
                card = page.get_by_text(form_no_env, exact=False).locator("..").locator("..").locator("..").first
                trash = card.locator("button, a").filter(has_not=page.get_by_text(iniciar_pat)).first
                trash.wait_for(state="visible", timeout=1000)
                trash.click()
                trash_clicked = True
            except Exception:
                try:
                    row = page.get_by_text(form_no_env, exact=False).locator("..").locator("..")
                    trash = row.locator("button, a").filter(has=page.locator("svg, [class*='trash'], [class*='eliminar']")).first
                    trash.wait_for(state="visible", timeout=1000)
                    trash.click()
                    trash_clicked = True
                except Exception:
                    pass
        if not trash_clicked:
            LOG.warning("Could not find/click draft trash icon")
            return False
        page.wait_for_timeout(200)
        eliminar_confirm = sat_ui.get("eliminar_confirm") or "¿Deseas eliminar esta declaración?"
        eliminar_si = sat_ui.get("eliminar_si") or "sí"
        try:
            page.get_by_text(eliminar_confirm, exact=False).wait_for(state="visible", timeout=1000)
        except Exception:
            LOG.warning("Delete confirmation popup did not appear")
            return True
        si_clicked = False
        if mapping.get("_popup_eliminar_si"):
            for sel in mapping["_popup_eliminar_si"]:
                try:
                    page.locator(sel).first.click(timeout=1000)
                    si_clicked = True
                    break
                except Exception:
                    continue
        if not si_clicked:
            si_pat = _ui_pattern(sat_ui, "eliminar_si", literal=True)
            try:
                page.get_by_role("button", name=si_pat).first.click(timeout=1000)
                si_clicked = True
            except Exception:
                pass
        if not si_clicked:
            try:
                page.get_by_text(eliminar_si, exact=True).first.click(timeout=1000)
                si_clicked = True
            except Exception:
                pass
        if si_clicked:
            LOG.info("Phase 2: Draft deleted (Sí confirmed); continuing to initial form")
        page.wait_for_timeout(300)
        return True
    except Exception as e:
        LOG.warning("Dismiss draft failed: %s", e)
        return False


def open_configuration_form(page: Page, mapping: dict) -> bool:
    """Open 'Configuración de la declaración' by clicking 'Presentar declaración'. If draft page loads (no select), avoid long wait so draft check runs quickly."""
    ok = _try_click(page, mapping, "_nav_presentar_declaracion")
    if ok:
        page.wait_for_timeout(1000)
        try:
            page.locator("select").first.wait_for(state="visible", timeout=1200)
        except Exception:
            pass
    return ok


def transition_initial_to_phase3(page: Page, mapping: dict, sat_ui: dict | None = None) -> bool:
    """After initial form: click SIGUIENTE, wait for loading to finish, then click CERRAR on the pre-fill info pop-up. Returns True if the full sequence succeeded."""
    sat_ui = sat_ui or DEFAULT_SAT_UI
    if not _try_click(page, mapping, "_btn_siguiente"):
        LOG.warning("Could not click SIGUIENTE after initial form")
        return False
    page.wait_for_timeout(500)
    # Wait for loading text to disappear (variable time)
    loading_text = sat_ui.get("loading_text") or "Cargando información"
    try:
        loading = page.get_by_text(loading_text, exact=False)
        for _ in range(PHASE3_LOADING_MAX_WAIT_SEC * 2):  # poll every 500ms
            if loading.count() == 0:
                break
            try:
                if not loading.first.is_visible():
                    break
            except Exception:
                break
            page.wait_for_timeout(500)
    except Exception:
        pass
    page.wait_for_timeout(300)
    # Wait for pre-fill pop-up to appear (CERRAR button can take several seconds after loading ends)
    cerrar_pat = _ui_pattern(sat_ui, "popup_cerrar")
    cerrar_btn = page.get_by_role("button", name=cerrar_pat).first
    try:
        cerrar_btn.wait_for(state="visible", timeout=PHASE3_POPUP_WAIT_FOR_CERRAR_SEC * 1000)
        LOG.info("Pre-fill pop-up visible, clicking CERRAR")
    except Exception as e:
        LOG.warning("Pre-fill pop-up CERRAR button did not appear within %ss: %s", PHASE3_POPUP_WAIT_FOR_CERRAR_SEC, e)
    cerrar_ok = False
    try:
        cerrar_btn.click(timeout=PHASE3_POPUP_CERRAR_CLICK_MS)
        cerrar_ok = True
        LOG.info("Clicked CERRAR on pre-fill pop-up")
    except Exception:
        pass
    if not cerrar_ok and mapping.get("_popup_cerrar"):
        for sel in mapping["_popup_cerrar"]:
            try:
                btn = page.locator(sel).first
                btn.wait_for(state="visible", timeout=PHASE3_POPUP_CERRAR_SELECTOR_MS)
                btn.click()
                cerrar_ok = True
                LOG.info("Clicked CERRAR on pre-fill pop-up (mapping)")
                break
            except Exception:
                continue
    if not cerrar_ok:
        try:
            page.get_by_role("button", name=cerrar_pat).first.click(timeout=PHASE3_POPUP_CERRAR_CLICK_MS)
            cerrar_ok = True
            LOG.info("Clicked CERRAR on pre-fill pop-up (fallback)")
        except Exception as e:
            LOG.warning("Could not click CERRAR on pop-up: %s", e)
    if cerrar_ok:
        page.wait_for_timeout(1500)
    return cerrar_ok


# Post-click wait for ISR/IVA obligation selection; match initial form pacing (80–100 ms between dropdowns).
_OBLIGATION_CLICK_WAIT_MS = 80

# Textbox fill (ISR/IVA): visible/scroll timeouts must allow SAT form to render; too low (e.g. 200ms) causes
# label-based fill to fail and forces slow position fallback, longer runtime and flaky fills (e.g. mixtas).
_TEXTBOX_FILL_VISIBLE_MS = 350
_TEXTBOX_FILL_WAIT_MS = 25
# Phase 4 ISR retenido popup: modal can render slower; use longer timeouts so fill succeeds (was broken by aggressive 350ms).
_PHASE4_POPUP_VISIBLE_MS = 1000
_PHASE4_POPUP_WAIT_MS = 80

# Test mode: how long to leave the browser open after a successful run (ms). 10s = 10_000 ms.
_TEST_INSPECTION_WAIT_MS = 10_000


def open_obligation_isr(page, mapping: dict) -> bool:
    """Select 'ISR simplificado de confianza. Personas físicas' (checkmark + label) to open the ISR section; then the Ingresos form loads. Returns True if clicked."""
    ok = _try_click(page, mapping, "_select_obligation_isr")
    if ok:
        page.wait_for_timeout(_OBLIGATION_CLICK_WAIT_MS)
    return ok


def click_administracion_declaracion(page: Page, mapping: dict, sat_ui: dict | None = None) -> bool:
    """Click the 'ADMINISTRACIÓN DE LA DECLARACIÓN' button (top right on ISR/IVA form) to return to the administración menu. Returns True if clicked."""
    sat_ui = sat_ui or DEFAULT_SAT_UI
    LOG.info("IVA: step — clicking 'ADMINISTRACIÓN DE LA DECLARACIÓN' to return to administración menu")
    ok = _try_click(page, mapping, "_btn_administracion_declaracion")
    if ok:
        LOG.info("IVA: clicked 'ADMINISTRACIÓN DE LA DECLARACIÓN' (mapping)")
        page.wait_for_timeout(_OBLIGATION_CLICK_WAIT_MS)
        return True
    admin_pat = _ui_pattern(sat_ui, "btn_administracion_declaracion")
    for btn in page.get_by_role("button", name=admin_pat).all():
        try:
            if btn.is_visible():
                btn.click(timeout=2000)
                LOG.info("IVA: clicked 'ADMINISTRACIÓN DE LA DECLARACIÓN' (button role)")
                page.wait_for_timeout(_OBLIGATION_CLICK_WAIT_MS)
                return True
        except Exception:
            continue
    for elem in page.get_by_text(admin_pat).all():
        try:
            if elem.is_visible() and elem.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").count() == 0:
                elem.click(timeout=2000)
                LOG.info("IVA: clicked 'ADMINISTRACIÓN DE LA DECLARACIÓN' (text)")
                page.wait_for_timeout(_OBLIGATION_CLICK_WAIT_MS)
                return True
        except Exception:
            continue
    LOG.warning("IVA: could not click 'ADMINISTRACIÓN DE LA DECLARACIÓN'")
    return False


def open_obligation_iva(page: Page, mapping: dict, sat_ui: dict | None = None) -> bool:
    """Select 'IVA simplificado de confianza' on the administración de la declaración page to open the IVA form. Returns True if clicked."""
    sat_ui = sat_ui or DEFAULT_SAT_UI
    LOG.info("IVA: step — selecting 'IVA simplificado de confianza' on administración page")
    ok = _try_click(page, mapping, "_select_obligation_iva")
    if ok:
        LOG.info("IVA: clicked 'IVA simplificado de confianza' (mapping)")
        page.wait_for_timeout(_OBLIGATION_CLICK_WAIT_MS)
        return True
    iva_pat = _ui_pattern(sat_ui, "select_obligation_iva")
    for elem in page.get_by_text(iva_pat).all():
        try:
            if elem.is_visible() and elem.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").count() == 0:
                elem.click(timeout=2000)
                LOG.info("IVA: clicked 'IVA simplificado de confianza' (text)")
                page.wait_for_timeout(_OBLIGATION_CLICK_WAIT_MS)
                return True
        except Exception:
            continue
    LOG.warning("IVA: could not click 'IVA simplificado de confianza'")
    return False


def _set_dropdown_by_label(page: Page, label_substring: str, value: str, timeout_ms: int = 3000) -> bool:
    """Set a <select> that is associated with a label containing label_substring. value is option label (e.g. Sí/No)."""
    try:
        loc = page.get_by_label(re.compile(re.escape(label_substring), re.I))
        if loc.count() == 0:
            return False
        first = loc.first
        first.wait_for(state="visible", timeout=timeout_ms)
        tag = first.evaluate("el => el.tagName.toLowerCase()")
        if tag == "select":
            first.select_option(label=value)
            return True
    except Exception:
        pass
    return False


def _get_isr_ingresos_scope(page: Page) -> Page:
    """Return the page; ISR Ingresos form is a tab panel (id=tab457maincontainer1) on the main page, not an iframe."""
    return page


def _read_sat_total_ingresos_cobrados(page: Page, scope: Page, mapping: dict | None = None, sat_ui: dict | None = None) -> float:
    """Read the value from SAT form section 'Total de ingresos efectivamente cobrados' (textbox/field in Ingresos tab). Ignores hidden modal. Returns 0.0 if not found."""
    sat_ui = sat_ui or DEFAULT_SAT_UI
    sat_label = sat_ui.get("total_ingresos_cobrados") or "Total de ingresos efectivamente cobrados"
    total_cobrados_pat = re.compile(re.escape(sat_label), re.I)
    # If mapping has a selector for this field, try it first
    if mapping:
        for sel in mapping.get("_isr_ingresos_total_cobrados") or []:
            try:
                el = scope.locator(sel).first
                el.wait_for(state="visible", timeout=600)
                raw = el.get_attribute("value") or el.input_value() or el.inner_text()
                parsed = _parse_currency(raw)
                if parsed != 0.0 or (raw and re.search(r"[\d,]", str(raw))):
                    LOG.info("Phase 3: SAT value: section=%r (mapping %r), raw=%r, parsed=%.2f", sat_label, sel, raw, parsed)
                    return parsed
            except Exception:
                continue
    # Prefer scope to main tab content so we don't match the hidden modal title (avoid slow .count())
    try:
        scope.locator("#tab457maincontainer1").first.wait_for(state="attached", timeout=500)
        tab_scope = scope.locator("#tab457maincontainer1").first
    except Exception:
        tab_scope = scope
    try:
        label_el = None
        for el in tab_scope.get_by_text(total_cobrados_pat).all():
            try:
                if el.is_visible():
                    label_el = el
                    break
            except Exception:
                continue
        if label_el is None:
            for el in scope.get_by_text(total_cobrados_pat).all():
                try:
                    if el.is_visible():
                        label_el = el
                        break
                except Exception:
                    continue
        if label_el is None:
            raise RuntimeError("No visible label %r found" % (sat_label,))
        # Label's "for" attribute points to input id (SAT often uses this)
        try:
            for_id = label_el.get_attribute("for")
            if for_id and for_id.strip():
                inp = tab_scope.locator(f"#{re.escape(for_id.strip())}").first
                if inp.count() > 0:
                    inp.wait_for(state="visible", timeout=600)
                    raw = inp.get_attribute("value") or inp.input_value() or inp.inner_text()
                    parsed = _parse_currency(raw)
                    if parsed != 0.0 or (raw and re.search(r"[\d,]", str(raw))):
                        LOG.info("Phase 3: SAT value: section=%r (label for=), raw=%r, parsed=%.2f", sat_label, raw, parsed)
                        return parsed
        except Exception:
            pass
        # Input(s) after the label in DOM order (SAT form: label then input with value 66,264; sometimes 2nd or 3rd input)
        for input_index in range(1, 5):
            try:
                following_input = label_el.locator(f"xpath=following::input[{input_index}]")
                if following_input.count() > 0:
                    following_input.first.wait_for(state="visible", timeout=500)
                    raw = following_input.first.get_attribute("value") or following_input.first.input_value() or following_input.first.inner_text()
                    parsed = _parse_currency(raw)
                    if parsed > 0 or (raw and re.search(r"[\d,]", str(raw))):
                        LOG.info("Phase 3: SAT value: section=%r (following::input[%s]), raw=%r, parsed=%.2f", sat_label, input_index, raw, parsed)
                        return parsed
            except Exception:
                continue
        # First try whole row text (works when value is in span/div next to label)
        for _ in range(2):
            try:
                row = label_el.locator("xpath=ancestor::tr[1]").first
                row.wait_for(state="visible", timeout=600)
                row_text = row.inner_text()
                # Match number with optional commas: 66,264 or 66264
                m = re.search(r"[\d][\d,]*\.?\d*", row_text.replace(" ", ""))
                if m:
                    raw = m.group(0)
                    parsed = _parse_currency(raw)
                    if parsed > 0 or re.search(r"[\d]", raw):
                        LOG.info("Phase 3: SAT value: section=%r (from row text), raw=%r, parsed=%.2f", sat_label, raw, parsed)
                        return parsed
            except Exception:
                break
        # Try input in same row
        for xpath in [
            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*//input",
            "xpath=(ancestor::tr[1])//input",
            "xpath=following-sibling::*//input",
            "xpath=..//input",
        ]:
            try:
                inp = label_el.locator(xpath).first
                inp.wait_for(state="visible", timeout=500)
                raw = inp.get_attribute("value") or inp.inner_text()
                parsed = _parse_currency(raw)
                if parsed != 0.0 or (raw and re.search(r"[\d,]", str(raw))):
                    LOG.info("Phase 3: SAT value: section=%r, raw=%r, parsed=%.2f", sat_label, raw, parsed)
                    return parsed
            except Exception:
                continue
        # Sibling/div with number
        for xpath in [
            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*",
            "xpath=following-sibling::*",
        ]:
            try:
                val_el = label_el.locator(xpath).first
                val_el.wait_for(state="visible", timeout=500)
                raw = val_el.inner_text()
                parsed = _parse_currency(raw)
                if parsed != 0.0 or re.search(r"[\d,]", str(raw)):
                    LOG.info("Phase 3: SAT value: section=%r (sibling), raw=%r, parsed=%.2f", sat_label, raw, parsed)
                    return parsed
            except Exception:
                continue
        # Try get_by_label in full scope (control may not be inside tab container)
        for try_scope in (scope, tab_scope):
            try:
                control = try_scope.get_by_label(total_cobrados_pat).first
                control.wait_for(state="visible", timeout=500)
                raw = control.get_attribute("value")
                if not raw or (isinstance(raw, str) and not raw.strip()):
                    try:
                        raw = control.input_value()
                    except Exception:
                        raw = control.inner_text() or ""
                raw = raw or control.inner_text()
                parsed = _parse_currency(raw)
                LOG.info("Phase 3: SAT value: section=%r (get_by_label), raw=%r, parsed=%.2f", sat_label, raw, parsed)
                return parsed
            except Exception:
                continue
        # SAT tab 457: try any visible input in tab that has a numeric value (prefilled total is often the only one)
        try:
            for inp in tab_scope.locator("input[type='text'], input[type='number'], input:not([type])").all():
                try:
                    if not inp.is_visible():
                        continue
                    raw = inp.get_attribute("value") or inp.input_value() or ""
                    parsed = _parse_currency(raw)
                    if parsed > 0 and re.search(r"[\d,]", str(raw)):
                        LOG.info("Phase 3: SAT value: section=%r (tab input), raw=%r, parsed=%.2f", sat_label, raw, parsed)
                        return parsed
                except Exception:
                    continue
        except Exception:
            pass
        # Last resort: any visible input/span in same table or container that contains a number
        try:
            container = label_el.locator("xpath=ancestor::table[1] | ancestor::*[contains(@class,'tab')][1] | ancestor::*[@id][1]").first
            for sel in ["input[type='text']", "input[type='number']", "span", "[role='textbox']"]:
                for node in container.locator(sel).all():
                    try:
                        if not node.is_visible():
                            continue
                        raw = node.get_attribute("value") or node.inner_text()
                        parsed = _parse_currency(raw)
                        if parsed > 0 and re.search(r"[\d,]", str(raw)):
                            LOG.info("Phase 3: SAT value: section=%r (container), raw=%r, parsed=%.2f", sat_label, raw, parsed)
                            return parsed
                    except Exception:
                        continue
        except Exception:
            pass
        LOG.warning("Phase 3: SAT value: %r not found in form (all strategies failed); using 0.0 for comparison", sat_label)
        return 0.0
    except Exception as e:
        LOG.info("Phase 3: SAT value: section=%r — could not read (returning 0.0): %s", sat_label, e)
        return 0.0


def _click_capturar_iva_acreditable(page_or_scope) -> bool:
    """Click the CAPTURAR next to '*IVA acreditable del periodo' using ISR logic: iterate all CAPTURAR, find row that contains the label text, click that one. Works when form uses div rows (no tr)."""
    try:
        for cap in page_or_scope.locator("a, button").filter(has_text=re.compile(r"CAPTURAR", re.I)).all():
            try:
                row = cap.locator("xpath=ancestor::tr[1] | ancestor::div[contains(@class,'row')][1]").first
                if row.count() == 0:
                    continue
                row_text = (row.inner_text(timeout=500) or "").lower()
                if "iva acreditable del periodo" in row_text or "acreditable del periodo" in row_text:
                    if "acreditable por actividades" in row_text and "acreditable del periodo" not in row_text:
                        continue
                    cap.scroll_into_view_if_needed(timeout=500)
                    cap.click()
                    return True
            except Exception:
                continue
    except Exception:
        pass
    return False


def _click_capturar_next_to_label(page_or_scope: Page | Frame, label_substring: str, *, occurrence: int = 0) -> bool:
    """Find the label containing label_substring (use occurrence=1 for second match, e.g. *Ingresos a disminuir section), then click the CAPTURAR in the same row only (not the first CAPTURAR in a large container). Returns True if clicked."""
    try:
        loc = page_or_scope.get_by_text(re.compile(re.escape(label_substring), re.I))
        if loc.count() <= occurrence:
            return False
        label_el = loc.nth(occurrence)
        label_el.wait_for(state="visible", timeout=500)
    except Exception:
        return False
    for i in range(1, 10):
        try:
            container = label_el.locator(f"xpath=(ancestor::*)[{i}]")
            if container.count() == 0:
                break
            caps = container.first.locator("a, button").filter(has_text=re.compile(r"CAPTURAR", re.I))
            n = caps.count()
            if n == 0:
                continue
            # Prefer container that has exactly one CAPTURAR (same row as label); else we might hit a big div and get the wrong (first) CAPTURAR
            if n == 1:
                capturar = caps.first
            else:
                # Multiple CAPTURARs: pick the one in the same row as label_el (ISR: use tr or div.row)
                for j in range(n):
                    cap = caps.nth(j)
                    row = cap.locator("xpath=ancestor::tr[1] | ancestor::div[contains(@class,'row')][1]").first
                    if row.count() == 0:
                        continue
                    row_text = row.inner_text(timeout=200) or ""
                    # Same row as our label if it contains the label text; exclude question row "¿Tienes ingresos a disminuir?"
                    if label_substring.lower() in row_text.lower() and "¿Tienes" not in row_text:
                        capturar = cap
                        break
                else:
                    continue
            capturar.wait_for(state="visible", timeout=500)
            try:
                capturar.scroll_into_view_if_needed(timeout=250)
            except Exception:
                pass
            capturar.click()
            return True
        except Exception:
            continue
    return False


def _fill_input_next_to_label(
    page_or_scope: Page | Frame, page_for_wait: Page, label_substring: str, value_str: str, *, occurrence: int = 0
) -> bool:
    """Find the label containing label_substring, then the input in same row or following in DOM, and fill it (same technique as Importe in ISR). Returns True if filled."""
    def _values_match(got: str, expected: str) -> bool:
        if got == expected:
            return True
        # SAT often displays numbers with commas (e.g. 58,085); compare as numbers
        g = (got or "").replace(",", "").strip()
        e = (expected or "").replace(",", "").strip()
        if g == e:
            return True
        try:
            return float(g) == float(e)
        except ValueError:
            return False

    def _set_value(inp, val: str) -> bool:
        def _get_val():
            try:
                return (inp.input_value() or inp.get_attribute("value") or "").strip()
            except Exception:
                return ""
        try:
            inp.scroll_into_view_if_needed(timeout=_TEXTBOX_FILL_VISIBLE_MS)
            inp.click()
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS)
            inp.fill("")
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS)
            inp.fill(val)
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS * 2)
            if _values_match(_get_val(), val):
                return True
            inp.click()
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS)
            page_for_wait.keyboard.press("Control+a")
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS)
            page_for_wait.keyboard.type(val, delay=20)
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS * 2)
            if _values_match(_get_val(), val):
                return True
            inp.click()
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS)
            inp.press_sequentially(val, delay=20)
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS * 2)
            if _values_match(_get_val(), val):
                return True
            inp.evaluate("""el => { el.value = arguments[0]; el.dispatchEvent(new Event('input', { bubbles: true })); el.dispatchEvent(new Event('change', { bubbles: true })); }""", val)
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS)
            return _values_match(_get_val(), val)
        except Exception:
            return False

    # Strategy 0: get_by_label with exact=False (same as ISR _try_fill when selector is "label=...")
    try:
        control = page_or_scope.get_by_label(label_substring, exact=False).first
        control.wait_for(state="visible", timeout=_TEXTBOX_FILL_VISIBLE_MS)
        tag = control.evaluate("el => (el && el.tagName) ? el.tagName.toLowerCase() : ''")
        if tag in ("input", "textarea", ""):
            if not control.get_attribute("disabled") and _set_value(control, value_str):
                return True
    except Exception:
        pass

    try:
        loc = page_or_scope.get_by_text(re.compile(re.escape(label_substring), re.I))
        if loc.count() <= occurrence:
            return False
        label_el = loc.nth(occurrence)
        label_el.wait_for(state="visible", timeout=_TEXTBOX_FILL_VISIBLE_MS)
    except Exception:
        return False
    # Only skip when label is inside a modal/dialog and we are searching the full page (avoid matching elements in a popup when we want main page). When scope is the dialog, we want to match labels inside it.
    if page_or_scope is page_for_wait and label_el.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").count() > 0:
        return False

    # Strategy 1: Label "for" attribute → input id (same as ISR: label for="xxx" → #xxx)
    try:
        for_id = label_el.get_attribute("for")
        if not for_id:
            try:
                anc_label = label_el.locator("xpath=ancestor::label[1]").first
                if anc_label.count() > 0:
                    for_id = anc_label.get_attribute("for")
            except Exception:
                pass
        if for_id and str(for_id).strip():
            inp = page_or_scope.locator(f"#{re.escape(str(for_id).strip())}").first
            if inp.count() > 0:
                inp.wait_for(state="visible", timeout=_TEXTBOX_FILL_VISIBLE_MS)
                if not inp.get_attribute("disabled") and _set_value(inp, value_str):
                    return True
    except Exception:
        pass

    # Strategy 2: Same-row input (SAT ISR uses <tr>; IVA may use tr or div.row)
    try:
        row_inputs = label_el.locator("xpath=(ancestor::tr[1])//input")
        n = row_inputs.count()
        for i in range(n):
            try:
                inp = row_inputs.nth(i)
                inp.wait_for(state="visible", timeout=_TEXTBOX_FILL_VISIBLE_MS)
                if inp.get_attribute("disabled"):
                    continue
                if _set_value(inp, value_str):
                    return True
            except Exception:
                continue
    except Exception:
        pass

    # Strategy 3: First ancestor that contains an input (works for div-based rows: label and input in same container)
    try:
        container = label_el.locator("xpath=ancestor::*[.//input][1]").first
        if container.count() > 0:
            for inp in container.locator("input").all():
                try:
                    if not inp.is_visible() or inp.get_attribute("disabled"):
                        continue
                    if _set_value(inp, value_str):
                        return True
                except Exception:
                    continue
    except Exception:
        pass

    for input_index in range(1, 6):
        try:
            following_input = label_el.locator(f"xpath=following::input[{input_index}]")
            if following_input.count() > 0:
                inp = following_input.first
                inp.wait_for(state="visible", timeout=_TEXTBOX_FILL_VISIBLE_MS)
                if inp.get_attribute("disabled"):
                    continue
                if _set_value(inp, value_str):
                    return True
        except Exception:
            continue
    for xpath in [
        "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*//input",
        "xpath=(ancestor::tr[1])//input",
        "xpath=following-sibling::*//input",
    ]:
        try:
            inp = label_el.locator(xpath).first
            if inp.count() == 0:
                continue
            inp.wait_for(state="visible", timeout=_TEXTBOX_FILL_VISIBLE_MS)
            if inp.get_attribute("disabled"):
                continue
            if _set_value(inp, value_str):
                return True
        except Exception:
            continue
    try:
        control = page_or_scope.get_by_label(re.compile(re.escape(label_substring), re.I)).first
        control.wait_for(state="visible", timeout=_TEXTBOX_FILL_VISIBLE_MS)
        tag = control.evaluate("el => (el && el.tagName) ? el.tagName.toLowerCase() : ''")
        if tag == "input" or tag == "":
            if _set_value(control, value_str):
                return True
    except Exception:
        pass
    return False


def _click_capturar_ingresos_a_disminuir(page: Page, scope: Page | Frame) -> bool:
    """Click the CAPTURAR for '*Ingresos a disminuir' only (the second CAPTURAR on the form; the first is Descuentos). Order: Sí → section appears → this CAPTURAR → popup AGREGAR → Concepto → Importe (int) → GUARDAR → CERRAR."""
    try:
        scope.locator("#tab457maincontainer1").first.wait_for(state="attached", timeout=500)
        tab_scope = scope.locator("#tab457maincontainer1").first
    except Exception:
        tab_scope = scope
    # Strategy 1: CAPTURAR that immediately follows the dropdown "¿Tienes ingresos a disminuir?" (id 457select48) in DOM — that is the *Ingresos a disminuir row's CAPTURAR, not Descuentos
    try:
        for sel in ["#457select48", "select[id='457select48']", "[id='457select48']"]:
            try:
                dd = tab_scope.locator(sel).first
                dd.wait_for(state="visible", timeout=800)
                cap = dd.locator("xpath=following::a[contains(., 'CAPTURAR')]").first
                cap.wait_for(state="visible", timeout=1200)
                try:
                    cap.scroll_into_view_if_needed(timeout=400)
                except Exception:
                    pass
                cap.click()
                return True
            except Exception:
                continue
    except Exception:
        pass
    # Strategy 2: of all CAPTURARs in tab, click the one whose row contains "*Ingresos a disminuir" / "Ingresos a disminuir" but NOT "¿Tienes ingresos a disminuir" (so we get the section row, not the question)
    try:
        for cap in tab_scope.locator("a, button").filter(has_text=re.compile(r"CAPTURAR", re.I)).all():
            try:
                row = cap.locator("xpath=ancestor::tr[1] | ancestor::div[contains(@class,'row')][1]").first
                if row.count() == 0:
                    continue
                row_text = row.inner_text(timeout=500) or ""
                if "Ingresos a disminuir" in row_text and "¿Tienes ingresos a disminuir" not in row_text:
                    cap.scroll_into_view_if_needed(timeout=500)
                    cap.click()
                    return True
            except Exception:
                continue
    except Exception:
        pass
    # Strategy 3: second label "Ingresos a disminuir" (section input) → its row → CAPTURAR in that row
    try:
        inp = tab_scope.get_by_label(re.compile(r"Ingresos a disminuir", re.I))
        if inp.count() >= 2:
            row = inp.nth(1).locator("xpath=ancestor::tr[1] | ancestor::div[contains(@class,'row')][1]").first
            cap = row.locator("a, button").filter(has_text=re.compile(r"CAPTURAR", re.I)).first
            cap.wait_for(state="visible", timeout=800)
            try:
                cap.scroll_into_view_if_needed(timeout=500)
            except Exception:
                pass
            cap.click()
            return True
    except Exception:
        pass
    # Strategy 4: label-based with occurrence=1 (uses same-row logic to avoid clicking Descuentos CAPTURAR)
    if _click_capturar_next_to_label(tab_scope, "Ingresos a disminuir", occurrence=1):
        return True
    return False


def _click_capturar_ingresos_adicionales(page: Page, scope: Page | Frame) -> bool:
    """Click the CAPTURAR for '*Ingresos adicionales' only (the CAPTURAR to the right of the '*Ingresos adicionales' label).
    Order: dropdown ¿Tienes ingresos adicionales? = Sí → *Ingresos adicionales CAPTURAR → popup AGREGAR → Concepto → Importe → GUARDAR → CERRAR."""
    try:
        scope.locator("#tab457maincontainer1").first.wait_for(state="attached", timeout=500)
        tab_scope = scope.locator("#tab457maincontainer1").first
    except Exception:
        tab_scope = scope

    # Strategy 1: label "*Ingresos adicionales" → its row → CAPTURAR in that row
    try:
        loc = tab_scope.get_by_text(re.compile(r"\*?\s*Ingresos adicionales", re.I))
        if loc.count() > 0:
            for idx in range(loc.count()):
                try:
                    label_el = loc.nth(idx)
                    label_el.wait_for(state="visible", timeout=500)
                    row = label_el.locator("xpath=ancestor::tr[1] | ancestor::div[contains(@class,'row')][1]").first
                    if row.count() == 0:
                        continue
                    cap = row.locator("a, button").filter(has_text=re.compile(r"CAPTURAR", re.I)).first
                    cap.wait_for(state="visible", timeout=800)
                    try:
                        cap.scroll_into_view_if_needed(timeout=400)
                    except Exception:
                        pass
                    cap.click()
                    return True
                except Exception:
                    continue
    except Exception:
        pass

    # Strategy 2: of all CAPTURARs in tab, click the one whose row contains "Ingresos adicionales" but not the question row "¿Tienes ingresos adicionales"
    try:
        for cap in tab_scope.locator("a, button").filter(has_text=re.compile(r"CAPTURAR", re.I)).all():
            try:
                row = cap.locator("xpath=ancestor::tr[1] | ancestor::div[contains(@class,'row')][1]").first
                if row.count() == 0:
                    continue
                row_text = row.inner_text(timeout=500) or ""
                if "Ingresos adicionales" in row_text and "¿Tienes ingresos adicionales" not in row_text:
                    cap.scroll_into_view_if_needed(timeout=500)
                    cap.click()
                    return True
            except Exception:
                continue
    except Exception:
        pass

    # Strategy 3: fallback to generic label-based helper on tab scope
    if _click_capturar_next_to_label(tab_scope, "Ingresos adicionales", occurrence=0):
        return True
    return False


def _click_capturar_total_percibidos(page: Page, scope: Page | Frame) -> bool:
    """Click the CAPTURAR to the right of 'Total de ingresos percibidos por la actividad' only (the third CAPTURAR; do not click Ingresos a disminuir CAPTURAR)."""
    try:
        scope.locator("#tab457maincontainer1").first.wait_for(state="attached", timeout=500)
        tab_scope = scope.locator("#tab457maincontainer1").first
    except Exception:
        tab_scope = scope
    try:
        label_el = tab_scope.get_by_text("Total de ingresos percibidos por la actividad", exact=False).first
        label_el.wait_for(state="visible", timeout=800)
        for i in range(1, 10):
            try:
                container = label_el.locator(f"xpath=(ancestor::*)[{i}]")
                if container.count() == 0:
                    break
                caps = container.first.locator("a, button").filter(has_text=re.compile(r"CAPTURAR", re.I))
                n = caps.count()
                if n == 0:
                    continue
                if n == 1:
                    capturar = caps.first
                else:
                    for j in range(n):
                        cap = caps.nth(j)
                        cap_row = cap.locator("xpath=ancestor::tr[1]").first
                        if cap_row.count() == 0:
                            continue
                        if "Total de ingresos percibidos" in (cap_row.inner_text(timeout=200) or ""):
                            capturar = cap
                            break
                    else:
                        continue
                capturar.wait_for(state="visible", timeout=500)
                try:
                    capturar.scroll_into_view_if_needed(timeout=250)
                except Exception:
                    pass
                capturar.click()
                return True
            except Exception:
                continue
    except Exception:
        pass
    return False


def _click_ver_detalle_isr_retenido(page_or_scope: Page | Frame, sat_ui: dict | None = None) -> bool:
    """Same logic as Phase 3 CAPTURAR: of all VER DETALLE on page, click the one whose row contains the configured row label. Returns True if clicked."""
    sat_ui = sat_ui or DEFAULT_SAT_UI
    required_in_row = sat_ui.get("isr_retenido_row_label") or "ISR retenido por personas morales"
    ver_text = sat_ui.get("ver_detalle_button") or "VER DETALLE"
    ver_pat = re.compile(re.escape(ver_text), re.I)
    try:
        for ver in page_or_scope.locator("a, button, [role='button']").filter(has_text=ver_pat).all():
            try:
                row = ver.locator("xpath=ancestor::tr[1] | ancestor::div[contains(@class,'row')][1]").first
                if row.count() == 0:
                    continue
                row_text = row.inner_text(timeout=600) or ""
                if required_in_row not in row_text:
                    continue
                if not ver.is_visible():
                    continue
                ver.scroll_into_view_if_needed(timeout=500)
                page_for_click = page_or_scope.page if hasattr(page_or_scope, "page") else page_or_scope
                if hasattr(page_for_click, "wait_for_timeout"):
                    page_for_click.wait_for_timeout(200)
                ver.click(timeout=2500)
                return True
            except Exception:
                continue
    except Exception:
        pass
    return False


def _click_ver_detalle_next_to_label(page_or_scope: Page | Frame, label_substring: str, sat_ui: dict | None = None) -> bool:
    """Find the label containing label_substring (prefer visible one when multiple), then click the 'VER DETALLE' in the same row. Returns True if clicked."""
    sat_ui = sat_ui or DEFAULT_SAT_UI
    ver_text = sat_ui.get("ver_detalle_button") or "VER DETALLE"
    ver_pat = re.compile(re.escape(ver_text), re.I)
    try:
        loc = page_or_scope.get_by_text(re.compile(re.escape(label_substring), re.I))
        if loc.count() == 0:
            return False
        label_el = None
        for idx in range(loc.count()):
            try:
                el = loc.nth(idx)
                el.wait_for(state="attached", timeout=500)
                if el.is_visible():
                    label_el = el
                    break
            except Exception:
                continue
        if label_el is None:
            label_el = loc.first
            label_el.wait_for(state="visible", timeout=1500)
    except Exception:
        return False
    for i in range(1, 10):
        try:
            container = label_el.locator(f"xpath=(ancestor::*)[{i}]")
            if container.count() == 0:
                break
            ver_btns = container.first.locator("a, button, [role='button']").filter(has_text=ver_pat)
            if ver_btns.count() == 0:
                continue
            ver_btn = ver_btns.first
            ver_btn.wait_for(state="visible", timeout=800)
            ver_btn.scroll_into_view_if_needed(timeout=400)
            page_for_click = page_or_scope.page if hasattr(page_or_scope, "page") else page_or_scope
            if hasattr(page_for_click, "wait_for_timeout"):
                page_for_click.wait_for_timeout(200)
            ver_btn.click(timeout=2500)
            return True
        except Exception:
            continue
    return False


def _set_dropdown_by_label_scope(scope: Page | Frame, page_for_click: Page, label_substring: str, value: str, timeout_ms: int = 5000) -> bool:
    """Set a <select> in scope by label containing label_substring. value is option label (e.g. Sí/No)."""
    try:
        loc = scope.get_by_label(re.compile(re.escape(label_substring), re.I))
        if loc.count() == 0:
            return False
        first = loc.first
        first.wait_for(state="visible", timeout=timeout_ms)
        tag = first.evaluate("el => el.tagName.toLowerCase()")
        if tag == "select":
            first.select_option(label=value)
            return True
    except Exception:
        pass
    return False


def fill_isr_ingresos_form(page: Page, mapping: dict, data: dict, sat_ui: dict | None = None) -> None:
    """Fill ISR simplificado Ingresos form. Phase 3 (ISR simplificado de confianza. Personas físicas):
    - *¿Los ingresos fueron obtenidos a través de copropiedad? → always "No"
    - Total de ingresos efectivamente cobrados → skip (prefilled, no action)
    - Descuentos, devoluciones y bonificaciones → CAPTURAR → popup: set *Descuentos...integrantes por copropiedad to 0 → CERRAR
    - ¿Tienes ingresos a disminuir? → if (SAT Total - Excel Total) > 1 then Sí + CAPTURAR popup (AGREGAR → Concepto → Importe → GUARDAR → CERRAR); else No
    - ¿Tienes ingresos adicionales? → if (Excel - SAT) > 1 then Sí + CAPTURAR popup; else No
    - Total percibidos CAPTURAR popup.
    Phase 4 (after Ingresos completed): GUARDAR → Determinación tab → VER DETALLE (ISR retenido por personas morales) → popup fill "ISR retenido no acreditable" from Excel (label "ISR retenido") → CERRAR → GUARDAR.
    Phase 5 (after Phase 4): Pago tab → *¿Tienes compensaciones por aplicar? → No, *¿Tienes estímulos por aplicar? → No → GUARDAR → wait for load."""
    LOG.info("Phase 3: Filling Ingresos form...")
    label_map = data.get("label_map") or {}
    # Always "No" per requirements for *¿Los ingresos fueron obtenidos a través de copropiedad?
    copropiedad = "No"
    # Descuentos popup: value for *Descuentos, devoluciones y bonificaciones de integrantes por copropiedad (0 if not in Excel)
    descuentos_copropiedad = label_map.get("Descuentos devoluciones y bonificaciones de integrantes por copropiedad")
    if descuentos_copropiedad is None:
        descuentos_copropiedad = 0
    # Excel value for "Total de ingresos" (Impuestos tab col D→E); compared with SAT "Total de ingresos efectivamente cobrados"
    # Primary: row with "Base gravable del pago provisional" (col D = label, col E = value); fallbacks for other templates
    excel_label_1 = "Base gravable del pago provisional"
    excel_label_2 = "Ingresos cobrados y amparados por factura del mes"
    excel_label_3 = "Total de ingresos acumulados"
    excel_label_4 = "Ingresos nominales facturados"
    # Use first label that has a value (including 0.0); 'or' would treat 0.0 as missing
    _v1, _v2, _v3, _v4 = label_map.get(excel_label_1), label_map.get(excel_label_2), label_map.get(excel_label_3), label_map.get(excel_label_4)
    used_key = (
        excel_label_1 if _v1 is not None
        else (excel_label_2 if _v2 is not None
              else (excel_label_3 if _v3 is not None
                    else (excel_label_4 if _v4 is not None else None)))
    )
    excel_total_cobrados_raw = _v1 if _v1 is not None else (_v2 if _v2 is not None else (_v3 if _v3 is not None else _v4))
    excel_total_cobrados = _parse_currency(excel_total_cobrados_raw)
    workbook_path = data.get("workbook_path") or "(workbook path not set)"
    LOG.info(
        "Phase 3: Excel: workbook=%s | Sheet=Impuestos (layout: D/E or E/F). "
        "Row used: label=%r, value=%s → parsed=%.2f",
        workbook_path, used_key, excel_total_cobrados_raw, excel_total_cobrados,
    )
    LOG.info(
        "Phase 3: Excel values found: %r=%s, %r=%s, %r=%s, %r=%s",
        excel_label_1, _v1, excel_label_2, _v2, excel_label_3, _v3, excel_label_4, _v4,
    )
    importe_total = _v1 if _v1 is not None else (_v2 if _v2 is not None else (_v3 if _v3 is not None else _v4))
    if importe_total is None:
        importe_total = 0
    try:
        n = float(importe_total)
        importe_str = f"{n:,.2f}".replace(",", "")
    except (TypeError, ValueError):
        importe_str = str(importe_total)
    concepto_label = data.get("isr_ingresos_concepto") or "Actividad empresarial"

    page.wait_for_timeout(150)
    LOG.info("")
    LOG.info("===== Phase 3: sección 1 - ¿Los ingresos fueron obtenidos a través de copropiedad? =====")
    # Wait for Ingresos form by label text (minimal wait for Phase 3)
    scope = _get_isr_ingresos_scope(page)
    try:
        scope.get_by_text("copropiedad", exact=False).first.wait_for(state="visible", timeout=400)
    except Exception:
        try:
            scope.get_by_text("ingresos fueron obtenidos", exact=False).first.wait_for(state="visible", timeout=400)
        except Exception:
            LOG.warning("Phase 3: label 'copropiedad' not visible")
    # 1. ¿Los ingresos fueron obtenidos a través de copropiedad? — use unique label (avoid matching "integrantes por copropiedad" in Descuentos)
    _si = copropiedad.strip().lower() in ("sí", "si", "yes")
    si_no_label = "Sí" if _si else "No"
    copropiedad_ok = _fill_select_next_to_label(
        scope,
        page,
        "ingresos fueron obtenidos a través de copropiedad",
        si_no_label,
        mapping=mapping,
        initial_dropdown_key="_isr_ingresos_copropiedad",
    )
    if not copropiedad_ok:
        copropiedad_ok = _set_dropdown_by_label_scope(scope, page, "ingresos fueron obtenidos a través de copropiedad", si_no_label, timeout_ms=1500)
    if copropiedad_ok:
        LOG.info("Phase 3: copropiedad = %s (dropdown)", si_no_label)
    else:
        LOG.warning("Phase 3: could not set copropiedad dropdown")
    page.wait_for_timeout(PHASE3_SECTION_GAP_MS)
    # 2. Total de ingresos efectivamente cobrados — no need to fill or do anything (prefilled, skip)
    LOG.info("")
    LOG.info("===== Phase 3: sección 3 - Descuentos, devoluciones y bonificaciones =====")
    # 3. Descuentos, devoluciones y bonificaciones: press CAPTURAR → popup: add "0" on *Descuentos, devoluciones y bonificaciones de integrantes por copropiedad → CERRAR
    try:
        capturar_clicked = _try_click(page, mapping, "_isr_ingresos_capturar_descuentos")
        if not capturar_clicked:
            capturar_clicked = _click_capturar_next_to_label(scope, "Descuentos, devoluciones y bonificaciones")
        if not capturar_clicked:
            capturar_clicked = _click_capturar_next_to_label(scope, "Descuentos")
        if capturar_clicked:
            page.wait_for_timeout(80)
            descuentos_value = str(descuentos_copropiedad)
            filled = False
            try:
                page.get_by_text("Devoluciones, descuentos y bonificaciones facturadas", exact=False).first.wait_for(state="visible", timeout=400)
                LOG.info("Phase 3: Descuentos popup (Devoluciones, descuentos y bonificaciones facturadas) visible")
            except Exception:
                pass
            page.wait_for_timeout(80)
            dialog = None
            for try_dialog in [
                lambda: page.get_by_role("dialog").first,
                lambda: page.locator(".modal, [role='dialog']").first,
                lambda: page.get_by_text("Devoluciones, descuentos y bonificaciones facturadas", exact=False).first.locator("xpath=ancestor::*[contains(@class,'modal') or contains(@class,'dialog') or @role='dialog'][1]"),
            ]:
                try:
                    d = try_dialog()
                    d.wait_for(state="visible", timeout=400)
                    dialog = d
                    break
                except Exception:
                    continue
            if dialog is None:
                dialog = page
            # Try multiple label texts (SAT wording may vary); shorter timeouts for faster fail and next section
            for label_pattern in [
                r"integrantes por copropiedad",
                r"Descuentos.*integrantes por copropiedad",
                r"Descuentos.*de integrantes por copropiedad",
                r"devoluciones.*copropiedad",
                r"copropiedad",
            ]:
                if filled:
                    break
                try:
                    label_el = dialog.get_by_text(re.compile(label_pattern, re.I)).first
                    label_el.wait_for(state="visible", timeout=250)
                    for xpath in [
                        "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*//input",
                        "xpath=(ancestor::tr[1])//input",
                        "xpath=following-sibling::*//input",
                        "xpath=..//input",
                    ]:
                        try:
                            inp = label_el.locator(xpath).first
                            inp.wait_for(state="visible", timeout=150)
                            if inp.get_attribute("disabled"):
                                continue
                            inp.click()
                            inp.clear()
                            inp.fill(descuentos_value)
                            filled = True
                            LOG.info("Phase 3: Descuentos popup filled (label=%r, xpath)", label_pattern)
                            break
                        except Exception:
                            continue
                    if filled:
                        break
                    # following::input[n] like SAT value read elsewhere
                    for input_index in [1, 2]:
                        try:
                            following_input = label_el.locator(f"xpath=following::input[{input_index}]")
                            if following_input.count() > 0:
                                inp = following_input.first
                                inp.wait_for(state="visible", timeout=150)
                                if inp.get_attribute("disabled"):
                                    continue
                                inp.click()
                                inp.clear()
                                inp.fill(descuentos_value)
                                filled = True
                                LOG.info("Phase 3: Descuentos popup filled (following::input[%s])", input_index)
                                break
                        except Exception:
                            continue
                except Exception:
                    continue
            if not filled:
                try:
                    inp = dialog.get_by_label(re.compile(r"Descuentos.*integrantes por copropiedad", re.I)).first
                    inp.wait_for(state="visible", timeout=250)
                    inp.click()
                    inp.clear()
                    inp.fill(descuentos_value)
                    filled = True
                except Exception:
                    pass
            if not filled:
                try:
                    modal_inputs = dialog.locator("input[type='text'], input[type='number'], input:not([type])")
                    n = min(modal_inputs.count(), 12)
                    for idx in range(n):
                        inp = modal_inputs.nth(idx)
                        try:
                            if inp.get_attribute("disabled"):
                                continue
                            inp.wait_for(state="visible", timeout=150)
                            row = inp.locator("xpath=ancestor::tr[1] | ancestor::div[contains(@class,'row')][1]")
                            if row.count() > 0:
                                row_text = row.first.inner_text() or ""
                                if re.search(r"integrantes por copropiedad|copropiedad|descuentos.*integrantes", row_text, re.I):
                                    inp.click()
                                    inp.clear()
                                    inp.fill(descuentos_value)
                                    filled = True
                                    LOG.info("Phase 3: Descuentos popup filled (modal_inputs by row text)")
                                    break
                        except Exception:
                            continue
                except Exception:
                    pass
            if not filled:
                LOG.warning("Phase 3: could not find Descuentos popup textbox (*Descuentos...integrantes por copropiedad)")
            page.wait_for_timeout(60)
            page.get_by_role("button", name=re.compile(r"CERRAR", re.I)).first.click(timeout=300)
            LOG.info("Phase 3: Descuentos popup filled and closed")
        else:
            LOG.warning("Phase 3: could not click Descuentos CAPTURAR link")
        page.wait_for_timeout(PHASE3_SECTION_GAP_MS)
    except Exception as e:
        LOG.warning("Phase 3: Descuentos CAPTURAR/popup failed: %s", e)
    LOG.info("")
    LOG.info("===== Phase 3: sección 4 - ¿Tienes ingresos a disminuir? / *Ingresos a disminuir =====")
    # 4. ¿Tienes ingresos a disminuir? — compare SAT "Total de ingresos efectivamente cobrados" vs Excel (Total de ingresos acumulados); if SAT - Excel > 1 → Sí + CAPTURAR popup
    sat_total_cobrados = _read_sat_total_ingresos_cobrados(page, scope, mapping, sat_ui)
    sat_total_cobrados = float(sat_total_cobrados) if sat_total_cobrados is not None else 0.0
    excel_total_cobrados = float(excel_total_cobrados) if excel_total_cobrados is not None else 0.0
    LOG.info("Phase 3: Comparison for *¿Tienes ingresos a disminuir? / *¿Tienes ingresos adicionales?: SAT=%.2f, Excel=%.2f", sat_total_cobrados, excel_total_cobrados)
    diferencia = sat_total_cobrados - excel_total_cobrados
    need_ingresos_a_disminuir = diferencia > 1
    si_no_disminuir_lbl = "Sí" if need_ingresos_a_disminuir else "No"
    page.wait_for_timeout(100)
    if _fill_select_next_to_label(
        scope,
        page,
        "ingresos a disminuir",
        si_no_disminuir_lbl,
        mapping=mapping,
        initial_dropdown_key="_isr_ingresos_disminuir",
    ):
        LOG.info("Phase 3: ingresos a disminuir = %s (SAT=%.2f Excel=%.2f diff=%.2f)", si_no_disminuir_lbl, sat_total_cobrados, excel_total_cobrados, diferencia)
    else:
        LOG.warning("Phase 3: could not set ingresos a disminuir dropdown")
    page.wait_for_timeout(100)
    if need_ingresos_a_disminuir:
        # *Ingresos a disminuir (section below dropdown) appears: wait for it, then CAPTURAR → popup AGREGAR → Concepto → Importe (int, no decimals) → GUARDAR → CERRAR
        page.wait_for_timeout(100)
        try:
            scope.locator("#tab457maincontainer1").first.wait_for(state="attached", timeout=400)
            tab_scope = scope.locator("#tab457maincontainer1").first
        except Exception:
            tab_scope = scope
        try:
            tab_scope.get_by_text(re.compile(r"Ingresos a disminuir", re.I)).nth(1).wait_for(state="visible", timeout=500)
        except Exception:
            pass
        page.wait_for_timeout(80)
        try:
            capturar_clicked = _click_capturar_ingresos_a_disminuir(page, scope)
            if capturar_clicked:
                LOG.info("Phase 3: Ingresos a disminuir CAPTURAR pressed")
                page.wait_for_timeout(80)
                try:
                    page.get_by_role("button", name=re.compile(r"AGREGAR", re.I)).first.click(timeout=600)
                    page.wait_for_timeout(50)
                except Exception as e_ag:
                    LOG.warning("Phase 3: could not click AGREGAR in Ingresos a disminuir popup: %s", e_ag)
                # Importe = difference with no decimals (e.g. 8178.81 → 8178)
                importe_val = int(round(diferencia))
                importe_str = str(importe_val)
                # Scope to the "Ingresos a disminuir" popup: modal has title + Concepto + Importe + GUARDAR; use .last for div so we get the modal, not main form
                dialog = None
                for use_last, dialog_loc, to_ms in [
                    (False, page.get_by_role("dialog"), 800),
                    (False, page.locator("[role='dialog']"), 500),
                    (True, page.locator("div").filter(has_text=re.compile(r"Ingresos a disminuir", re.I)).filter(has_text="Concepto").filter(has_text="Importe").filter(has_text="GUARDAR"), 400),
                    (True, page.locator("div").filter(has_text=re.compile(r"Ingresos a disminuir", re.I)).filter(has=page.locator("select")), 350),
                ]:
                    try:
                        d = dialog_loc.last if use_last else dialog_loc.first
                        d.wait_for(state="visible", timeout=to_ms)
                        dialog = d
                        break
                    except Exception:
                        continue
                if dialog is None:
                    dialog = page
                # Click AGREGAR inside the popup so Concepto/Importe row is ready
                if dialog != page:
                    try:
                        dialog.get_by_role("button", name=re.compile(r"AGREGAR", re.I)).first.click(timeout=400)
                        page.wait_for_timeout(50)
                    except Exception:
                        pass
                # Option text variants (SAT may use accents)
                concepto_option_labels = [
                    "Ingresos facturados pendientes de cancelacion con aceptacion del receptor",
                    "Ingresos facturados pendientes de cancelación con aceptación del receptor",
                ]
                # Strategy 1: *Concepto is label, dropdown on the right; then select option
                concepto_ok = False
                for opt_label in concepto_option_labels:
                    if concepto_ok:
                        break
                    try:
                        loc_concepto = dialog.get_by_text(re.compile(r"\*?\s*Concepto", re.I))
                        if loc_concepto.count() == 0:
                            loc_concepto = dialog.get_by_text("Concepto", exact=False)
                        label_concepto = loc_concepto.first
                        label_concepto.wait_for(state="visible", timeout=300)
                        for xpath in [
                            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*//select",
                            "xpath=(ancestor::tr[1])//select",
                            "xpath=following-sibling::*//select",
                            "xpath=..//select",
                        ]:
                            try:
                                concepto_dd = label_concepto.locator(xpath).first
                                concepto_dd.wait_for(state="visible", timeout=200)
                                concepto_dd.select_option(label=opt_label, timeout=2000)
                                concepto_ok = True
                                page.wait_for_timeout(30)
                                break
                            except Exception:
                                continue
                    except Exception:
                        continue
                if not concepto_ok:
                    try:
                        sel = dialog.locator("select").first
                        sel.wait_for(state="visible", timeout=300)
                        for opt_label in concepto_option_labels:
                            try:
                                sel.select_option(label=opt_label, timeout=2000)
                                concepto_ok = True
                                page.wait_for_timeout(30)
                                break
                            except Exception:
                                continue
                    except Exception as e_c:
                        LOG.warning("Phase 3: Concepto in popup: %s", e_c)
                if concepto_ok:
                    LOG.info("Phase 3: Ingresos a disminuir popup: selected *Concepto (Ingresos facturados pendientes de cancelación con aceptación del receptor)")
                # Importe: same row as Concepto dropdown (form row), then GUARDAR → CERRAR
                importe_ok = False
                # Strategy A: input in the same row as the Concepto select we just used (most reliable)
                try:
                    sel_first = dialog.locator("select").first
                    sel_first.wait_for(state="visible", timeout=200)
                    for row_xpath in ["xpath=ancestor::tr[1]", "xpath=ancestor::*[.//input][1]"]:
                        try:
                            row = sel_first.locator(row_xpath)
                            if row.count() == 0:
                                continue
                            inp = row.locator("input[type='text'], input[type='number'], input:not([type])").first
                            inp.wait_for(state="visible", timeout=180)
                            if inp.get_attribute("disabled") or inp.get_attribute("readonly"):
                                continue
                            inp.click()
                            inp.fill(importe_str)
                            importe_ok = True
                            page.wait_for_timeout(25)
                            break
                        except Exception:
                            continue
                except Exception:
                    pass
                if not importe_ok:
                    try:
                        label_importe = dialog.get_by_text("Importe", exact=False).first
                        label_importe.wait_for(state="visible", timeout=300)
                        for xpath in [
                            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*//input[not(@disabled)]",
                            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*//input",
                            "xpath=(ancestor::tr[1])//input[not(@disabled)]",
                            "xpath=(ancestor::tr[1])//input",
                            "xpath=following-sibling::*//input",
                        ]:
                            try:
                                importe_inp = label_importe.locator(xpath).first
                                importe_inp.wait_for(state="visible", timeout=180)
                                if importe_inp.get_attribute("disabled") or importe_inp.get_attribute("readonly"):
                                    continue
                                importe_inp.click()
                                importe_inp.fill(importe_str)
                                importe_ok = True
                                page.wait_for_timeout(25)
                                break
                            except Exception:
                                continue
                    except Exception as e_i:
                        LOG.warning("Phase 3: Importe in popup: %s", e_i)
                if not importe_ok:
                    try:
                        for inp in dialog.locator("input[type='text'], input[type='number'], input:not([type])").all():
                            try:
                                if inp.get_attribute("disabled") or inp.get_attribute("readonly"):
                                    continue
                                inp.wait_for(state="visible", timeout=120)
                                inp.fill(importe_str)
                                importe_ok = True
                                break
                            except Exception:
                                continue
                    except Exception:
                        pass
                if importe_ok:
                    LOG.info("Phase 3: Ingresos a disminuir popup: filled Importe=%s", importe_str)
                if concepto_ok and importe_ok:
                    try:
                        btn_scope = dialog if dialog != page else page
                        print(f"{_debug_ts()} Phase 3: Ingresos a disminuir popup flags: concepto_ok={concepto_ok}, importe_ok={importe_ok}, importe_str={importe_str}")
                        print(f"{_debug_ts()} Phase 3: Ingresos a disminuir: clicking GUARDAR")
                        btn_scope.get_by_role("button", name=re.compile(r"GUARDAR", re.I)).first.click(timeout=600)
                        page.wait_for_timeout(50)
                        # Small confirmation popup ("Captura la información requerida") with ACEPTAR button
                        try:
                            confirm_btn = page.get_by_role("button", name=re.compile(r"ACEPTAR", re.I)).first
                            confirm_btn.wait_for(state="visible", timeout=500)
                            print(f"{_debug_ts()} Phase 3: Ingresos a disminuir: clicking ACEPTAR")
                            confirm_btn.click()
                            page.wait_for_timeout(60)
                        except Exception:
                            pass
                        print(f"{_debug_ts()} Phase 3: Ingresos a disminuir: clicking CERRAR")
                        btn_scope.get_by_role("button", name=re.compile(r"CERRAR", re.I)).first.click(timeout=600)
                        LOG.info("Phase 3: Ingresos a disminuir popup filled (importe=%s), closed", importe_str)
                        page.wait_for_timeout(50)
                    except Exception as e_btn:
                        LOG.warning("Phase 3: GUARDAR/CERRAR in Ingresos a disminuir popup: %s", e_btn)
                else:
                    LOG.warning("Phase 3: Ingresos a disminuir popup not filled (concepto=%s, importe=%s); not clicking GUARDAR", concepto_ok, importe_ok)
            else:
                LOG.warning("Phase 3: could not click Ingresos a disminuir CAPTURAR")
        except Exception as e:
            LOG.warning("Phase 3: Ingresos a disminuir CAPTURAR/popup failed: %s", e)
    page.wait_for_timeout(PHASE3_SECTION_GAP_MS)
    LOG.info("")
    LOG.info("===== Phase 3: sección 5 - ¿Tienes ingresos adicionales? =====")
    # 5. ¿Tienes ingresos adicionales? — if Excel > SAT (difference > 1) → Sí + CAPTURAR popup with "Ingresos no considerados en el prellenado" and Importe = difference
    diferencia_adicionales = (excel_total_cobrados or 0.0) - (sat_total_cobrados or 0.0)
    need_ingresos_adicionales = diferencia_adicionales > 1
    si_no_adicionales_lbl = "Sí" if need_ingresos_adicionales else "No"
    if _fill_select_next_to_label(
        scope,
        page,
        "¿Tienes ingresos adicionales",
        si_no_adicionales_lbl,
        mapping=mapping,
        initial_dropdown_key="_isr_ingresos_adicionales",
    ):
        LOG.info("Phase 3: ingresos adicionales = %s (SAT=%.2f Excel=%.2f diff=%.2f)", si_no_adicionales_lbl, sat_total_cobrados, excel_total_cobrados, diferencia_adicionales)
    else:
        LOG.warning("Phase 3: could not set ingresos adicionales dropdown")
    page.wait_for_timeout(100)
    if need_ingresos_adicionales:
        # *Ingresos adicionales appears: CAPTURAR → popup AGREGAR → Concepto "Ingresos no considerados en el prellenado" → Importe = difference (int) → GUARDAR → CERRAR
        try:
            # Prefer specialized helper that targets the CAPTURAR to the right of "*Ingresos adicionales"
            capturar_clicked = _click_capturar_ingresos_adicionales(page, scope)
            if not capturar_clicked:
                capturar_clicked = _click_capturar_next_to_label(scope, "Ingresos adicionales")
            if capturar_clicked:
                LOG.info("Phase 3: Ingresos adicionales CAPTURAR pressed")
                page.wait_for_timeout(80)
                # Importe = difference with no decimals (e.g. 100.00 → 100)
                diferencia_adic_val = int(round(diferencia_adicionales))
                diferencia_adic_str = str(diferencia_adic_val)
                # Scope to the "Ingresos adicionales" popup: modal has title + Concepto + Importe + GUARDAR
                dialog = None
                for use_last, dialog_loc, to_ms in [
                    (False, page.get_by_role("dialog"), 800),
                    (False, page.locator("[role='dialog']"), 500),
                    (True, page.locator("div").filter(has_text=re.compile(r"Ingresos adicionales", re.I)).filter(has_text="Concepto").filter(has_text="Importe").filter(has_text="GUARDAR"), 400),
                    (True, page.locator("div").filter(has_text=re.compile(r"Ingresos adicionales", re.I)).filter(has=page.locator("select")), 350),
                ]:
                    try:
                        d = dialog_loc.last if use_last else dialog_loc.first
                        d.wait_for(state="visible", timeout=to_ms)
                        dialog = d
                        break
                    except Exception:
                        continue
                if dialog is None:
                    dialog = page
                # Click AGREGAR inside the popup so Concepto/Importe row is ready
                if dialog != page:
                    try:
                        dialog.get_by_role("button", name=re.compile(r"AGREGAR", re.I)).first.click(timeout=600)
                        page.wait_for_timeout(50)
                    except Exception:
                        pass
                else:
                    try:
                        page.get_by_role("button", name=re.compile(r"AGREGAR", re.I)).first.click(timeout=600)
                        page.wait_for_timeout(50)
                    except Exception:
                        pass
                # Concepto: fixed option "Ingresos no considerados en el prellenado"
                # Reuse the generic helper that finds the select next to the "Concepto" label inside this dialog.
                concepto_ok = _fill_select_next_to_label(
                    dialog,
                    page,
                    "Concepto",
                    "Ingresos no considerados en el prellenado",
                    mapping=None,
                    initial_dropdown_key=None,
                )
                # Importe: same row as Concepto dropdown (form row), then GUARDAR → CERRAR
                importe_ok = False
                try:
                    sel_first = dialog.locator("select").first
                    sel_first.wait_for(state="visible", timeout=200)
                    for row_xpath in ["xpath=ancestor::tr[1]", "xpath=ancestor::*[.//input][1]"]:
                        try:
                            row = sel_first.locator(row_xpath)
                            if row.count() == 0:
                                continue
                            inp = row.locator("input[type='text'], input[type='number'], input:not([type])").first
                            inp.wait_for(state="visible", timeout=180)
                            if inp.get_attribute("disabled") or inp.get_attribute("readonly"):
                                continue
                            inp.click()
                            inp.fill(diferencia_adic_str)
                            importe_ok = True
                            page.wait_for_timeout(25)
                            break
                        except Exception:
                            continue
                except Exception:
                    pass
                if not importe_ok:
                    try:
                        label_importe = dialog.get_by_text("Importe", exact=False).first
                        label_importe.wait_for(state="visible", timeout=300)
                        for xpath in [
                            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*//input[not(@disabled)]",
                            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*//input",
                            "xpath=(ancestor::tr[1])//input[not(@disabled)]",
                            "xpath=(ancestor::tr[1])//input",
                            "xpath=following-sibling::*//input",
                        ]:
                            try:
                                importe_inp = label_importe.locator(xpath).first
                                importe_inp.wait_for(state="visible", timeout=180)
                                if importe_inp.get_attribute("disabled") or importe_inp.get_attribute("readonly"):
                                    continue
                                importe_inp.click()
                                importe_inp.fill(diferencia_adic_str)
                                importe_ok = True
                                page.wait_for_timeout(25)
                                break
                            except Exception:
                                continue
                    except Exception:
                        pass
                if not importe_ok:
                    try:
                        for inp in dialog.locator("input[type='text'], input[type='number'], input:not([type])").all():
                            try:
                                if inp.get_attribute("disabled") or inp.get_attribute("readonly"):
                                    continue
                                inp.wait_for(state="visible", timeout=120)
                                inp.fill(diferencia_adic_str)
                                importe_ok = True
                                break
                            except Exception:
                                continue
                    except Exception:
                        pass
                if concepto_ok and importe_ok:
                    try:
                        btn_scope = dialog if dialog != page else page
                        btn_scope.get_by_role("button", name=re.compile(r"GUARDAR", re.I)).first.click(timeout=1500)
                        page.wait_for_timeout(150)
                        try:
                            confirm_btn = page.get_by_role("button", name=re.compile(r"ACEPTAR", re.I)).first
                            confirm_btn.wait_for(state="visible", timeout=800)
                            confirm_btn.click()
                            page.wait_for_timeout(80)
                        except Exception:
                            pass
                        btn_scope.get_by_role("button", name=re.compile(r"CERRAR", re.I)).first.click(timeout=1500)
                        LOG.info("Phase 3: Ingresos adicionales popup filled (diff=%.2f), closed", diferencia_adicionales)
                        page.wait_for_timeout(150)
                    except Exception as e_btn:
                        LOG.warning("Phase 3: Ingresos adicionales GUARDAR/CERRAR: %s", e_btn)
                else:
                    LOG.warning("Phase 3: Ingresos adicionales popup not filled (concepto=%s, importe=%s); not clicking GUARDAR", concepto_ok, importe_ok)
            else:
                LOG.warning("Phase 3: could not click Ingresos adicionales CAPTURAR")
        except Exception as e:
            LOG.warning("Phase 3: Ingresos adicionales CAPTURAR/popup failed: %s", e)
        page.wait_for_timeout(PHASE3_SECTION_GAP_MS)
    else:
        page.wait_for_timeout(PHASE3_SECTION_GAP_MS)
    LOG.info("")
    LOG.info("===== Phase 3: sección 6 - Total de ingresos percibidos por la actividad =====")
    # 6. Total de ingresos percibidos por la actividad: press CAPTURAR only when diff > 0 (when diff=0 the CAPTURAR option does not appear on SAT). Then popup "Total de ingresos efectivamente cobrados" → for each Excel row (Actividad empresarial, Actividad profesional, Uso o goce temporal) with value != "-": AGREGAR → Concepto → Importe → GUARDAR → ACEPTAR → then CERRAR
    if not need_ingresos_adicionales:
        LOG.info("Phase 3: Total percibidos skipped (diff=%.2f), CAPTURAR not shown on SAT", diferencia_adicionales)
    else:
        try:
            capturar_clicked = _try_click(page, mapping, "_isr_ingresos_capturar_total")
            if not capturar_clicked:
                capturar_clicked = _click_capturar_total_percibidos(page, scope)
            if not capturar_clicked:
                capturar_clicked = _click_capturar_next_to_label(scope, "Total de ingresos percibidos por la actividad")
            if not capturar_clicked:
                capturar_clicked = _click_capturar_next_to_label(scope, "Total de ingresos percibidos")
            if not capturar_clicked:
                raise RuntimeError("Could not click Total percibidos CAPTURAR link")
            LOG.info("Phase 3: Total percibidos CAPTURAR pressed")
            page.wait_for_timeout(200)
            # Resolve "Total de ingresos efectivamente cobrados" popup dialog
            dialog = None
            for use_last, dialog_loc, to_ms in [
                (False, page.get_by_role("dialog"), 900),
                (False, page.locator("[role='dialog']"), 600),
                (True, page.locator("div").filter(has_text=re.compile(r"Total de ingresos efectivamente cobrados", re.I)).filter(has_text=re.compile(r"Concepto|AGREGAR|Monto", re.I)), 450),
                (True, page.locator("div").filter(has_text=re.compile(r"Total de ingresos efectivamente cobrados", re.I)).filter(has=page.locator("select")), 400),
            ]:
                try:
                    d = dialog_loc.last if use_last else dialog_loc.first
                    d.wait_for(state="visible", timeout=to_ms)
                    dialog = d
                    break
                except Exception:
                    continue
            if dialog is None:
                dialog = page
            # Three Excel labels (col D or E) → SAT Concepto dropdown option; only add when value is not "-" or missing
            total_percibidos_entries = [
                ("Actividad empresarial", "Actividad empresarial"),
                ("Actividad profesional (honorarios)", "Servicios profesionales (Honorarios)"),
                ("Uso o goce temporal de bienes (arrendamiento)", "Uso o goce temporal de bienes"),
            ]
            for excel_label, sat_concepto in total_percibidos_entries:
                raw = label_map.get(excel_label)
                if raw is None:
                    continue
                # Skip when Excel shows "-" or the parsed numeric value is zero
                if isinstance(raw, str) and str(raw).strip() == "-":
                    continue
                parsed = _parse_currency(raw)
                if abs(parsed) < 0.005:
                    continue
                importe_str = str(int(round(parsed)))
                LOG.info("Phase 3: Total percibidos adding entry: excel_label=%r, sat_concepto=%r, importe=%s", excel_label, sat_concepto, importe_str)
                try:
                    if dialog != page:
                        dialog.get_by_role("button", name=re.compile(r"AGREGAR", re.I)).first.click(timeout=800)
                        page.wait_for_timeout(50)
                except Exception as e_ag:
                    LOG.warning("Phase 3: Total percibidos AGREGAR for %r: %s", excel_label, e_ag)
                    continue
                concepto_ok = False
                try:
                    loc_concepto = dialog.get_by_text(re.compile(r"\*?\s*Concepto", re.I))
                    if loc_concepto.count() == 0:
                        loc_concepto = dialog.get_by_text("Concepto", exact=False)
                    if loc_concepto.count() > 0:
                        label_concepto = loc_concepto.first
                        label_concepto.wait_for(state="visible", timeout=300)
                        for xpath in [
                            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*//select",
                            "xpath=(ancestor::tr[1])//select",
                            "xpath=following-sibling::*//select",
                            "xpath=..//select",
                        ]:
                            try:
                                concepto_dd = label_concepto.locator(xpath).first
                                concepto_dd.wait_for(state="visible", timeout=200)
                                concepto_dd.select_option(label=sat_concepto, timeout=2000)
                                concepto_ok = True
                                page.wait_for_timeout(30)
                                break
                            except Exception:
                                continue
                except Exception:
                    pass
                if not concepto_ok:
                    try:
                        sel = dialog.locator("select").first
                        sel.wait_for(state="visible", timeout=300)
                        sel.select_option(label=sat_concepto, timeout=2000)
                        concepto_ok = True
                        page.wait_for_timeout(30)
                    except Exception as e_c:
                        LOG.warning("Phase 3: Total percibidos Concepto %r: %s", sat_concepto, e_c)
                if concepto_ok:
                    LOG.info(
                        "Phase 3: Total percibidos popup: selected Concepto=%r for excel_label=%r",
                        sat_concepto,
                        excel_label,
                    )
                importe_ok = False
                try:
                    sel_first = dialog.locator("select").first
                    sel_first.wait_for(state="visible", timeout=200)
                    for row_xpath in ["xpath=ancestor::tr[1]", "xpath=ancestor::*[.//input][1]"]:
                        try:
                            row = sel_first.locator(row_xpath)
                            if row.count() == 0:
                                continue
                            inp = row.locator("input[type='text'], input[type='number'], input:not([type])").first
                            inp.wait_for(state="visible", timeout=180)
                            if inp.get_attribute("disabled") or inp.get_attribute("readonly"):
                                continue
                            inp.click()
                            inp.fill(importe_str)
                            importe_ok = True
                            page.wait_for_timeout(25)
                            break
                        except Exception:
                            continue
                except Exception:
                    pass
                if not importe_ok:
                    try:
                        label_importe = dialog.get_by_text("Importe", exact=False).first
                        label_importe.wait_for(state="visible", timeout=250)
                        for xpath in [
                            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*//input[not(@disabled)]",
                            "xpath=(ancestor::tr[1])//input[not(@disabled)]",
                            "xpath=following-sibling::*//input",
                        ]:
                            try:
                                importe_inp = label_importe.locator(xpath).first
                                importe_inp.wait_for(state="visible", timeout=180)
                                if importe_inp.get_attribute("readonly"):
                                    continue
                                importe_inp.click()
                                importe_inp.fill(importe_str)
                                importe_ok = True
                                page.wait_for_timeout(25)
                                break
                            except Exception:
                                continue
                    except Exception:
                        pass
                if importe_ok:
                    LOG.info(
                        "Phase 3: Total percibidos popup: filled Importe=%s for excel_label=%r",
                        importe_str,
                        excel_label,
                    )
                if concepto_ok and importe_ok:
                    try:
                        btn_scope = dialog if dialog != page else page
                        btn_scope.get_by_role("button", name=re.compile(r"GUARDAR", re.I)).first.click(timeout=600)
                        page.wait_for_timeout(60)
                        try:
                            confirm_btn = page.get_by_role("button", name=re.compile(r"ACEPTAR", re.I)).first
                            confirm_btn.wait_for(state="visible", timeout=500)
                            confirm_btn.click()
                            page.wait_for_timeout(70)
                        except Exception:
                            pass
                    except Exception as e_btn:
                        LOG.warning("Phase 3: Total percibidos GUARDAR/ACEPTAR for %r: %s", excel_label, e_btn)
                else:
                    LOG.warning("Phase 3: Total percibidos entry not filled (concepto=%s, importe=%s) for %r", concepto_ok, importe_ok, excel_label)
            # Close the popup
            try:
                btn_scope = dialog if dialog != page else page
                btn_scope.get_by_role("button", name=re.compile(r"CERRAR", re.I)).first.click(timeout=800)
                LOG.info("Phase 3: Total percibidos popup filled and closed")
            except Exception as e_close:
                LOG.warning("Phase 3: Total percibidos CERRAR: %s", e_close)
            page.wait_for_timeout(50)
        except Exception as e:
            LOG.warning("Phase 3: Total percibidos CAPTURAR/popup failed: %s", e)
    LOG.info("Phase 3: Ingresos form fill completed")

    # Phase 4: GUARDAR → Determinación tab (to the right of Ingresos) → VER DETALLE (ISR retenido por personas morales) → popup fill "ISR retenido no acreditable" → CERRAR → GUARDAR
    LOG.info("")
    LOG.info("===== Phase 4: Determinación (GUARDAR → Determinación tab → ISR retenido VER DETALLE) =====")
    sat_ui = sat_ui or DEFAULT_SAT_UI
    _det_tab = sat_ui.get("determinacion_tab_name") or "Determinación"
    _isr_row = sat_ui.get("isr_retenido_row_label") or "ISR retenido por personas morales"
    _isr_no_acred = sat_ui.get("isr_retenido_no_acreditable_label") or "ISR retenido no acreditable"
    _isr_excel_key = sat_ui.get("isr_retenido_excel_label") or "ISR retenido"
    _ver_det = sat_ui.get("ver_detalle_button") or "VER DETALLE"
    try:
        page.wait_for_timeout(400)
        guardar_clicked = False
        LOG.info("Phase 4: clicking GUARDAR (save Ingresos form)")
        # Main form GUARDAR (top-right); SAT may use button, input, or link — try visible one first
        for attempt, (loc, desc) in enumerate([
            (page.get_by_role("button", name=re.compile(r"GUARDAR", re.I)), "button"),
            (page.get_by_text("GUARDAR", exact=True), "text"),
            (page.locator("input[type='submit'][value*='GUARDAR'], input[type='button'][value*='GUARDAR']"), "input"),
        ]):
            try:
                if attempt == 0:
                    first_btn = loc.first
                    first_btn.wait_for(state="visible", timeout=8000)
                    if first_btn.get_attribute("disabled"):
                        continue
                    first_btn.click(timeout=5000)
                else:
                    for elem in loc.all():
                        try:
                            if elem.is_visible() and elem.get_attribute("disabled") != "true":
                                elem.click(timeout=5000)
                                guardar_clicked = True
                                LOG.info("Phase 4: GUARDAR clicked (save Ingresos form, %s)", desc)
                                break
                        except Exception:
                            continue
                    if guardar_clicked:
                        break
                guardar_clicked = True
                LOG.info("Phase 4: GUARDAR clicked (save Ingresos form, %s)", desc)
                break
            except Exception:
                continue
        if not guardar_clicked:
            raise RuntimeError("Could not click GUARDAR (main form save button)")
        page.wait_for_timeout(400)
        LOG.info("Phase 4: GUARDAR clicked, waiting for load")
        page.wait_for_load_state("domcontentloaded", timeout=5000)
        page.wait_for_timeout(200)
        LOG.info("Phase 4: clicking Determinación tab (to the right of Ingresos)")
        # Reuse same pattern as Phase 5 (Pago): check if section already visible first, then click tab only if needed.
        det_clicked = False
        try:
            if page.get_by_text(re.compile(re.escape(_isr_row), re.I)).first.is_visible(timeout=800):
                det_clicked = True
                LOG.info("Phase 4: Determinación section already visible (%r), skipping tab click", _isr_row)
            elif page.get_by_text(_ver_det, exact=False).first.is_visible(timeout=500):
                det_clicked = True
                LOG.info("Phase 4: Determinación section already visible (VER DETALLE), skipping tab click")
        except Exception:
            pass
        if not det_clicked:
            # Avoid matching hidden modal title "Determinación de la Base gravable"; click the visible tab in the tab bar only
            def _try_click_det_tab(elem, *, force: bool = False) -> bool:
                try:
                    elem.wait_for(state="visible", timeout=2000)
                    elem.scroll_into_view_if_needed(timeout=2000)
                    if force:
                        elem.click(force=True, timeout=3000)
                    else:
                        elem.click(timeout=3000)
                    page.wait_for_timeout(600)
                    # Verify section switched: row label or VER DETALLE should appear
                    if page.get_by_text(re.compile(re.escape(_isr_row), re.I)).first.is_visible(timeout=1500):
                        return True
                    if page.get_by_text(_ver_det, exact=False).first.is_visible(timeout=800):
                        return True
                    return False
                except Exception:
                    return False

            # 1) Prefer role=tab (proper tab widget)
            tab_loc = page.get_by_role("tab", name=re.compile(re.escape(_det_tab), re.I))
            if tab_loc.count() > 0 and _try_click_det_tab(tab_loc.first):
                det_clicked = True
                LOG.info("Phase 4: Determinación tab clicked (role=tab)")
            if not det_clicked:
                # 2) Tab bar: element with Determinación inside tablist/tab container
                for container in page.locator("[role='tablist'], .nav-tabs, ul.tabs, .tabs, [class*='tab']").all():
                    try:
                        if not container.is_visible():
                            continue
                        elem = container.get_by_text(_det_tab, exact=True).first
                        if elem.count() > 0 and _try_click_det_tab(elem):
                            det_clicked = True
                            LOG.info("Phase 4: Determinación tab clicked (inside tablist)")
                            break
                    except Exception:
                        continue
            if not det_clicked:
                for elem in page.get_by_text(_det_tab, exact=True).all():
                    try:
                        if not elem.is_visible():
                            continue
                        if elem.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").count() > 0:
                            continue
                        if _try_click_det_tab(elem):
                            det_clicked = True
                            LOG.info("Phase 4: Determinación tab clicked")
                            break
                    except Exception:
                        continue
            if not det_clicked:
                for elem in page.locator("a, button, [role='tab'], li, span").filter(has_text=re.compile(r"^%s$" % re.escape(_det_tab), re.I)).all():
                    try:
                        if not elem.is_visible():
                            continue
                        if elem.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").count() > 0:
                            continue
                        if _try_click_det_tab(elem):
                            det_clicked = True
                            LOG.info("Phase 4: Determinación tab clicked (fallback)")
                            break
                    except Exception:
                        continue
            if not det_clicked:
                # Last resort: force click first visible Determinación tab (exclude modals)
                for elem in page.get_by_text(_det_tab, exact=True).all():
                    try:
                        if not elem.is_visible():
                            continue
                        if elem.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").count() > 0:
                            continue
                        elem.scroll_into_view_if_needed(timeout=2000)
                        elem.click(force=True, timeout=3000)
                        page.wait_for_timeout(400)
                        if page.get_by_text(re.compile(re.escape(_isr_row), re.I)).first.is_visible(timeout=1500):
                            det_clicked = True
                            LOG.info("Phase 4: Determinación tab clicked (force)")
                            break
                    except Exception:
                        continue
            if not det_clicked:
                raise RuntimeError("Could not click Determinación tab (visible tab not found or section did not load)")
        page.wait_for_timeout(400)
        LOG.info("Phase 4: Determinación tab clicked, waiting for section to load")
        page.wait_for_load_state("domcontentloaded", timeout=5000)
        page.wait_for_timeout(200)
        LOG.info("Phase 4: Determinación section loaded; next: VER DETALLE on same row as label 'ISR retenido por personas morales'")

        # Excel value for "ISR retenido no acreditable": row with label from config (e.g. ISR retenido) in col D or E
        LOG.info("Phase 4: getting value from Excel: row with label %r in col D or E (parse as Base gravable)", _isr_excel_key)
        label_map = data.get("label_map") or {}
        isr_retenido_raw = label_map.get(_isr_excel_key)
        isr_retenido_parsed = _parse_currency(isr_retenido_raw) if isr_retenido_raw is not None else 0.0
        workbook_path = data.get("workbook_path") or "(workbook not set)"
        LOG.info(
            "Phase 4: Excel value for ISR retenido no acreditable: label=%r | workbook=%s | raw=%s → parsed=%.2f",
            _isr_excel_key, workbook_path, isr_retenido_raw, isr_retenido_parsed,
        )
        isr_retenido_str = str(int(round(isr_retenido_parsed))) if isr_retenido_parsed is not None else "0"

        ver_detalle_clicked = False
        LOG.info("Phase 4: clicking VER DETALLE (same row as label %r, same logic as Phase 3 CAPTURAR)", _isr_row)
        try:
            # Same logic as Phase 3: of all VER DETALLE, click the one whose row contains the configured row label
            ver_detalle_clicked = _click_ver_detalle_isr_retenido(page, sat_ui)
            if not ver_detalle_clicked:
                ver_detalle_clicked = _click_ver_detalle_next_to_label(page, _isr_row, sat_ui)
            if not ver_detalle_clicked:
                try:
                    scope = _get_isr_ingresos_scope(page)
                    scope.locator("#tab457maincontainer1").first.wait_for(state="attached", timeout=500)
                    tab_scope = scope.locator("#tab457maincontainer1").first
                    ver_detalle_clicked = _click_ver_detalle_isr_retenido(tab_scope, sat_ui) or _click_ver_detalle_next_to_label(tab_scope, _isr_row, sat_ui)
                except Exception:
                    pass
            if ver_detalle_clicked:
                LOG.info("Phase 4: VER DETALLE clicked; waiting for popup to appear")
                page.wait_for_timeout(200)
                popup_visible = False
                for _ in range(6):
                    try:
                        if page.get_by_role("dialog").first.is_visible():
                            popup_visible = True
                            break
                        if page.get_by_text(re.compile(re.escape(_isr_no_acred), re.I)).first.is_visible():
                            popup_visible = True
                            break
                    except Exception:
                        pass
                    page.wait_for_timeout(200)
                if not popup_visible:
                    LOG.warning("Phase 4: VER DETALLE was clicked but popup did not appear (dialog not visible); treating as failed")
                    ver_detalle_clicked = False
                else:
                    LOG.info("Phase 4: popup appeared")
            else:
                LOG.warning("Phase 4: could not click VER DETALLE next to %s", _isr_row)
        except Exception as e:
            LOG.warning("Phase 4: VER DETALLE click failed: %s", e)

        if ver_detalle_clicked:
            page.wait_for_timeout(200)
            LOG.info("Phase 4: looking for popup with label %r", _isr_no_acred)
            try:
                page.get_by_text(re.compile(re.escape(_isr_no_acred), re.I)).first.wait_for(state="visible", timeout=150)
                LOG.info("Phase 4: ISR retenido popup (%s) visible", _isr_no_acred)
            except Exception:
                pass
            page.wait_for_timeout(10)
            # Resolve popup dialog — same logic as Phase 3 Descuentos popup
            dialog = None
            for try_dialog in [
                lambda: page.get_by_role("dialog").first,
                lambda: page.locator(".modal, [role='dialog']").first,
                lambda: page.get_by_text(re.compile(re.escape(_isr_no_acred), re.I)).first.locator("xpath=ancestor::*[contains(@class,'modal') or contains(@class,'dialog') or @role='dialog'][1]"),
            ]:
                try:
                    d = try_dialog()
                    d.wait_for(state="visible", timeout=80)
                    dialog = d
                    break
                except Exception:
                    continue
            if dialog is None:
                dialog = page
            filled = False

            def _set_isr_no_acreditable_value(inp, value_str: str) -> bool:
                """Fill input with value; SAT control may ignore .fill(), so verify and fall back to keyboard + JS. Uses Phase 4 popup timeouts (modal needs more time)."""
                def _get_val():
                    try:
                        return (inp.input_value() or inp.get_attribute("value") or "").strip()
                    except Exception:
                        return ""
                try:
                    inp.scroll_into_view_if_needed(timeout=_PHASE4_POPUP_VISIBLE_MS)
                    inp.click()
                    page.wait_for_timeout(_PHASE4_POPUP_WAIT_MS)
                    inp.fill("")
                    page.wait_for_timeout(_PHASE4_POPUP_WAIT_MS)
                    inp.fill(value_str)
                    page.wait_for_timeout(_PHASE4_POPUP_WAIT_MS * 2)
                    if _get_val() == value_str:
                        return True
                    inp.click()
                    page.wait_for_timeout(_PHASE4_POPUP_WAIT_MS)
                    page.keyboard.press("Control+a")
                    page.wait_for_timeout(_PHASE4_POPUP_WAIT_MS)
                    page.keyboard.type(value_str, delay=40)
                    page.wait_for_timeout(_PHASE4_POPUP_WAIT_MS * 2)
                    if _get_val() == value_str:
                        return True
                    inp.click()
                    page.wait_for_timeout(_PHASE4_POPUP_WAIT_MS)
                    inp.press_sequentially(value_str, delay=40)
                    page.wait_for_timeout(_PHASE4_POPUP_WAIT_MS * 2)
                    if _get_val() == value_str:
                        return True
                    inp.evaluate("""el => {
                        el.value = arguments[0];
                        el.dispatchEvent(new Event('input', { bubbles: true }));
                        el.dispatchEvent(new Event('change', { bubbles: true }));
                    }""", value_str)
                    page.wait_for_timeout(_PHASE4_POPUP_WAIT_MS)
                    return _get_val() == value_str
                except Exception:
                    return False

            try:
                # Strategy 1: Same as Descuentos popup (section 3) — label then following::input[1] and following::input[2]
                try:
                    label_el = dialog.get_by_text(re.compile(re.escape(_isr_no_acred), re.I)).first
                    label_el.wait_for(state="visible", timeout=_PHASE4_POPUP_VISIBLE_MS)
                    for input_index in [1, 2]:
                        try:
                            following_input = label_el.locator(f"xpath=following::input[{input_index}]")
                            if following_input.count() > 0:
                                inp = following_input.first
                                inp.wait_for(state="visible", timeout=_PHASE4_POPUP_VISIBLE_MS)
                                if inp.get_attribute("disabled"):
                                    continue
                                if _set_isr_no_acreditable_value(inp, isr_retenido_str):
                                    filled = True
                                    LOG.info("Phase 4: filled 'ISR retenido no acreditable' (following::input[%s], same as Descuentos) with value=%s", input_index, isr_retenido_str)
                                    break
                        except Exception:
                            continue
                except Exception:
                    pass

                if not filled:
                    # Strategy 2: By position — dialog has 3 ISR inputs: (0) tipo ingreso, (1) a adicionar, (2) no acreditable
                    try:
                        modal_inputs = dialog.locator("input[type='text'], input[type='number'], input:not([type])")
                        if modal_inputs.count() >= 3:
                            inp = modal_inputs.nth(2)
                            inp.wait_for(state="visible", timeout=_PHASE4_POPUP_VISIBLE_MS)
                            row = inp.locator("xpath=ancestor::tr[1] | ancestor::div[contains(@class,'row')][1]")
                            if row.count() > 0:
                                row_text = (row.first.inner_text() or "").lower()
                                if "no acreditable" in row_text and "a adicionar" not in row_text:
                                    if _set_isr_no_acreditable_value(inp, isr_retenido_str):
                                        filled = True
                                        LOG.info("Phase 4: filled 'ISR retenido no acreditable' (3rd input, row text) with value=%s", isr_retenido_str)
                    except Exception:
                        pass

                if not filled:
                    # Strategy 3: Label "no acreditable" + xpaths (exclude row with "a adicionar")
                    for label_el in dialog.get_by_text(re.compile(re.escape(_isr_no_acred), re.I)).all():
                        try:
                            if not label_el.is_visible():
                                continue
                            # Same row must contain "no acreditable" and not be the "a adicionar" row
                            row = label_el.locator("xpath=ancestor::tr[1]")
                            if row.count() > 0 and row.first.get_by_text(re.compile(r"a adicionar", re.I)).count() > 0:
                                continue
                            for xpath in [
                                "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*//input",
                                "xpath=(ancestor::tr[1])//input",
                                "xpath=following-sibling::*//input",
                                "xpath=..//input",
                            ]:
                                try:
                                    inp = label_el.locator(xpath).first
                                    if inp.count() == 0:
                                        continue
                                    inp.wait_for(state="visible", timeout=_PHASE4_POPUP_VISIBLE_MS)
                                    if inp.get_attribute("disabled"):
                                        continue
                                    if _set_isr_no_acreditable_value(inp, isr_retenido_str):
                                        filled = True
                                        LOG.info("Phase 4: filled 'ISR retenido no acreditable' textbox with value=%s (from Excel)", isr_retenido_str)
                                        break
                                except Exception:
                                    continue
                            if filled:
                                break
                        except Exception:
                            continue

                if not filled:
                    try:
                        inp = dialog.get_by_label(re.compile(re.escape(_isr_no_acred), re.I)).first
                        inp.wait_for(state="visible", timeout=_PHASE4_POPUP_VISIBLE_MS)
                        if _set_isr_no_acreditable_value(inp, isr_retenido_str):
                            filled = True
                            LOG.info("Phase 4: filled 'ISR retenido no acreditable' (by label) with value=%s from Excel", isr_retenido_str)
                    except Exception:
                        pass
                if not filled:
                    # Strategy 4: Iterate inputs, pick the one in a row with "no acreditable" and without "a adicionar"
                    try:
                        modal_inputs = dialog.locator("input[type='text'], input[type='number'], input:not([type])")
                        n = min(modal_inputs.count(), 12)
                        for idx in range(n):
                            inp = modal_inputs.nth(idx)
                            try:
                                inp.wait_for(state="visible", timeout=_PHASE4_POPUP_VISIBLE_MS)
                                row = inp.locator("xpath=ancestor::tr[1] | ancestor::div[contains(@class,'row')][1]")
                                if row.count() == 0:
                                    continue
                                row_text = (row.first.inner_text() or "").lower()
                                if "no acreditable" not in row_text or "a adicionar" in row_text:
                                    continue
                                if _set_isr_no_acreditable_value(inp, isr_retenido_str):
                                    filled = True
                                    LOG.info("Phase 4: filled 'ISR retenido no acreditable' (by row text) with value=%s from Excel", isr_retenido_str)
                                    break
                            except Exception:
                                continue
                    except Exception:
                        pass
            except Exception as e_fill:
                LOG.warning("Phase 4: could not fill ISR retenido no acreditable: %s", e_fill)
            page.wait_for_timeout(150)
            LOG.info("Phase 4: clicking CERRAR in ISR retenido popup")
            page.get_by_role("button", name=re.compile(r"CERRAR", re.I)).first.click(timeout=1500)
            page.wait_for_timeout(200)
            LOG.info("Phase 4: CERRAR clicked, popup closed; next: GUARDAR and wait for load")
        LOG.info("Phase 4: clicking GUARDAR (after Determinación / ISR retenido popup)")
        guardar2_clicked = False
        for loc in [
            page.get_by_role("button", name=re.compile(r"GUARDAR", re.I)),
            page.locator("input[type='submit'][value*='GUARDAR'], input[type='button'][value*='GUARDAR']"),
            page.get_by_text("GUARDAR", exact=True),
        ]:
            try:
                first_btn = loc.first
                first_btn.wait_for(state="visible", timeout=6000)
                if first_btn.get_attribute("disabled"):
                    continue
                first_btn.click(timeout=4000)
                guardar2_clicked = True
                break
            except Exception:
                continue
        if guardar2_clicked:
            page.wait_for_timeout(400)
            LOG.info("Phase 4: GUARDAR clicked, waiting for load")
            page.wait_for_load_state("domcontentloaded", timeout=5000)
            page.wait_for_timeout(200)
            LOG.info("Phase 4: load complete after GUARDAR")
        LOG.info("Phase 4: Determinación (ISR retenido no acreditable) completed")
    except Exception as e:
        LOG.warning("Phase 4 (Determinación / ISR retenido VER DETALLE) failed: %s", e)

    # Phase 5: Pago tab → *¿Tienes compensaciones por aplicar? → No, *¿Tienes estímulos por aplicar? → No → GUARDAR → wait for load
    # ISR Pago section detection reuses same pattern as Phase 4 (Determinación): 1) check if section already visible by
    # section-specific labels; 2) if not, click tab then verify section content (labels) appeared. Success = any of:
    # compensaciones/estímulos question text OR all three labels "A cargo", "Total de contribuciones", "Total de aplicaciones".
    LOG.info("")
    LOG.info("===== Phase 5: Pago (Pago tab → compensaciones/estímulos → No → GUARDAR) =====")
    _pago_tab = sat_ui.get("pago_tab_name") or "Pago"
    _comp_q = sat_ui.get("isr_pago_compensaciones_question") or "compensaciones por aplicar"
    _estim_q = sat_ui.get("isr_pago_estimulos_question") or "estímulos por aplicar"
    _pago_a_cargo = sat_ui.get("isr_pago_a_cargo_label") or "A cargo"
    _pago_total_contrib = sat_ui.get("isr_pago_total_contribuciones_label") or "Total de contribuciones"
    _pago_total_aplic = sat_ui.get("isr_pago_total_aplicaciones_label") or "Total de aplicaciones"
    _btn_guardar = sat_ui.get("btn_guardar") or "GUARDAR"

    def _pago_section_loaded(page_ctx, timeout_ms: int = 800) -> bool:
        """True if ISR Pago section content is visible: compensaciones/estímulos OR (A cargo AND Total de contribuciones AND Total de aplicaciones). Uses element visibility first; fallback: body inner_text (draft-style) when get_by_text().first matches hidden node."""
        # 1) Element-based: visible node with these labels
        try:
            if page_ctx.get_by_text(re.compile(re.escape(_comp_q), re.I)).first.is_visible(timeout=timeout_ms):
                return True
            if page_ctx.get_by_text(re.compile(re.escape(_estim_q), re.I)).first.is_visible(timeout=timeout_ms):
                return True
            if (
                page_ctx.get_by_text(re.compile(re.escape(_pago_a_cargo), re.I)).first.is_visible(timeout=timeout_ms)
                and page_ctx.get_by_text(re.compile(re.escape(_pago_total_contrib), re.I)).first.is_visible(timeout=timeout_ms)
                and page_ctx.get_by_text(re.compile(re.escape(_pago_total_aplic), re.I)).first.is_visible(timeout=timeout_ms)
            ):
                return True
        except Exception:
            pass
        # 2) Reuse draft-style detection: body text contains labels (robust when get_by_text matches hidden element)
        try:
            body = (page_ctx.locator("body").inner_text(timeout=timeout_ms) or "").lower()
            if _comp_q.lower() in body or _estim_q.lower() in body:
                return True
            if (
                _pago_a_cargo.lower() in body
                and _pago_total_contrib.lower() in body
                and _pago_total_aplic.lower() in body
            ):
                LOG.info("Phase 5: Pago section detected via body text (A cargo + Total de contribuciones + Total de aplicaciones)")
                return True
        except Exception:
            pass
        # Check iframes like draft detection (SAT may render form in iframe)
        try:
            for frame in page_ctx.frames:
                if frame == page_ctx.main_frame:
                    continue
                iframe_body = (frame.locator("body").inner_text(timeout=min(timeout_ms, 500)) or "").lower()
                if _comp_q.lower() in iframe_body or _estim_q.lower() in iframe_body:
                    return True
                if _pago_a_cargo.lower() in iframe_body and _pago_total_contrib.lower() in iframe_body and _pago_total_aplic.lower() in iframe_body:
                    LOG.info("Phase 5: Pago section detected via iframe body text")
                    return True
        except Exception:
            pass
        return False

    try:
        page.wait_for_timeout(400)
        # If Pago section is already visible (e.g. tab was clicked in a previous attempt or already selected), treat as success.
        pago_clicked = False
        LOG.info("Phase 5: checking if Pago section already visible (%r / %r or %r + %r + %r)", _comp_q, _estim_q, _pago_a_cargo, _pago_total_contrib, _pago_total_aplic)
        try:
            if _pago_section_loaded(page, timeout_ms=1500):
                pago_clicked = True
                LOG.info("Phase 5: Pago section already visible, skipping tab click")
            else:
                LOG.info("Phase 5: Pago section not visible yet, will try clicking tab")
        except Exception as e:
            LOG.debug("Phase 5: early visibility check failed: %s", e)
        if not pago_clicked:
            LOG.info("Phase 5: clicking Pago tab (to the right of Determinación)")

            def _try_click_pago_tab(elem, *, force: bool = False, attempt_name: str = "") -> bool:
                """Click Pago tab element; success = section content visible (comp/estim or A cargo + Total de contribuciones + Total de aplicaciones)."""
                try:
                    elem.wait_for(state="visible", timeout=2000)
                    elem.scroll_into_view_if_needed(timeout=2000)
                    if force:
                        elem.click(force=True, timeout=3000)
                    else:
                        elem.click(timeout=3000)
                    page.wait_for_timeout(600)
                    if _pago_section_loaded(page):
                        return True
                    LOG.info("Phase 5: attempt %s: tab clicked but section load check failed (%r/%r or %r+%r+%r not visible in time)", attempt_name or "?", _comp_q, _estim_q, _pago_a_cargo, _pago_total_contrib, _pago_total_aplic)
                    return False
                except Exception as e:
                    LOG.info("Phase 5: attempt %s: failed — %s", attempt_name or "?", e)
                    return False

            attempt = 0
            tab_loc = page.get_by_role("tab", name=re.compile(re.escape(_pago_tab), re.I))
            if tab_loc.count() > 0:
                attempt += 1
                LOG.info("Phase 5: attempt %s — role=tab name=%r", attempt, _pago_tab)
                if _try_click_pago_tab(tab_loc.first, attempt_name="role=tab"):
                    pago_clicked = True
                    LOG.info("Phase 5: Pago tab clicked (role=tab)")
            if not pago_clicked:
                for container in page.locator("[role='tablist'], .nav-tabs, ul.tabs, .tabs, [class*='tab']").all():
                    try:
                        if not container.is_visible():
                            continue
                        elem = container.get_by_text(_pago_tab, exact=True).first
                        if elem.count() > 0:
                            attempt += 1
                            LOG.info("Phase 5: attempt %s — inside tablist", attempt)
                            if _try_click_pago_tab(elem, attempt_name="tablist"):
                                pago_clicked = True
                                LOG.info("Phase 5: Pago tab clicked (inside tablist)")
                                break
                    except Exception as e:
                        LOG.debug("Phase 5: tablist container attempt: %s", e)
            if not pago_clicked:
                for elem in page.get_by_text(_pago_tab, exact=True).all():
                    try:
                        if not elem.is_visible():
                            continue
                        if elem.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").count() > 0:
                            continue
                        attempt += 1
                        LOG.info("Phase 5: attempt %s — get_by_text exact", attempt)
                        if _try_click_pago_tab(elem, attempt_name="get_by_text"):
                            pago_clicked = True
                            LOG.info("Phase 5: Pago tab clicked")
                            break
                    except Exception as e:
                        LOG.debug("Phase 5: get_by_text attempt: %s", e)
            if not pago_clicked:
                for elem in page.locator("a, button, [role='tab'], li, span").filter(has_text=re.compile(r"^%s$" % re.escape(_pago_tab), re.I)).all():
                    try:
                        if not elem.is_visible():
                            continue
                        if elem.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").count() > 0:
                            continue
                        attempt += 1
                        LOG.info("Phase 5: attempt %s — filter Pago", attempt)
                        if _try_click_pago_tab(elem, attempt_name="filter"):
                            pago_clicked = True
                            LOG.info("Phase 5: Pago tab clicked (fallback)")
                            break
                    except Exception as e:
                        LOG.debug("Phase 5: filter attempt: %s", e)
            if not pago_clicked:
                for elem in page.get_by_text(_pago_tab, exact=True).all():
                    try:
                        if not elem.is_visible():
                            continue
                        if elem.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").count() > 0:
                            continue
                        attempt += 1
                        LOG.info("Phase 5: attempt %s — force click", attempt)
                        elem.scroll_into_view_if_needed(timeout=2000)
                        elem.click(force=True, timeout=3000)
                        page.wait_for_timeout(1000)
                        if _pago_section_loaded(page):
                            pago_clicked = True
                            LOG.info("Phase 5: Pago tab clicked (force)")
                            break
                        LOG.info("Phase 5: attempt %s (force): tab clicked but section load check failed", attempt)
                    except Exception as e:
                        LOG.info("Phase 5: attempt %s (force): failed — %s", attempt, e)
            if not pago_clicked:
                LOG.warning("Phase 5: all %s attempt(s) failed; success = %r/%r or %r+%r+%r visible after click", attempt, _comp_q, _estim_q, _pago_a_cargo, _pago_total_contrib, _pago_total_aplic)
                raise RuntimeError("Could not click Pago tab (visible tab not found or section did not load)")
        page.wait_for_timeout(800)
        LOG.info("Phase 5: Pago tab clicked, waiting for section to load")
        page.wait_for_load_state("domcontentloaded", timeout=5000)
        page.wait_for_timeout(500)
        LOG.info("Phase 5: Pago section loaded; next: dropdowns *¿Tienes compensaciones por aplicar? and *¿Tienes estímulos por aplicar? → No")

        LOG.info("Phase 5: selecting *%s → No", _comp_q)
        # Pago uses custom dropdown widgets; go directly to the Pago-specific helper to avoid slow generic resolution.
        comp_ok = _fill_pago_custom_dropdown(page, _comp_q, "No")
        if comp_ok:
            LOG.info("Phase 5: dropdown *%s set to No", _comp_q)
        else:
            LOG.warning("Phase 5: could not set *%s to No", _comp_q)
        page.wait_for_timeout(120)
        LOG.info("Phase 5: selecting *%s → No", _estim_q)
        # Same for estímulos: use Pago-specific helper directly.
        estim_ok = _fill_pago_custom_dropdown(page, _estim_q, "No")
        if estim_ok:
            LOG.info("Phase 5: dropdown *%s set to No", _estim_q)
        else:
            LOG.warning("Phase 5: could not set *%s to No", _estim_q)
        page.wait_for_timeout(200)

        LOG.info("Phase 5: clicking GUARDAR")
        guardar_pago_clicked = False
        _guardar_pat = re.compile(re.escape(_btn_guardar), re.I)
        for loc in [
            page.get_by_role("button", name=_guardar_pat),
            page.locator("input[type='submit'][value*='%s'], input[type='button'][value*='%s']" % (_btn_guardar, _btn_guardar)),
            page.get_by_text(_btn_guardar, exact=True),
        ]:
            try:
                first_btn = loc.first
                first_btn.wait_for(state="visible", timeout=4000)
                if first_btn.get_attribute("disabled"):
                    continue
                first_btn.click(timeout=3000)
                guardar_pago_clicked = True
                break
            except Exception:
                continue
        if guardar_pago_clicked:
            page.wait_for_timeout(400)
            LOG.info("Phase 5: GUARDAR clicked, waiting for load")
            page.wait_for_load_state("domcontentloaded", timeout=5000)
            page.wait_for_timeout(200)
            LOG.info("Phase 5: load complete after GUARDAR")
        LOG.info("Phase 5: Pago completed")
    except Exception as e:
        LOG.warning("Phase 5 (Pago tab / compensaciones / estímulos / GUARDAR) failed: %s", e)


def fill_iva_simplificado(page: Page, mapping: dict, data: dict, sat_ui: dict | None = None, iva_determinacion_fields: list[tuple[str, str]] | None = None) -> None:
    """
    Central entry for IVA simplificado de confianza. Run after fill_isr_ingresos_form.
    Navigates: ADMINISTRACIÓN DE LA DECLARACIÓN → IVA simplificado de confianza → fill Determinación form → GUARDAR → Pago tab.
    """
    sat_ui = sat_ui or DEFAULT_SAT_UI
    LOG.info("")
    LOG.info("===== IVA simplificado de confianza (central) =====")
    if not click_administracion_declaracion(page, mapping, sat_ui):
        LOG.warning("IVA simplificado: could not click ADMINISTRACIÓN DE LA DECLARACIÓN")
    page.wait_for_timeout(400)
    if not open_obligation_iva(page, mapping, sat_ui):
        LOG.warning("IVA simplificado: could not click IVA simplificado de confianza")
    page.wait_for_timeout(400)
    fill_iva_simplificado_determinacion(page, mapping, data, iva_determinacion_fields=iva_determinacion_fields)
    LOG.info("===== IVA simplificado de confianza (central) complete =====")
    LOG.info("")


def _fill_iva_popup_input_by_position(dialog_scope, page_for_wait: Page, step_1based: int, value_str: str) -> bool:
    """Fallback for IVA acreditable popup: fill the nth visible editable input in dialog (1-based). Reuses ISR modal pattern (same-row / position)."""
    def _set(inp, val: str) -> bool:
        try:
            cur = (inp.input_value() or inp.get_attribute("value") or "").strip()
            if cur == val or (cur.replace(",", "") == val.replace(",", "")):
                return True
            inp.scroll_into_view_if_needed(timeout=_TEXTBOX_FILL_VISIBLE_MS)
            inp.click()
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS)
            inp.fill("")
            inp.fill(val)
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS * 2)
            return (inp.input_value() or "").strip() == val or (inp.input_value() or "").replace(",", "") == val.replace(",", "")
        except Exception:
            return False

    try:
        inputs = dialog_scope.locator(
            "input[type='text'], input[type='number'], input:not([type='submit']):not([type='button']):not([type='hidden'])"
        )
        idx = step_1based - 1
        n = 0
        for i in range(min(inputs.count(), 12)):
            inp = inputs.nth(i)
            try:
                if not inp.is_visible(timeout=_TEXTBOX_FILL_VISIBLE_MS):
                    continue
                if inp.get_attribute("disabled") or inp.get_attribute("readonly"):
                    continue
                if n == idx:
                    if _set(inp, value_str):
                        LOG.info("IVA Determinación popup: filled input by position (index %s)", step_1based)
                        return True
                    return False
                n += 1
            except Exception:
                continue
    except Exception:
        pass
    return False


def _fill_iva_input_by_position(
    page_for_wait: Page, scope, step_1based: int, total_fields: int, value_str: str
) -> bool:
    """Fallback: fill the nth visible editable input in scope (1-based). Used when label-based locate fails."""
    def _set(inp, val: str) -> bool:
        try:
            cur = (inp.input_value() or inp.get_attribute("value") or "").strip()
            if cur == val or (cur.replace(",", "") == val.replace(",", "")):
                return True
            inp.scroll_into_view_if_needed(timeout=_TEXTBOX_FILL_VISIBLE_MS)
            inp.click()
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS)
            inp.fill("")
            inp.fill(val)
            page_for_wait.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS * 2)
            return (inp.input_value() or "").strip() == val or (inp.input_value() or "").replace(",", "") == val.replace(",", "")
        except Exception:
            return False

    try:
        inputs = scope.locator(
            "input[type='text'], input[type='number'], input:not([type='submit']):not([type='button']):not([type='hidden'])"
        )
        idx = step_1based - 1
        n = 0
        for i in range(min(inputs.count(), 12)):
            inp = inputs.nth(i)
            try:
                if not inp.is_visible(timeout=_TEXTBOX_FILL_VISIBLE_MS):
                    continue
                if inp.get_attribute("disabled") or inp.get_attribute("readonly"):
                    continue
                if n == idx:
                    if _set(inp, value_str):
                        LOG.info("IVA Determinación: filled input by position (index %s)", step_1based)
                        return True
                    return False
                n += 1
            except Exception:
                continue
    except Exception:
        pass
    return False


def fill_iva_simplificado_determinacion(page: Page, mapping: dict, data: dict, iva_determinacion_fields: list[tuple[str, str]] | None = None) -> None:
    """
    Fill IVA simplificado de confianza Determinación form (run after navigating from administración to IVA).
    Main form: fields from iva_determinacion_fields (excel_label, form_label) or default list.
    Then CAPTURAR next to *IVA acreditable del periodo → popup: fill IVA acreditable por actividades gravadas 16%/8%/0% (from Excel), actividades mixtas = 0 → CERRAR → GUARDAR → click Pago tab.
    """
    LOG.info("")
    LOG.info("===== IVA simplificado de confianza: Determinación form =====")
    label_map = data.get("label_map") or {}
    scope = page
    main_fields = iva_determinacion_fields if iva_determinacion_fields else list(DEFAULT_IVA_DETERMINACION_FIELDS)

    # Wait for IVA form (Determinación)
    LOG.info("IVA Determinación: waiting for IVA form to load")
    page.wait_for_timeout(100)
    try:
        page.get_by_text(re.compile(r"IVA\s+simplificado\s+de\s+confianza", re.I)).first.wait_for(state="visible", timeout=1200)
        LOG.info("IVA Determinación: IVA form title visible")
    except Exception:
        LOG.debug("IVA Determinación: IVA form title wait skipped or timed out")
    page.wait_for_timeout(50)

    # Scope to Determinación tab content to avoid matching hidden/duplicate elements from other tabs
    iva_scope = scope
    try:
        for candidate in [
            page.get_by_role("tabpanel").filter(has=page.get_by_text("Actividades gravadas a la tasa del 16%")).first,
            page.locator("[role='tabpanel']").filter(has=page.get_by_text("Actividades gravadas", exact=False)).first,
            page.locator(".tab-pane, [class*='tabpanel'], [class*='tab-content']").filter(has=page.get_by_text("Actividades gravadas", exact=False)).first,
        ]:
            if candidate.count() > 0 and candidate.first.is_visible(timeout=_TEXTBOX_FILL_VISIBLE_MS):
                iva_scope = candidate.first
                LOG.info("IVA Determinación: using tabpanel scope for field lookup")
                break
    except Exception:
        pass

    for step, (excel_label, form_label) in enumerate(main_fields, start=1):
        raw = label_map.get(excel_label)
        val = _parse_currency(raw) if raw is not None else 0.0
        value_str = str(int(round(val))) if val is not None else "0"
        if form_label == "IVA retenido" and val is not None and val < 0:
            value_str = str(abs(int(round(val))))
        LOG.info("IVA Determinación: step %s/%s — filling %r (excel %r) with value %s", step, len(main_fields), form_label, excel_label, value_str)
        filled = _try_fill(iva_scope, page, mapping, form_label, value_str)
        if not filled:
            filled = _fill_input_next_to_label(iva_scope, page, form_label, value_str)
        if not filled:
            filled = _fill_iva_input_by_position(page, iva_scope, step, len(main_fields), value_str)
        if filled:
            LOG.info("IVA Determinación: step %s/5 — filled %r", step, form_label)
        else:
            LOG.warning("IVA Determinación: step %s/%s — could not fill %r", step, len(main_fields), form_label)
        page.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS)

    # Click CAPTURAR next to *IVA acreditable del periodo (reuse ISR logic: iterate CAPTURARs, match row by label text)
    LOG.info("IVA Determinación: step — clicking CAPTURAR next to '*IVA acreditable del periodo'")
    capturar_ok = _click_capturar_iva_acreditable(iva_scope)
    if not capturar_ok:
        capturar_ok = _click_capturar_next_to_label(iva_scope, "IVA acreditable del periodo", occurrence=0)
    if not capturar_ok:
        capturar_ok = _click_capturar_next_to_label(iva_scope, "*IVA acreditable del periodo", occurrence=0)
    if not capturar_ok:
        capturar_ok = _click_capturar_iva_acreditable(page)
    if not capturar_ok:
        LOG.warning("IVA Determinación: could not click CAPTURAR for IVA acreditable del periodo")
    else:
        LOG.info("IVA Determinación: CAPTURAR clicked; waiting for popup")
        page.wait_for_timeout(80)
        # Wait for popup
        dialog = None
        for _ in range(8):
            try:
                if page.get_by_role("dialog").first.is_visible(timeout=_TEXTBOX_FILL_VISIBLE_MS):
                    dialog = page.get_by_role("dialog").first
                    LOG.info("IVA Determinación: popup visible (dialog role)")
                    break
                if page.get_by_text(re.compile(r"IVA\s+acreditable\s+del\s+periodo", re.I)).first.is_visible():
                    dialog = page.get_by_text(re.compile(r"IVA\s+acreditable\s+del\s+periodo", re.I)).first.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog'][1]").first
                    LOG.info("IVA Determinación: popup visible (IVA acreditable del periodo title)")
                    break
            except Exception:
                pass
            page.wait_for_timeout(50)
        if dialog is None:
            dialog = page
            LOG.debug("IVA Determinación: using page as popup scope")
        # Popup: *IVA acreditable por actividades gravadas a tasa 16% u 8% y tasa 0% <- Excel "IVA acreditable del periodo"
        iva_acred_raw = label_map.get("IVA acreditable del periodo")
        iva_acred_val = _parse_currency(iva_acred_raw) if iva_acred_raw is not None else 0.0
        iva_acred_str = str(int(round(iva_acred_val))) if iva_acred_val is not None else "0"
        label_grav = "IVA acreditable por actividades gravadas a tasa 16% u 8% y tasa 0%"
        LOG.info("IVA Determinación popup: filling '%s' with %s (from Excel 'IVA acreditable del periodo')", label_grav, iva_acred_str)
        filled_grav = _fill_input_next_to_label(dialog, page, label_grav, iva_acred_str)
        if not filled_grav:
            filled_grav = _fill_input_next_to_label(dialog, page, "*" + label_grav, iva_acred_str)
        if not filled_grav:
            filled_grav = _fill_iva_popup_input_by_position(dialog, page, 1, iva_acred_str)
        if filled_grav:
            LOG.info("IVA Determinación popup: filled '%s'", label_grav)
        else:
            LOG.warning("IVA Determinación popup: could not fill '%s'", label_grav)
        page.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS)
        # *IVA acreditable por actividades mixtas = 0
        label_mixtas = "IVA acreditable por actividades mixtas"
        LOG.info("IVA Determinación popup: filling '%s' with 0", label_mixtas)
        filled_mixtas = _fill_input_next_to_label(dialog, page, label_mixtas, "0")
        if not filled_mixtas:
            filled_mixtas = _fill_input_next_to_label(dialog, page, "*" + label_mixtas, "0")
        if not filled_mixtas:
            filled_mixtas = _fill_iva_popup_input_by_position(dialog, page, 2, "0")
        if filled_mixtas:
            LOG.info("IVA Determinación popup: filled '%s' with 0", label_mixtas)
        else:
            LOG.warning("IVA Determinación popup: could not fill '%s'", label_mixtas)
        page.wait_for_timeout(_TEXTBOX_FILL_WAIT_MS)
        LOG.info("IVA Determinación popup: clicking CERRAR")
        page.get_by_role("button", name=re.compile(r"CERRAR", re.I)).first.click(timeout=1500)
        LOG.info("IVA Determinación popup: CERRAR clicked; popup closed")
        page.wait_for_timeout(200)  # Let modal fully close so main form GUARDAR is the first match

    # GUARDAR main form (fail fast: 2s per attempt so we don't burn ~1 min on wrong/modal GUARDAR)
    LOG.info("IVA Determinación: clicking GUARDAR (main form)")
    guardar_ok = False
    _guardar_wait_ms = 2000
    _guardar_click_ms = 2000
    for loc in [
        page.get_by_role("button", name=re.compile(r"GUARDAR", re.I)),
        page.locator("input[type='submit'][value*='GUARDAR'], input[type='button'][value*='GUARDAR']"),
        page.get_by_text("GUARDAR", exact=True),
    ]:
        try:
            first_btn = loc.first
            first_btn.wait_for(state="visible", timeout=_guardar_wait_ms)
            # Skip if inside modal (cap check at 500ms so we don't hang)
            try:
                if first_btn.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").first.is_visible(timeout=500):
                    continue
            except Exception:
                pass
            if first_btn.get_attribute("disabled"):
                continue
            first_btn.click(timeout=_guardar_click_ms)
            guardar_ok = True
            LOG.info("IVA Determinación: GUARDAR clicked")
            break
        except Exception:
            continue
    if not guardar_ok:
        LOG.warning("IVA Determinación: could not click GUARDAR")
    else:
        page.wait_for_timeout(400)
        page.wait_for_load_state("domcontentloaded", timeout=4000)
        page.wait_for_timeout(200)
        LOG.info("IVA Determinación: waiting for load after GUARDAR complete")
    # Click Pago tab (next to Determinación)
    LOG.info("IVA: step — clicking Pago tab (next to Determinación)")
    pago_clicked = False
    def _try_click_pago(elem, *, force: bool = False) -> bool:
        try:
            elem.wait_for(state="visible", timeout=1200)
            elem.scroll_into_view_if_needed(timeout=1000)
            if force:
                elem.click(force=True, timeout=2000)
            else:
                elem.click(timeout=2000)
            page.wait_for_timeout(200)
            return True
        except Exception:
            return False
    for tab_loc in [page.get_by_role("tab", name=re.compile(r"Pago", re.I)), page.get_by_text("Pago", exact=True)]:
        try:
            if tab_loc.count() > 0 and _try_click_pago(tab_loc.first):
                pago_clicked = True
                break
        except Exception:
            continue
    if not pago_clicked:
        for elem in page.get_by_text("Pago", exact=True).all():
            try:
                if elem.is_visible() and elem.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").count() == 0:
                    if _try_click_pago(elem):
                        pago_clicked = True
                        break
            except Exception:
                continue
    if pago_clicked:
        LOG.info("IVA: Pago tab clicked")
        page.wait_for_timeout(300)
        page.wait_for_load_state("domcontentloaded", timeout=4000)
        page.wait_for_timeout(200)
        LOG.info("IVA Pago: selecting *¿Tienes compensaciones por aplicar? → No (reuse ISR dropdown logic)")
        comp_ok = _fill_pago_custom_dropdown(page, "compensaciones por aplicar", "No")
        if comp_ok:
            LOG.info("IVA Pago: dropdown *¿Tienes compensaciones por aplicar? set to No")
        else:
            LOG.warning("IVA Pago: could not set *¿Tienes compensaciones por aplicar? to No")
        page.wait_for_timeout(60)
        LOG.info("IVA Pago: selecting *¿Tienes estímulos por aplicar? → No (reuse ISR dropdown logic)")
        estim_ok = _fill_pago_custom_dropdown(page, "estímulos por aplicar", "No")
        if estim_ok:
            LOG.info("IVA Pago: dropdown *¿Tienes estímulos por aplicar? set to No")
        else:
            LOG.warning("IVA Pago: could not set *¿Tienes estímulos por aplicar? to No")
        page.wait_for_timeout(80)
        LOG.info("IVA Pago: clicking GUARDAR")
        guardar_pago_ok = False
        for loc in [
            page.get_by_role("button", name=re.compile(r"GUARDAR", re.I)),
            page.locator("input[type='submit'][value*='GUARDAR'], input[type='button'][value*='GUARDAR']"),
            page.get_by_text("GUARDAR", exact=True),
        ]:
            try:
                first_btn = loc.first
                first_btn.wait_for(state="visible", timeout=2500)
                if first_btn.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").count() > 0:
                    continue
                if first_btn.get_attribute("disabled"):
                    continue
                first_btn.click(timeout=2000)
                guardar_pago_ok = True
                LOG.info("IVA Pago: GUARDAR clicked")
                break
            except Exception:
                continue
        if guardar_pago_ok:
            page.wait_for_timeout(300)
            page.wait_for_load_state("domcontentloaded", timeout=4000)
            page.wait_for_timeout(150)
            LOG.info("IVA Pago: load complete after GUARDAR; IVA simplificado flow complete (then logout/end script)")
        else:
            LOG.warning("IVA Pago: could not click GUARDAR")
    else:
        LOG.warning("IVA: could not click Pago tab")
    LOG.info("===== IVA simplificado de confianza: Determinación form complete =====")
    LOG.info("")


def _fill_pago_custom_dropdown(page: Page, label_substring: str, option_text: str) -> bool:
    """Handle Pago section dropdowns: try native <select> after label, then custom trigger (combobox / 'Selecciona') + click option. Returns True if option was set/clicked."""
    try:
        # 1) get_by_label: control may be associated by label text or aria-label
        try:
            control = page.get_by_label(re.compile(re.escape(label_substring), re.I)).first
            control.wait_for(state="visible", timeout=700)
            tag = control.evaluate("el => el.tagName")
            if tag and str(tag).upper() == "SELECT":
                control.select_option(label=option_text, timeout=1500)
                LOG.info("Phase 5: set dropdown (get_by_label + select_option) for %r → %s", label_substring, option_text)
                return True
            control.click(timeout=1000)
            page.wait_for_timeout(250)
            page.get_by_role("option", name=re.compile(re.escape(option_text), re.I)).first.click(timeout=1200)
            return True
        except Exception:
            pass

        label_el = None
        for el in page.get_by_text(re.compile(re.escape(label_substring), re.I)).all():
            try:
                if el.is_visible(timeout=800):
                    if el.locator("xpath=ancestor::*[contains(@class,'modal') or @role='dialog']").count() > 0:
                        continue
                    label_el = el
                    break
            except Exception:
                continue
        if label_el is None:
            return False
        label_el.scroll_into_view_if_needed(timeout=800)
        page.wait_for_timeout(60)

        # 2) First <select> that follows the label in DOM (same row or next cell)
        for xpath_select in [
            "xpath=following::select[1]",
            "xpath=ancestor::tr[1]//select",
            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*//select",
        ]:
            try:
                sel = label_el.locator(xpath_select).first
                sel.wait_for(state="visible", timeout=400)
                sel.select_option(label=option_text, timeout=1500)
                LOG.info("Phase 5: set dropdown (select after label) for %r → %s", label_substring, option_text)
                return True
            except Exception:
                continue

        # 3) Custom dropdown: find trigger (combobox, 'Selecciona' box, or immediate next sibling)
        trigger = None
        for xpath in [
            "xpath=ancestor::tr[1]//*[@role='combobox']",
            "xpath=ancestor::tr[1]//*[contains(@class,'select') or contains(@class,'dropdown')]",
            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*[1]",
            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*[1]//*[@role='combobox']",
            "xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*[1]//*[contains(., 'Selecciona')]",
            "xpath=ancestor::tr[1]//*[contains(., 'Selecciona')]",
            "xpath=following-sibling::*[1]",
            "xpath=following-sibling::*[1]//*[@role='combobox']",
            "xpath=following-sibling::*[1]//*[contains(., 'Selecciona')]",
            "xpath=..//*[@role='combobox']",
            "xpath=..//*[contains(., 'Selecciona')]",
        ]:
            try:
                loc = label_el.locator(xpath)
                if loc.count() == 0:
                    continue
                for i in range(loc.count()):
                    el = loc.nth(i)
                    if el.is_visible(timeout=600):
                        trigger = el
                        break
                if trigger is not None:
                    break
            except Exception:
                continue
        if trigger is None:
            return False
        trigger.click(timeout=1500)
        page.wait_for_timeout(250)
        option_clicked = False
        # Prefer option inside listbox (opened dropdown) to avoid clicking another "No" on the page
        for listbox_selector in ["[role='listbox']", "[role='menu']", ".dropdown-menu", "[class*='listbox']", "[class*='dropdown']"]:
            try:
                listbox = page.locator(listbox_selector).first
                listbox.wait_for(state="visible", timeout=500)
                opt = listbox.get_by_role("option", name=re.compile(re.escape(option_text), re.I)).first
                opt.wait_for(state="visible", timeout=500)
                opt.click(timeout=1200)
                option_clicked = True
                break
            except Exception:
                continue
        if not option_clicked:
            for opt_loc in [
                page.get_by_role("option", name=re.compile(re.escape(option_text), re.I)),
                page.get_by_text(option_text, exact=True),
                page.locator("[role='option']").filter(has_text=re.compile(re.escape(option_text), re.I)),
                page.locator("li, div[role='option'], [class*='option']").filter(has_text=re.compile(r"^" + re.escape(option_text) + r"$", re.I)),
            ]:
                try:
                    first_opt = opt_loc.first
                    first_opt.wait_for(state="visible", timeout=900)
                    first_opt.click(timeout=1200)
                    option_clicked = True
                    break
                except Exception:
                    continue
        page.wait_for_timeout(100)
        return option_clicked
    except Exception as e:
        LOG.debug("Phase 5 custom dropdown (%s → %s) failed: %s", label_substring, option_text, e)
        return False


def logout_sat(page: Page, mapping: dict) -> None:
    """Click 'Cerrar' (next to Inicio) in the SAT nav bar to log out. Safe to call if not logged in or element missing."""
    try:
        if _try_click(page, mapping, "_nav_cerrar"):
            LOG.info("Logged out from SAT (Cerrar clicked)")
            page.wait_for_timeout(1500)  # Let logout navigation complete before caller may close browser
        else:
            # Fallback: cap wait so we don't block 30s (Playwright default) if Cerrar is missing
            cerrar = page.get_by_role("link", name=re.compile(r"Cerrar", re.I)).first
            cerrar.wait_for(state="visible", timeout=5000)
            cerrar.click(timeout=2000)
            page.wait_for_timeout(1500)  # Let logout navigation complete before caller may close browser
            LOG.info("Logged out from SAT (Cerrar clicked via fallback)")
    except Exception as e:
        LOG.debug("Logout (Cerrar) skipped or failed: %s", e)


def _debug_ts() -> str:
    """Timestamp for debug prints (same format as login_sat)."""
    now = datetime.now()
    return now.strftime("%Y-%m-%d %H:%M:%S") + f",{now.microsecond // 1000:03d}"


def _try_fill_select_by_index(scope: Page | Frame, index: int, value_str: str) -> bool:
    """Try to set a <select> by its index in scope (0-based). Uses value then label. Returns True if successful."""
    try:
        loc = scope.locator("select")
        sel = loc.nth(index)
        sel.wait_for(state="visible", timeout=4000)
        try:
            sel.select_option(value=value_str)
            return True
        except Exception:
            sel.select_option(label=value_str)
            return True
    except Exception:
        return False


def _fill_select_by_mapping(
    scope: Page | Frame, page_for_wait: Page, selector_list: list, value_str: str
) -> bool:
    """Try each selector in the list (e.g. from form_field_mapping) and set the select by value then label. Returns True if any succeeds."""
    if not selector_list:
        return False
    for sel_str in selector_list:
        try:
            dropdown = scope.locator(sel_str).first
            dropdown.wait_for(state="visible", timeout=1200)
            try:
                dropdown.select_option(value=value_str, timeout=2000)
                return True
            except Exception:
                dropdown.select_option(label=value_str, timeout=2000)
                return True
        except Exception:
            continue
    return False


def _fill_select_next_to_label(
    scope: Page | Frame,
    page_for_wait: Page,
    label_text: str,
    value_str: str,
    mapping: dict | None = None,
    initial_dropdown_key: str | None = None,
    log_prefix: str = "",
) -> bool:
    """Locate dropdown (by label or mapping), press dropdown, scroll if needed, select option matching Excel value (fallback to select_option for native <select>)."""
    scope_type = "iframe" if isinstance(scope, Frame) else "page"
    if log_prefix:
        LOG.info("%sLabel=%r value=%r scope=%s", log_prefix, label_text, value_str, scope_type)
    elif LOG.isEnabledFor(logging.DEBUG):
        LOG.debug("Label=%r value=%r scope=%s", label_text, value_str, scope_type)

    _SCROLL_TIMEOUT_MS = 120
    _OPTION_CLICK_TIMEOUT_MS = 220

    def do_press_dropdown_then_click_option(dropdown) -> bool:
        """Try fast path (select_option) first; else open dropdown and click option."""
        try:
            dropdown.wait_for(state="visible", timeout=300)
            try:
                dropdown.scroll_into_view_if_needed(timeout=_SCROLL_TIMEOUT_MS)
            except Exception:
                pass
            try:
                dropdown.select_option(value=value_str, timeout=280)
                return True
            except Exception:
                pass
            try:
                dropdown.select_option(label=value_str, timeout=280)
                return True
            except Exception:
                pass
            page_for_wait.wait_for_timeout(80)
            dropdown.click(timeout=250)
            page_for_wait.wait_for_timeout(80)
            opt_by_value = dropdown.locator(f"option[value={repr(value_str)}]")
            if opt_by_value.count() > 0:
                option = opt_by_value.first
            else:
                option = dropdown.locator("option").filter(has_text=re.compile(re.escape(value_str), re.I)).first
            option.wait_for(state="attached", timeout=200)
            try:
                option.click(timeout=_OPTION_CLICK_TIMEOUT_MS)
                return True
            except Exception:
                try:
                    dropdown.select_option(value=value_str, timeout=350)
                    return True
                except Exception:
                    dropdown.select_option(label=value_str, timeout=350)
                    return True
        except Exception as e2:
            if log_prefix:
                LOG.info("%s  Strategy 1 (press dropdown + select) failed: %s", log_prefix, e2)
            elif LOG.isEnabledFor(logging.DEBUG):
                LOG.debug("Strategy 1 (press dropdown + select) failed: %s", e2)
            return False

    def resolve_dropdown_from_label(label_el):
        """Return a visible <select> near the label (prefer dropdown to the right of label; skip hidden selects)."""
        def first_visible_select(loc):
            if loc.count() == 0:
                return None
            for i in range(loc.count()):
                try:
                    el = loc.nth(i)
                    el.wait_for(state="visible", timeout=200)
                    return el
                except Exception:
                    continue
            return None
        # 1) Prefer select in the next cell to the right (dropdown to the right of label)
        sel = label_el.locator("xpath=((ancestor::td | ancestor::th)[1])/following-sibling::*[1]//select")
        out = first_visible_select(sel)
        if out is not None:
            return out
        # 2) Select in same row (tr); pick first visible (phase-2 style)
        sel = label_el.locator("xpath=ancestor::tr[1]//select")
        out = first_visible_select(sel)
        if out is not None:
            return out
        # 3) Same cell
        sel = label_el.locator("xpath=((ancestor::td | ancestor::th)[1])//select")
        out = first_visible_select(sel)
        if out is not None:
            return out
        # 4) Ancestor with single select
        sel = label_el.locator("xpath=(ancestor::*[.//select and count(descendant::select)=1])[1]//select")
        out = first_visible_select(sel)
        if out is not None:
            return out
        sel = label_el.locator("xpath=..").locator("select")
        out = first_visible_select(sel)
        if out is not None:
            return out
        if sel.count() > 0:
            return sel.first
        return None

    # Locate dropdown: try mapping first when provided (fast id/selectors); else resolve by label (slow xpath).
    dropdown = None
    if initial_dropdown_key and mapping and mapping.get(initial_dropdown_key):
        sel_list = mapping[initial_dropdown_key] if isinstance(mapping[initial_dropdown_key], list) else [mapping[initial_dropdown_key]]
        for sel_str in sel_list:
            try:
                loc = scope.locator(sel_str).first
                loc.wait_for(state="visible", timeout=300)
                dropdown = loc
                break
            except Exception:
                continue
    if dropdown is None:
        try:
            xpath_contains_arg = "'" + label_text.replace("'", "''") + "'"
            xpath_label = (
                "//*[(self::label or self::td or self::th or self::span or self::div)"
                " and not(ancestor::select) and contains(., " + xpath_contains_arg + ")]"
            )
            label_el = scope.locator("xpath=" + xpath_label).last
            label_el.wait_for(state="attached", timeout=350)
            dropdown = resolve_dropdown_from_label(label_el)
        except Exception:
            pass
    try:
        if dropdown is not None:
            try:
                el_id = dropdown.get_attribute("id") or "(no id)"
                if log_prefix:
                    LOG.info("%s  Strategy 1 (press dropdown, scroll, select option): dropdown id=%r", log_prefix, el_id)
                elif LOG.isEnabledFor(logging.DEBUG):
                    LOG.debug("Strategy 1: dropdown id=%r", el_id)
            except Exception:
                pass
            if do_press_dropdown_then_click_option(dropdown):
                if log_prefix:
                    LOG.info("%s  Strategy 1: filled OK", log_prefix)
                elif LOG.isEnabledFor(logging.DEBUG):
                    LOG.debug("Strategy 1: filled OK")
                return True
    except Exception as e:
        if log_prefix:
            LOG.info("%s  Strategy 1 exception: %s", log_prefix, e)
        elif LOG.isEnabledFor(logging.DEBUG):
            LOG.debug("Strategy 1 exception: %s", e)

    if log_prefix:
        LOG.info("%s  Result: NOT filled for label=%r", log_prefix, label_text)
    elif LOG.isEnabledFor(logging.DEBUG):
        LOG.debug("Result: NOT filled for label=%r", label_text)
    return False


def _get_declaration_form_scope(page: Page) -> Page | Frame:
    """Return the page or the frame that contains the declaration form selects. SAT often loads the form in an iframe."""
    probe = "select[id*='TipoDeclaracion'], select[id*='EjercicioFiscal']"
    try:
        loc = page.locator(probe)
        if loc.count() > 0:
            return page
    except Exception:
        pass
    for frame in page.frames:
        if frame == page.main_frame:
            continue
        try:
            loc = frame.locator(probe)
            if loc.count() > 0:
                LOG.info("Phase 2: Declaration form found in iframe")
                return frame
        except Exception:
            continue
    return page


def fill_initial_form(page: Page, data: dict, mapping: dict) -> None:
    """Fill Configuración de la declaración in order: Ejercicio (2022–2026) → Periodicidad → Periodo (Enero–Diciembre, YTD) → Tipo de declaración (Normal / Normal por Corrección Fiscal). Each dropdown may appear after the previous is selected."""
    scope = _get_declaration_form_scope(page)
    page_for_wait = scope.page if isinstance(scope, Frame) else scope
    year = data.get("year")
    month = data.get("month")
    periodicidad = data.get("periodicidad", 1)
    tipo = data.get("tipo_declaracion", "Normal")
    try:
        p = int(periodicidad) if periodicidad is not None else 1
    except (TypeError, ValueError):
        p = 1
    periodicidad_value = _SAT_PERIODICIDAD_VALUE.get(p, "M")
    periodo_str = f"{month:02d}" if month is not None else "N/A"
    LOG.info("Phase 2: Scope: %s. Will fill in order: Ejercicio=%s, Periodicidad=%s (value %r), Periodo=%s, Tipo=%r", 'iframe' if isinstance(scope, Frame) else 'main page', year, periodicidad, periodicidad_value, periodo_str, tipo)
    # Wait for at least one select (Ejercicio) to be visible.
    try:
        loc = scope.locator("select")
        loc.first.wait_for(state="visible", timeout=12000)
        n_selects = loc.count()
        LOG.info("Phase 2: Selects in scope: %s", n_selects)
    except Exception as e:
        LOG.info("Phase 2: Wait for select failed: %s", e)
    page_for_wait.wait_for_timeout(60)
    # Order per updated SAT form: Ejercicio → Periodicidad → Periodo (appears after Periodicidad) → Tipo de declaración (appears after Periodo).
    if year is not None:
        LOG.info("Phase 2: --- Filling Ejercicio ---")
        ok = _fill_select_next_to_label(scope, page_for_wait, "Ejercicio", str(year), mapping=mapping, initial_dropdown_key="initial_ejercicio", log_prefix="Phase 2: ")
        if not ok and mapping.get("initial_ejercicio"):
            sel_list = mapping["initial_ejercicio"] if isinstance(mapping["initial_ejercicio"], list) else [mapping["initial_ejercicio"]]
            ok = _fill_select_by_mapping(scope, page_for_wait, sel_list, str(year))
            if ok:
                LOG.info("Phase 2: Ejercicio filled via mapping selectors")
        LOG.info("Phase 2: Ejercicio: %s%s", year, " (filled)" if ok else " (NOT filled — check selectors)")
        page_for_wait.wait_for_timeout(80)
    LOG.info("Phase 2: --- Filling Periodicidad ---")
    ok = _fill_select_next_to_label(scope, page_for_wait, "Periodicidad", periodicidad_value, mapping=mapping, initial_dropdown_key="initial_periodicidad", log_prefix="Phase 2: ")
    LOG.info("Phase 2: Periodicidad: %s%s", periodicidad, " (filled)" if ok else " (NOT filled — check selectors)")
    page_for_wait.wait_for_timeout(250)
    # Periodo dropdown appears after Periodicidad (Enero–Diciembre; SAT may show only YTD months). pstcdypisr uses label "Periodo" (no accent).
    if month is not None:
        LOG.info("Phase 2: --- Filling Periodo ---")
        periodo_value = _SAT_PERIODO_LABEL.get(month, "Enero")
        ok = _fill_select_next_to_label(scope, page_for_wait, "Periodo", periodo_value, mapping=mapping, initial_dropdown_key="initial_periodo", log_prefix="Phase 2: ")
        if not ok:
            ok = _fill_select_next_to_label(scope, page_for_wait, "Período", periodo_value, mapping=mapping, initial_dropdown_key="initial_periodo", log_prefix="Phase 2: ")
        if not ok and mapping.get("initial_periodo"):
            sel_list = mapping["initial_periodo"] if isinstance(mapping["initial_periodo"], list) else [mapping["initial_periodo"]]
            ok = _fill_select_by_mapping(scope, page_for_wait, sel_list, periodo_value)
            if ok:
                LOG.info("Phase 2: Periodo filled via mapping selectors")
        LOG.info("Phase 2: Periodo: %02d (%s)%s", month, periodo_value, " (filled)" if ok else " (NOT filled — check selectors)")
        page_for_wait.wait_for_timeout(200)
    # Tipo de declaración appears after Periodo (Normal / Normal por Corrección Fiscal). pstcdypisr uses label "Tipo de declaración" (lowercase d).
    LOG.info("Phase 2: --- Filling Tipo de Declaración ---")
    ok = _fill_select_next_to_label(scope, page_for_wait, "Tipo de declaración", str(tipo), mapping=mapping, initial_dropdown_key="initial_tipo_declaracion", log_prefix="Phase 2: ")
    if not ok:
        ok = _fill_select_next_to_label(scope, page_for_wait, "Tipo de Declaración", str(tipo), mapping=mapping, initial_dropdown_key="initial_tipo_declaracion", log_prefix="Phase 2: ")
    if not ok and mapping.get("initial_tipo_declaracion"):
        sel_list = mapping["initial_tipo_declaracion"] if isinstance(mapping["initial_tipo_declaracion"], list) else [mapping["initial_tipo_declaracion"]]
        ok = _fill_select_by_mapping(scope, page_for_wait, sel_list, str(tipo))
    LOG.info("Phase 2: Tipo de Declaración: %s%s", tipo, " (filled)" if ok else " (NOT filled — check selectors)")
    page_for_wait.wait_for_timeout(40)


def fill_obligation_section(page, mapping: dict, label_map: dict, labels: list[str]) -> None:
    """Fill form fields for given Excel labels (try each label's selectors with value from label_map).
    For ISR simplificado de confianza, the Ingresos form should match the field list and order in
    FILL THE FORM ON SAT.pdf (e.g. pages 25-42). Add or reorder labels/mappings if the PDF differs."""
    for label in labels:
        value = label_map.get(label)
        if value is None:
            continue
        if isinstance(value, float) and value == 0.0:
            continue
        _try_fill(page, page, mapping, label, value)
        page.wait_for_timeout(_OBLIGATION_CLICK_WAIT_MS)


def check_totals(page, data: dict, mapping: dict, tolerance: int) -> tuple[bool, str]:
    """
    Compare SAT summary (ISR a pagar, IVA a pagar, Total a pagar) with Excel.
    Tolerance ±tolerance pesos each. Returns (ok, message).
    """
    label_map = data["label_map"]
    excel_isr = _parse_currency(label_map.get("ISR a cargo"))
    excel_iva = _parse_currency(label_map.get("IVA a cargo"))
    excel_total = excel_isr + excel_iva

    # Read displayed totals from SAT (selectors are placeholders; implementation reads text)
    def _read_summary(selector_key: str) -> float:
        selectors = mapping.get(selector_key)
        if not selectors:
            return 0.0
        for sel in selectors:
            try:
                loc = page.locator(sel)
                if loc.count() == 0:
                    continue
                text = loc.first.text_content(timeout=2000) or ""
                # Extract number from text like "$ 95" or "1,480"
                n = re.sub(r"[^\d.]", "", text.replace(",", ""))
                if n:
                    return float(n)
            except Exception:
                continue
        return 0.0

    sat_isr = _read_summary("_summary_isr_a_pagar")
    sat_iva = _read_summary("_summary_iva_a_pagar")
    sat_total = _read_summary("_summary_total_a_pagar")
    if sat_total == 0.0 and (sat_isr != 0.0 or sat_iva != 0.0):
        sat_total = sat_isr + sat_iva

    ok_isr = abs(sat_isr - excel_isr) <= tolerance
    ok_iva = abs(sat_iva - excel_iva) <= tolerance
    ok_total = abs(sat_total - excel_total) <= tolerance
    if ok_isr and ok_iva and ok_total:
        return True, f"Totals OK: ISR {sat_isr}, IVA {sat_iva}, Total {sat_total}"
    msg = (
        f"Totals mismatch (tolerance ±{tolerance}): "
        f"Excel ISR={excel_isr} IVA={excel_iva} Total={excel_total}; "
        f"SAT ISR={sat_isr} IVA={sat_iva} Total={sat_total}"
    )
    return False, msg


def send_declaration(page, mapping: dict) -> bool:
    """Click Enviar declaración."""
    return _try_click(page, mapping, "_btn_enviar_declaracion")


def get_efirma_from_config(config: dict) -> dict:
    """
    Get e.firma from config (test mode: local .cer/.key paths and password, no DB).
    Accepts either root keys (test_cer_path, test_key_path, test_password) or nested config["test"] (cer_path, key_path, password).
    """
    t = config.get("test") or {}
    cer = config.get("test_cer_path") or t.get("cer_path") or ""
    key = config.get("test_key_path") or t.get("key_path") or ""
    pwd = config.get("test_password") or t.get("password") or ""
    rfc = config.get("test_rfc") or t.get("rfc") or ""
    if not cer or not key:
        raise ValueError(
            "Test mode requires test_cer_path and test_key_path in config.json (or test.cer_path, test.key_path). "
            "Add test_password (or test.password) for e.firma."
        )
    cer_path = os.path.abspath(cer)
    key_path = os.path.abspath(key)
    if not os.path.isfile(cer_path):
        raise FileNotFoundError(f"test cer path not found: {cer_path}")
    if not os.path.isfile(key_path):
        raise FileNotFoundError(f"test key path not found: {key_path}")
    return {
        "cer_path": cer_path,
        "key_path": key_path,
        "password": pwd,
        "rfc": rfc,
    }


def run(
    workbook_path: str | None = None,
    company_id: int | None = None,
    branch_id: int | None = None,
    config_path: str | None = None,
    mapping_path: str | None = None,
    test_login: bool = False,
    test_initial_form: bool = False,
    test_full: bool = False,
    test_phase3: bool = False,
    test_iva: bool = False,
    use_api_efirma: bool = False,
    efirma_from_api: dict | None = None,
    company_id_api: int | None = None,
    branch_id_api: int | None = None,
    declaration_id_api: int | None = None,
    api_base_url: str | None = None,
    pdf_download_callback: Callable[[str], None] | None = None,
    pdf_download_dir: str | None = None,
) -> bool:
    """
    Full flow: read Excel → get e.firma → login SAT → navigate → fill initial → fill ISR → fill IVA → check totals → send.
    If test_login=True: only open SAT and perform e.firma login (no DB; use test_* in config). Returns True if login succeeded.
    If test_initial_form=True: login then fill Declaración Provisional initial form (no Excel; use test_year, test_month, test_periodicidad in config).
    If test_full=True: login + fill initial form + phase 3 (ISR Ingresos) + phase 4 (Determinación, ISR retenido VER DETALLE) + phase 5 (Pago) + IVA simplificado de confianza (ADMINISTRACIÓN → IVA → Determinación → Pago tab). e.firma and data from config/Excel. No send.
    If test_phase3=True: login + fill initial form + select ISR simplificado de confianza + fill ISR section (Ingresos, etc.). Stops after ISR fill; no IVA/send. Requires --workbook.
    If test_iva=True: login + fill initial form + transition to phase 3 (SIGUIENTE, CERRAR) + select IVA simplificado de confianza + fill IVA Determinación form. No ISR. Requires --workbook.
    If use_api_efirma=True and efirma_from_api: use efirma_from_api for login (no DB/config). Optional pdf_download_callback called with local PDF path after send (API mode).
    Returns True if declaration was sent (or test step OK), False otherwise. Logs and prints outcome.
    """
    config = load_config(config_path)
    mapping = load_mapping(mapping_path)
    setup_logging(config.get("log_file"))

    def _on_sigint(_signum, _frame):
        # Do not call Playwright from here (can block). Raise KeyboardInterrupt so main thread
        # unwinds and the finally block runs (logout_sat, then close browser).
        signal.signal(signal.SIGINT, signal.SIG_DFL)  # second Ctrl+C kills immediately
        print("Ctrl+C: cleaning up (logout, close browser)...", file=sys.stderr)
        raise KeyboardInterrupt

    signal.signal(signal.SIGINT, _on_sigint)
    base_url = config.get("sat_portal_url", SAT_PORTAL_URL)

    global _run_context
    if test_login:
        if use_api_efirma and efirma_from_api:
            LOG.info("Test mode: login only (e.firma from API: downloaded CER/KEY, password from response)")
            efirma = efirma_from_api
        else:
            LOG.info("Test mode: login only (no DB, using test_cer_path / test_key_path / test_password from config)")
            efirma = get_efirma_from_config(config)
        sat_ui = get_sat_ui(config)
        for attempt in range(2):
            try:
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=False)
                    context = browser.new_context(accept_downloads=True)
                    page = context.new_page()
                    _run_context = {"page": page, "mapping": mapping}
                    try:
                        login_sat(page, efirma, mapping, base_url, sat_ui)
                        LOG.info("Test login: e.firma login completed. Browser will stay open 10s for inspection.")
                        page.wait_for_timeout(10000)
                        return True
                    finally:
                        _run_context = None
                        time.sleep(0.5)  # Let Playwright pending ops settle before close (avoids greenlet thread errors)
                        LOG.info("Cleanup: logging out from SAT, then closing browser.")
                        try:
                            logout_sat(page, mapping)
                            page.wait_for_timeout(2000)  # Let Cerrar click and navigation complete before closing
                        except Exception as e:
                            LOG.warning("Test cleanup logout or wait: %s", e)
                        try:
                            page.close()
                        except Exception as e:
                            LOG.debug("Page close: %s", e)
                        try:
                            context.close()
                        except Exception as e:
                            LOG.debug("Context close: %s", e)
                        try:
                            browser.close()
                        except Exception as e:
                            LOG.debug("Browser close: %s", e)
            except Exception as e:
                LOG.exception("Test login failed")
                print(str(e), file=sys.stderr)
                if attempt == 0:
                    LOG.info("Closing and retrying once in %s seconds...", RETRY_WAIT_SECONDS)
                    time.sleep(RETRY_WAIT_SECONDS)
                else:
                    return False
        return False

    if test_initial_form:
        LOG.info("Test mode: login + fill initial Declaración Provisional form (no Excel/DB; using test_* from config)")
        efirma = get_efirma_from_config(config)
        t = config.get("test") or {}
        year = config.get("test_year") or t.get("year")
        month = config.get("test_month") or t.get("month")
        periodicidad = config.get("test_periodicidad") or t.get("periodicidad") or 1
        if year is None:
            year = datetime.now().year
        if month is None:
            month = 1
        data = {"year": int(year), "month": int(month), "periodicidad": int(periodicidad), "tipo_declaracion": "Normal", "label_map": {}}
        LOG.info("Initial form test data: year=%s month=%s periodicidad=%s", data["year"], data["month"], data["periodicidad"])
        for attempt in range(2):
            try:
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=False)
                    context = browser.new_context(accept_downloads=True)
                    page = context.new_page()
                    _run_context = {"page": page, "mapping": mapping}
                    try:
                        sat_ui = get_sat_ui(config)
                        login_sat(page, efirma, mapping, base_url, sat_ui)
                        print()
                        LOG.info("\n")
                        LOG.info("Opening Configuración de la declaración (Presentar declaración)")
                        if not open_configuration_form(page, mapping):
                            LOG.warning("Could not click Presentar declaración; continuing to fill initial form.")
                        dismiss_draft_if_present(page, mapping, sat_ui)
                        fill_initial_form(page, data, mapping)
                        LOG.info("Test initial form: initial form step complete. Browser will stay open 10s for inspection.")
                        page.wait_for_timeout(10000)
                        return True
                    finally:
                        _run_context = None
                        time.sleep(0.5)  # Let Playwright pending ops settle before close (avoids greenlet thread errors)
                        try:
                            logout_sat(page, mapping)
                        except Exception as e:
                            LOG.debug("Test cleanup logout: %s", e)
                        try:
                            page.close()
                        except Exception as e:
                            LOG.debug("Page close: %s", e)
                        try:
                            context.close()
                        except Exception as e:
                            LOG.debug("Context close: %s", e)
                        try:
                            browser.close()
                        except Exception as e:
                            LOG.debug("Browser close: %s", e)
            except Exception as e:
                LOG.exception("Test initial form failed")
                print(str(e), file=sys.stderr)
                if attempt == 0:
                    LOG.info("Closing and retrying once in %s seconds...", RETRY_WAIT_SECONDS)
                    time.sleep(RETRY_WAIT_SECONDS)
                else:
                    return False
        return False

    # Full run in test mode: Excel for initial form + phase 3 (SIGUIENTE, CERRAR, ISR selection, ISR Ingresos fill). Config for e.firma (no DB). No IVA/send.
    if test_full:
        if not workbook_path:
            raise ValueError("Test full run requires --workbook")
        LOG.info("Test full run: e.firma from config (no DB); initial form from Excel; then phase 3 (ISR Ingresos) + phase 4 (Determinación, ISR retenido) + phase 5 (Pago); no IVA/send")
        data = read_impuestos(workbook_path)
        data["workbook_path"] = workbook_path
        label_map = data["label_map"]
        LOG.info("Period: %s-%s, periodicidad: %s", data.get("year"), data.get("month"), data.get("periodicidad"))
        efirma = efirma_from_api if (use_api_efirma and efirma_from_api) else get_efirma_from_config(config)
        sat_ui = get_sat_ui(config)
        isr_labels = get_isr_determinacion_labels(config)
        run_success = False
        for attempt in range(2):
            try:
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=False)
                    context = browser.new_context(accept_downloads=True)
                    page = context.new_page()
                    _run_context = {"page": page, "mapping": mapping}
                    try:
                        login_sat(page, efirma, mapping, base_url, sat_ui)
                        LOG.info("Phase 1: Logged in to SAT")
                        print()
                        LOG.info("\n")
                        if not open_configuration_form(page, mapping):
                            LOG.warning("Could not click Presentar declaración; continuing to fill initial form.")
                        dismiss_draft_if_present(page, mapping, sat_ui)
                        fill_initial_form(page, data, mapping)
                        page.wait_for_timeout(400)
                        print()
                        LOG.info("\n")
                        LOG.info("Phase 2→3: SIGUIENTE, wait for load, CERRAR pop-up")
                        if not transition_initial_to_phase3(page, mapping, sat_ui):
                            LOG.warning("Transition to phase 3 had issues; continuing.")
                        LOG.info("Phase 3: Selecting ISR simplificado de confianza and filling ISR section")
                        if not open_obligation_isr(page, mapping):
                            LOG.warning("Could not click ISR simplificado de confianza")
                        page.wait_for_timeout(500)
                        fill_isr_ingresos_form(page, mapping, data, sat_ui)
                        fill_obligation_section(page, mapping, label_map, isr_labels)
                        LOG.info("Test full run: ISR complete; running IVA simplificado de confianza (central)")
                        iva_det_fields = get_iva_determinacion_fields(config)
                        fill_iva_simplificado(page, mapping, data, sat_ui, iva_determinacion_fields=iva_det_fields)
                        LOG.info("Test full run: initial form + ISR + IVA Determinación + Pago tab complete. Browser will stay open %ss for inspection.", _TEST_INSPECTION_WAIT_MS // 1000)
                        page.wait_for_timeout(_TEST_INSPECTION_WAIT_MS)
                        run_success = True
                        return True
                    finally:
                        _run_context = None
                        time.sleep(0.5)  # Let Playwright pending ops settle before close (avoids greenlet thread errors)
                        LOG.info("Cleanup: logging out from SAT, then closing browser.")
                        try:
                            page.set_default_timeout(8000)  # Cap cleanup so logout/close don't block 30s+ if element missing
                            logout_sat(page, mapping)
                            # Let Cerrar click and navigation complete before closing the browser.
                            page.wait_for_timeout(2000)
                        except Exception as e:
                            LOG.warning("Test cleanup logout or wait: %s", e)
                        try:
                            page.close()
                        except Exception as e:
                            LOG.debug("Page close: %s", e)
                        try:
                            context.close()
                        except Exception as e:
                            LOG.debug("Context close: %s", e)
                        try:
                            browser.close()
                        except Exception as e:
                            LOG.debug("Browser close: %s", e)
                        # Force process exit so script does not hang (Playwright may leave driver/threads alive)
                        LOG.info("Test full run: closing browser done; exiting process.")
                        os._exit(0 if run_success else 1)
            except Exception as e:
                LOG.exception("Test full run failed")
                print(str(e), file=sys.stderr)
                if attempt == 0:
                    LOG.info("Closing and retrying once in %s seconds...", RETRY_WAIT_SECONDS)
                    time.sleep(RETRY_WAIT_SECONDS)
                else:
                    return False
        return False

    # Test IVA only: login, initial form, transition to phase 3, then IVA simplificado de confianza (no ISR).
    if test_iva:
        if not workbook_path:
            raise ValueError("Test IVA run requires --workbook")
        LOG.info("Test IVA: login + initial form + phase 2→3 (SIGUIENTE, CERRAR) + IVA simplificado de confianza (Determinación); e.firma and data from config/Excel; no ISR/send")
        data = read_impuestos(workbook_path)
        data["workbook_path"] = workbook_path
        LOG.info("Period: %s-%s, periodicidad: %s", data.get("year"), data.get("month"), data.get("periodicidad"))
        efirma = efirma_from_api if (use_api_efirma and efirma_from_api) else get_efirma_from_config(config)
        sat_ui = get_sat_ui(config)
        run_success = False
        for attempt in range(2):
            try:
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=False)
                    context = browser.new_context(accept_downloads=True)
                    page = context.new_page()
                    _run_context = {"page": page, "mapping": mapping}
                    try:
                        login_sat(page, efirma, mapping, base_url, sat_ui)
                        LOG.info("Phase 1: Logged in to SAT")
                        print()
                        LOG.info("\n")
                        if not open_configuration_form(page, mapping):
                            LOG.warning("Could not click Presentar declaración; continuing.")
                        dismiss_draft_if_present(page, mapping, sat_ui)
                        fill_initial_form(page, data, mapping)
                        page.wait_for_timeout(400)
                        print()
                        LOG.info("\n")
                        LOG.info("Phase 2→3: SIGUIENTE, wait for load, CERRAR pop-up")
                        if not transition_initial_to_phase3(page, mapping, sat_ui):
                            LOG.warning("Transition to phase 3 had issues; continuing.")
                        LOG.info("Test IVA: Selecting IVA simplificado de confianza (skip ISR)")
                        if not open_obligation_iva(page, mapping, sat_ui):
                            LOG.warning("Could not click IVA simplificado de confianza")
                        page.wait_for_timeout(1500)
                        fill_iva_simplificado_determinacion(page, mapping, data)
                        LOG.info("Test IVA: login + initial form + IVA Determinación + Pago tab complete. Browser will stay open 10s for inspection.")
                        page.wait_for_timeout(10000)
                        run_success = True
                        return True
                    finally:
                        _run_context = None
                        time.sleep(0.5)
                        LOG.info("Cleanup: logging out from SAT, then closing browser.")
                        try:
                            logout_sat(page, mapping)
                            page.wait_for_timeout(2000)  # Let Cerrar click and navigation complete before closing
                        except Exception as e:
                            LOG.warning("Test cleanup logout or wait: %s", e)
                        try:
                            page.close()
                        except Exception as e:
                            LOG.debug("Page close: %s", e)
                        try:
                            context.close()
                        except Exception as e:
                            LOG.debug("Context close: %s", e)
                        try:
                            browser.close()
                        except Exception as e:
                            LOG.debug("Browser close: %s", e)
            except Exception as e:
                LOG.exception("Test IVA failed")
                print(str(e), file=sys.stderr)
                if attempt == 0:
                    LOG.info("Closing and retrying once in %s seconds...", RETRY_WAIT_SECONDS)
                    time.sleep(RETRY_WAIT_SECONDS)
                else:
                    return False
        return False

    # Phase 3 test: login, initial form, select ISR simplificado de confianza, fill ISR section (pages 24–51); stop for inspection.
    if test_phase3:
        if not workbook_path:
            raise ValueError("Test phase 3 requires --workbook")
        LOG.info("Test phase 3: Phase 1 (login) + Phase 2 (initial form) + Phase 3 (select ISR simplificado, fill ISR section); e.firma from config, data from Excel; no IVA/send")
        data = read_impuestos(workbook_path)
        data["workbook_path"] = workbook_path
        label_map = data["label_map"]
        LOG.info("Period: %s-%s, periodicidad: %s", data.get("year"), data.get("month"), data.get("periodicidad"))
        efirma = efirma_from_api if (use_api_efirma and efirma_from_api) else get_efirma_from_config(config)
        sat_ui = get_sat_ui(config)
        isr_labels = get_isr_determinacion_labels(config)
        for attempt in range(2):
            try:
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=False)
                    context = browser.new_context(accept_downloads=True)
                    page = context.new_page()
                    _run_context = {"page": page, "mapping": mapping}
                    try:
                        LOG.info("Phase 1: Logging in to SAT (e.firma)")
                        login_sat(page, efirma, mapping, base_url, sat_ui)
                        print()
                        LOG.info("\n")
                        LOG.info("Phase 2: Opening Configuración (Presentar declaración), then filling Ejercicio → Periodicidad → Periodo → Tipo")
                        if not open_configuration_form(page, mapping):
                            LOG.warning("Could not click Presentar declaración")
                        dismiss_draft_if_present(page, mapping, sat_ui)
                        fill_initial_form(page, data, mapping)
                        page.wait_for_timeout(400)
                        print()
                        LOG.info("\n")
                        LOG.info("Phase 2→3: Clicking SIGUIENTE, waiting for load, closing pre-fill pop-up (CERRAR)")
                        if not transition_initial_to_phase3(page, mapping, sat_ui):
                            LOG.warning("Transition to phase 3 (SIGUIENTE/CERRAR) had issues; continuing.")
                        LOG.info("Phase 3: Selecting ISR simplificado de confianza and filling ISR section (Ingresos form per PDF pp 25-42)")
                        if not open_obligation_isr(page, mapping):
                            LOG.warning("Could not click ISR simplificado de confianza")
                        page.wait_for_timeout(500)
                        fill_isr_ingresos_form(page, mapping, data, sat_ui)
                        fill_obligation_section(page, mapping, label_map, isr_labels)
                        LOG.info("Test phase 3 complete (Phase 1 login + Phase 2 initial form + Phase 3 ISR section). Browser will stay open 10s for inspection.")
                        page.wait_for_timeout(10000)
                        return True
                    finally:
                        _run_context = None
                        time.sleep(0.5)  # Let Playwright pending ops settle before close (avoids greenlet thread errors)
                        try:
                            logout_sat(page, mapping)
                        except Exception as e:
                            LOG.debug("Test cleanup logout: %s", e)
                        try:
                            page.close()
                        except Exception as e:
                            LOG.debug("Page close: %s", e)
                        try:
                            context.close()
                        except Exception as e:
                            LOG.debug("Context close: %s", e)
                        try:
                            browser.close()
                        except Exception as e:
                            LOG.debug("Browser close: %s", e)
            except Exception as e:
                LOG.exception("Test phase 3 failed")
                print(str(e), file=sys.stderr)
                if attempt == 0:
                    LOG.info("Closing and retrying once in %s seconds...", RETRY_WAIT_SECONDS)
                    time.sleep(RETRY_WAIT_SECONDS)
                else:
                    return False
        return False

    # Normal flow (DB or API e.firma)
    if not workbook_path:
        raise ValueError("Normal run requires --workbook")
    if use_api_efirma and efirma_from_api:
        efirma = efirma_from_api
        LOG.info("Using e.firma from API (downloaded CER/KEY)")
    elif company_id is not None and branch_id is not None:
        LOG.info("Fetching e.firma from DB for company=%s branch=%s", company_id, branch_id)
        efirma = get_efirma_from_db(company_id, branch_id, config)
    else:
        raise ValueError("Normal run requires (--company-id and --branch-id) or API e.firma (use_api_efirma + efirma_from_api)")
    LOG.info("Reading workbook: %s", workbook_path)
    data = read_impuestos(workbook_path)
    data["workbook_path"] = workbook_path
    label_map = data["label_map"]
    LOG.info("Period: %s-%s, labels read: %d", data.get("year"), data.get("month"), len(label_map))
    tolerance = config.get("totals_tolerance_pesos", TOLERANCE_PESOS)
    sat_ui = get_sat_ui(config)
    isr_labels = get_isr_determinacion_labels(config)
    iva_determinacion_fields = get_iva_determinacion_fields(config)
    iva_pago_labels = get_iva_pago_labels(config)

    for attempt in range(2):
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=False)  # headless=False so user can see; set True for automation
                context = browser.new_context(accept_downloads=True)
                page = context.new_page()
                _run_context = {"page": page, "mapping": mapping}
                try:
                    login_sat(page, efirma, mapping, base_url, sat_ui)
                    LOG.info("Phase 1: Logged in to SAT")
                    print()
                    LOG.info("\n")
                    if not open_configuration_form(page, mapping):
                        LOG.warning("Could not click Presentar declaración")
                    dismiss_draft_if_present(page, mapping, sat_ui)
                    fill_initial_form(page, data, mapping)
                    page.wait_for_timeout(400)
                    print()
                    LOG.info("\n")
                    # Phase 2→3: SIGUIENTE, wait for load, click CERRAR on pre-fill pop-up
                    if not transition_initial_to_phase3(page, mapping, sat_ui):
                        LOG.warning("Transition to phase 3 (SIGUIENTE/CERRAR) had issues; continuing.")
                    # Phase 3: Select ISR simplificado de confianza, then fill ISR Ingresos form (per PDF pp 25-42)
                    if not open_obligation_isr(page, mapping):
                        LOG.warning("Could not click ISR simplificado de confianza; continuing to fill ISR fields anyway.")
                    page.wait_for_timeout(500)
                    fill_isr_ingresos_form(page, mapping, data, sat_ui)
                    fill_obligation_section(page, mapping, label_map, isr_labels)
                    page.wait_for_timeout(700)

                    # IVA simplificado de confianza (central: ADMINISTRACIÓN → IVA → Determinación → Pago tab)
                    fill_iva_simplificado(page, mapping, data, sat_ui, iva_determinacion_fields=iva_determinacion_fields)
                    page.wait_for_timeout(700)

                    # Fill any remaining IVA fields (e.g. Pago tab) via mapping
                    fill_obligation_section(page, mapping, label_map, iva_pago_labels)
                    page.wait_for_timeout(1000)

                    ok, msg = check_totals(page, data, mapping, tolerance)
                    LOG.info(msg)
                    if not ok:
                        LOG.error("Totals check failed. Not sending declaration.")
                        print(msg, file=sys.stderr)
                        return False

                    if pdf_download_callback is not None:
                        with context.expect_download(timeout=60000) as download_info:
                            if not send_declaration(page, mapping):
                                LOG.warning("Could not find/click Enviar declaración button")
                                return False
                            page.wait_for_timeout(3000)
                            try:
                                download = download_info.value
                                pdf_dir = pdf_download_dir or tempfile.gettempdir()
                                decl_id = declaration_id_api if declaration_id_api is not None else 0
                                pdf_path = os.path.join(pdf_dir, f"receipt_{decl_id}.pdf")
                                download.save_as(pdf_path)
                                LOG.info("PDF receipt saved: %s", pdf_path)
                                pdf_download_callback(pdf_path)
                            except Exception as e:
                                LOG.warning("Could not save PDF or run callback: %s", e)
                    else:
                        if not send_declaration(page, mapping):
                            LOG.warning("Could not find/click Enviar declaración button")
                            return False
                        page.wait_for_timeout(3000)
                    LOG.info("Declaration send clicked; complete any remaining steps in the browser.")
                    return True
                finally:
                    logout_sat(page, mapping)
                    _run_context = None
                    context.close()
                    browser.close()
        except Exception as e:
            LOG.exception("Error during run")
            print(str(e), file=sys.stderr)
            if attempt == 0:
                LOG.info("Closing and retrying once in %s seconds...", RETRY_WAIT_SECONDS)
                time.sleep(RETRY_WAIT_SECONDS)
            else:
                return False
    return False


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fill SAT provisional declaration from Contaayuda Excel workpaper (Impuestos tab)."
    )
    parser.add_argument("--workbook", "-w", help="Full path to the .xlsx workpaper (required unless --test-login)")
    parser.add_argument("--company-id", "-c", type=int, help="Company ID for e.firma lookup (required unless --test-login)")
    parser.add_argument("--branch-id", "-b", type=int, help="Branch ID for e.firma lookup (required unless --test-login)")
    parser.add_argument("--config", help="Path to config.json (default: script dir/config.json)")
    parser.add_argument("--mapping", help="Path to form_field_mapping.json (default: script dir)")
    parser.add_argument("--test-login", action="store_true", help="Test only: open SAT and log in with local .cer/.key and password from config (no DB)")
    parser.add_argument("--test-initial-form", action="store_true", help="Test phase 2: login then fill Declaración Provisional initial form (no Excel; use test_year, test_month, test_periodicidad in config)")
    parser.add_argument("--test-full", action="store_true", help="Test full: login + initial form + phase 3 (ISR Ingresos) + phase 4 (Determinación, ISR retenido VER DETALLE) + phase 5 (Pago); e.firma from config, data from Excel (--workbook); no IVA/send")
    parser.add_argument("--test-phase3", action="store_true", help="Test phase 3: login, initial form, select ISR simplificado de confianza, fill ISR section (Ingresos etc.); requires --workbook; stops before IVA/send")
    parser.add_argument("--test-iva", action="store_true", help="Test IVA only: login, initial form, transition to phase 3, then IVA simplificado de confianza (Determinación); requires --workbook; no ISR/send")
    parser.add_argument("--api", action="store_true", help="Production API mode: get pending declarations, download Excel/CER/KEY, run filler, MarkCompleted with PDF path")
    parser.add_argument("--test-api", action="store_true", help="Test API mode: same data from API as --api but run full flow without sending (no MarkProcessing/MarkCompleted)")
    parser.add_argument("--api-dry-run", action="store_true", help="Only call GetPendingDeclarations and print the list; no download or run")
    parser.add_argument("--api-download-only", action="store_true", help="API: get pending, download Excel/CER/KEY for first declaration, print paths and exit (no browser)")
    parser.add_argument("--test-api-login", action="store_true", help="API: get pending, download files + password, then login to SAT only (e.firma), no form fill")
    args = parser.parse_args()

    if args.test_login:
        success = run(
            workbook_path=None,
            company_id=None,
            branch_id=None,
            config_path=args.config,
            mapping_path=args.mapping,
            test_login=True,
        )
        sys.exit(0 if success else 1)

    if args.test_initial_form:
        success = run(
            workbook_path=None,
            company_id=None,
            branch_id=None,
            config_path=args.config,
            mapping_path=args.mapping,
            test_initial_form=True,
        )
        sys.exit(0 if success else 1)

    if args.test_full:
        workbook = args.workbook
        if not workbook:
            config_path = args.config or os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
            config = load_config(config_path)
            t = config.get("test") or {}
            workbook = config.get("test_workbook_path") or t.get("workbook_path")
        if not workbook:
            print("Error: --test-full requires --workbook or test_workbook_path in config.json.", file=sys.stderr)
            sys.exit(2)
        if not os.path.isfile(workbook):
            print(f"Error: Workbook not found: {workbook}", file=sys.stderr)
            sys.exit(2)
        success = run(
            workbook_path=workbook,
            company_id=None,
            branch_id=None,
            config_path=args.config,
            mapping_path=args.mapping,
            test_full=True,
        )
        # Force exit so process does not hang after browser close (Playwright may leave driver alive)
        os._exit(0 if success else 1)

    if args.test_phase3:
        workbook = args.workbook
        if not workbook:
            config_path = args.config or os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
            config = load_config(config_path)
            t = config.get("test") or {}
            workbook = config.get("test_workbook_path") or t.get("workbook_path")
        if not workbook:
            print("Error: --test-phase3 requires --workbook or test_workbook_path in config.json.", file=sys.stderr)
            sys.exit(2)
        if not os.path.isfile(workbook):
            print(f"Error: Workbook not found: {workbook}", file=sys.stderr)
            sys.exit(2)
        success = run(
            workbook_path=workbook,
            company_id=None,
            branch_id=None,
            config_path=args.config,
            mapping_path=args.mapping,
            test_phase3=True,
        )
        sys.exit(0 if success else 1)

    if args.test_iva:
        workbook = args.workbook
        if not workbook:
            config_path = args.config or os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
            config = load_config(config_path)
            t = config.get("test") or {}
            workbook = config.get("test_workbook_path") or t.get("workbook_path")
        if not workbook:
            print("Error: --test-iva requires --workbook or test_workbook_path in config.json.", file=sys.stderr)
            sys.exit(2)
        if not os.path.isfile(workbook):
            print(f"Error: Workbook not found: {workbook}", file=sys.stderr)
            sys.exit(2)
        success = run(
            workbook_path=workbook,
            company_id=None,
            branch_id=None,
            config_path=args.config,
            mapping_path=args.mapping,
            test_iva=True,
        )
        sys.exit(0 if success else 1)

    # API modes: require --company-id; --branch-id optional (default from config or 1)
    if args.api_dry_run or args.api_download_only or args.test_api_login or args.test_api or args.api:
        if args.company_id is None:
            print("Error: --api, --test-api, --test-api-login, --api-dry-run, and --api-download-only require --company-id.", file=sys.stderr)
            sys.exit(2)
        config_path = args.config or os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
        config = load_config(config_path)
        api_cfg = config.get("sat_automatic_declaration_api") or {}
        base_url = api_cfg.get("base_url") or SAT_AUTOMATIC_DECLARATION_API_BASE_URL
        branch_id = args.branch_id
        if branch_id is None:
            bid = api_cfg.get("branch_id", "1")
            branch_id = int(bid) if isinstance(bid, str) and bid.isdigit() else (int(bid) if isinstance(bid, (int, float)) else 1)
        download_dir = api_cfg.get("download_dir") or ""
        if not download_dir or not os.path.isdir(download_dir):
            download_dir = tempfile.mkdtemp(prefix="sat_declaration_api_")
        setup_logging(config.get("log_file"))
        LOG.info("")
        LOG.info("=== Get pending declarations ===")
        LOG.info("API mode: company_id=%s branch_id=%s base_url=%s download_dir=%s", args.company_id, branch_id, base_url, download_dir)

        if args.api_dry_run:
            pending = get_pending_declarations(args.company_id, branch_id, base_url)
            if not pending:
                LOG.info("No pending declarations for company_id=%s branch_id=%s", args.company_id, branch_id)
                print("No pending declarations.")
                sys.exit(0)
            LOG.info("Found %d pending declaration(s) for company_id=%s branch_id=%s", len(pending), args.company_id, branch_id)
            print(f"Found {len(pending)} pending declaration(s):")
            for i, dec in enumerate(pending):
                cid = dec.get("CUSTOMERID") or dec.get("CustomerId") or ""
                did = dec.get("DECLARATIONID") or dec.get("DeclarationId") or ""
                LOG.info("  Pending %d: CUSTOMERID=%s DECLARATIONID=%s", i + 1, cid, did)
                print(f"  {i + 1}. CUSTOMERID={cid} DECLARATIONID={did}")
            print("(DECLARATIONID is used for MarkProcessing / MarkCompleted when you run --api.)")
            sys.exit(0)

        if args.api_download_only:
            pending = get_pending_declarations(args.company_id, branch_id, base_url)
            if not pending:
                LOG.info("No pending declarations for company_id=%s branch_id=%s", args.company_id, branch_id)
                print("No pending declarations.")
                sys.exit(0)
            LOG.info("Found %d pending declaration(s), downloading first", len(pending))
            for i, dec in enumerate(pending):
                cid = dec.get("CUSTOMERID") or dec.get("CustomerId") or ""
                did = dec.get("DECLARATIONID") or dec.get("DeclarationId") or ""
                LOG.info("  Pending %d: CUSTOMERID=%s DECLARATIONID=%s", i + 1, cid, did)
            LOG.info("")
            LOG.info("=== Download files ===")
            declaration = pending[0]
            try:
                workbook_path, efirma = prepare_declaration_from_api(declaration, download_dir)
            except Exception as e:
                print(f"Error downloading declaration files: {e}", file=sys.stderr)
                sys.exit(2)
            print("Downloaded files (first pending declaration):")
            print(f"  Workbook: {workbook_path}")
            print(f"  CER:      {efirma['cer_path']}")
            print(f"  KEY:      {efirma['key_path']}")
            print(f"  (password: from API response, in memory only — present: {'yes' if efirma.get('password') else 'no'})")
            sys.exit(0)

        pending = get_pending_declarations(args.company_id, branch_id, base_url)
        if not pending:
            LOG.info("No pending declarations for company_id=%s branch_id=%s", args.company_id, branch_id)
            print("No pending declarations.")
            sys.exit(0)
        LOG.info("Found %d pending declaration(s) for company_id=%s branch_id=%s", len(pending), args.company_id, branch_id)
        for i, dec in enumerate(pending):
            cid = dec.get("CUSTOMERID") or dec.get("CustomerId") or ""
            did = dec.get("DECLARATIONID") or dec.get("DeclarationId") or ""
            LOG.info("  Pending %d: CUSTOMERID=%s DECLARATIONID=%s", i + 1, cid, did)
        LOG.info("")
        LOG.info("=== Download files ===")
        declaration = pending[0]
        decl_id = declaration.get("DECLARATIONID") or declaration.get("DeclarationId")
        if decl_id is None:
            print("Error: Declaration missing DECLARATIONID.", file=sys.stderr)
            sys.exit(2)
        declaration_id = int(decl_id) if not isinstance(decl_id, int) else decl_id
        try:
            workbook_path, efirma = prepare_declaration_from_api(declaration, download_dir)
        except Exception as e:
            print(f"Error preparing declaration from API: {e}", file=sys.stderr)
            sys.exit(2)
        LOG.info("")
        LOG.info("=== Login to SAT ===")

        if args.test_api_login:
            success = run(
                workbook_path=None,
                company_id=None,
                branch_id=None,
                config_path=args.config,
                mapping_path=args.mapping,
                test_login=True,
                use_api_efirma=True,
                efirma_from_api=efirma,
            )
            sys.exit(0 if success else 1)

        if args.test_api:
            success = run(
                workbook_path=workbook_path,
                company_id=None,
                branch_id=None,
                config_path=args.config,
                mapping_path=args.mapping,
                test_full=True,
                use_api_efirma=True,
                efirma_from_api=efirma,
            )
            sys.exit(0 if success else 1)

        # --api: production flow with MarkProcessing and MarkCompleted
        mark_processing_flag = api_cfg.get("mark_processing", True)
        if mark_processing_flag:
            mark_processing(args.company_id, branch_id, declaration_id, base_url)

        def _on_pdf(pdf_path: str) -> None:
            mark_completed(args.company_id, branch_id, declaration_id, pdf_path, base_url)

        success = run(
            workbook_path=workbook_path,
            company_id=args.company_id,
            branch_id=branch_id,
            config_path=args.config,
            mapping_path=args.mapping,
            use_api_efirma=True,
            efirma_from_api=efirma,
            company_id_api=args.company_id,
            branch_id_api=branch_id,
            declaration_id_api=declaration_id,
            api_base_url=base_url,
            pdf_download_callback=_on_pdf,
            pdf_download_dir=download_dir,
        )
        sys.exit(0 if success else 1)

    if not args.workbook or args.company_id is None or args.branch_id is None:
        print("Error: --workbook, --company-id, and --branch-id are required unless using a test mode.", file=sys.stderr)
        sys.exit(2)
    if not os.path.isfile(args.workbook):
        print(f"Error: Workbook not found: {args.workbook}", file=sys.stderr)
        sys.exit(2)
    success = run(
        workbook_path=args.workbook,
        company_id=args.company_id,
        branch_id=args.branch_id,
        config_path=args.config,
        mapping_path=args.mapping,
        test_login=False,
    )
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
